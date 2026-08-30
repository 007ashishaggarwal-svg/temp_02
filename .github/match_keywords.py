#!/usr/bin/env python3
"""
High-Precision Biopharma 3-Gate Matching, Scoring & Signal Detection Engine
===========================================================================
Features:
1. 3-Gate Matching Formula:
   Matched = (Primary Indication OR Asset) AND (Secondary Context) NOT (Preclinical / Animal / Noise)
2. Negative Preclinical & Non-Human Rejector Bank:
   Strips mouse model, rat model, murine, in vitro, cell culture, cell line, xenograft,
   zebrafish, drosophila, caenorhabditis, review articles, systematic reviews, meta analyses.
3. Retail Stock & Options Trading Suppressors:
   Strips options trading, Motley Fool, Zacks, GF Value screener headlines, senior notes/debt.
4. Granular Signal Type Classification:
   'clinical_pos', 'clinical_neg', 'regulatory', 'corporate', 'leadership_change', 'commercial', 'general'
5. Dynamic 0-100 Numeric Relevance Scorer:
   Calculates precise integer score based on title match, phase 3/regulatory prominence, and project whitelists.
6. Deterministic Event Clustering & Entity Extraction:
   Outputs Cluster ID and Cluster Hint JSON.
"""

import os
import re
import sys
import json
import html
import hashlib
from datetime import datetime, timezone
import openpyxl

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

def normalize_headline_key(headline: str) -> str:
    """Strip punctuation, publisher suffix, lowercase, and collapse whitespace."""
    if not headline:
        return ""
    # Strip publisher suffix (e.g. ' - BioSpace', ' | Reuters')
    h = re.sub(r"\s+[-–—|]\s+[A-Za-z0-9\s\.\&]+$", "", headline).strip()
    tokens = re.findall(r"\b[a-zA-Z0-9]+\b", h.lower())
    return " ".join(tokens)


class SmartDeduplicator:
    """3-Tier Cross-Channel Deduplication & Canonical URL Upgrade Engine."""
    def __init__(self):
        self.seen_urls = {}
        self.seen_headline_keys = {}

    def is_duplicate_or_merge(self, item: dict) -> tuple[bool, dict]:
        url = item.get("url", "").strip()
        headline = item.get("headline", "").strip()
        h_key = normalize_headline_key(headline)

        if url and url in self.seen_urls:
            existing = self.seen_urls[url]
            if len(item.get("full_text", "")) > len(existing.get("full_text", "")):
                existing["full_text"] = item.get("full_text", "")
            return True, existing

        if h_key and h_key in self.seen_headline_keys:
            existing = self.seen_headline_keys[h_key]
            if "news.google.com" in existing.get("url", "") and "news.google.com" not in url:
                existing["url"] = url
                existing["source"] = item.get("source", existing.get("source", ""))
            if len(item.get("full_text", "")) > len(existing.get("full_text", "")):
                existing["full_text"] = item.get("full_text", "")
            return True, existing

        tokens1 = set(h_key.split())
        for existing_key, existing_item in self.seen_headline_keys.items():
            tokens2 = set(existing_key.split())
            if tokens1 and tokens2:
                overlap = len(tokens1 & tokens2) / max(len(tokens1), len(tokens2))
                if overlap >= 0.85:
                    if "news.google.com" in existing_item.get("url", "") and "news.google.com" not in url:
                        existing_item["url"] = url
                        existing_item["source"] = item.get("source", existing_item.get("source", ""))
                    if len(item.get("full_text", "")) > len(existing_item.get("full_text", "")):
                        existing_item["full_text"] = item.get("full_text", "")
                    return True, existing_item

        stored = dict(item)
        if url:
            self.seen_urls[url] = stored
        if h_key:
            self.seen_headline_keys[h_key] = stored
        return False, stored


# Ambiguous tokens that are common English words or reused trial names across unrelated diseases
AMBIGUOUS_TERMS_REQUIRING_STRICT_CONTEXT = {
    "SURPASS": {
        "required_context": [r"\blilly\b", r"\btirzepatide\b", r"\bmounjaro\b", r"\bzepbound\b", r"\bdiabetes\b", r"\bt2d\b", r"\bobesity\b", r"\bgip\b"],
        "forbidden_context": [r"\bipf\b", r"\bpulmonary\b", r"\bcelea\b", r"\bdeupirfenidone\b", r"\bpuretech\b"]
    },
    "SURMOUNT": {
        "required_context": [r"\blilly\b", r"\btirzepatide\b", r"\bzepbound\b", r"\bobesity\b", r"\bweight\b", r"\bosas\b", r"\bhfpef\b"],
        "forbidden_context": []
    },
    "ACHIEVE": {
        "required_context": [r"\blilly\b", r"\borforglipron\b", r"\bly3502970\b", r"\bglp-?1\b", r"\bdiabetes\b", r"\bt2d\b"],
        "forbidden_context": [r"\bveterinary\b", r"\bcanine\b", r"\bbeverage\b", r"\bcloud\b", r"\brevenue\b", r"\barr\b"]
    },
    "ATTAIN": {
        "required_context": [r"\blilly\b", r"\borforglipron\b", r"\bly3502970\b", r"\bobesity\b", r"\bweight\b", r"\bglp-?1\b"],
        "forbidden_context": [r"\bveterinary\b", r"\bcanine\b", r"\bcarbon\b", r"\bneutrality\b", r"\biso\b"]
    },
    "FOCUS": {
        "required_context": [r"\bsemaglutide\b", r"\bnovo\b", r"\bretinopathy\b", r"\bdiabetes\b"],
        "forbidden_context": [r"\bgrocery\b", r"\bconsumer\b", r"\bpackaging\b", r"\bshopping\b"]
    },
    "SELECT": {
        "required_context": [r"\bsemaglutide\b", r"\bwegovy\b", r"\bnovo\b", r"\bcvot\b", r"\bcardiovascular\b"],
        "forbidden_context": [r"\bchairman\b", r"\bbanking\b", r"\bgovernance\b", r"\bboard\b"]
    },
    "ADC": {
        "required_context": [r"\bantibody\b", r"\bdrug\s+conjugate\b", r"\boncology\b", r"\bcancer\b", r"\btumor\b", r"\bpayload\b", r"\bher2\b", r"\btrop2\b", r"\bclaudin\b"],
        "forbidden_context": [r"\baudio\b", r"\bconverter\b", r"\bdigital\b", r"\banalog\b", r"\bsound\b", r"\bchip\b"]
    },
    "SMA": {
        "required_context": [r"\bspinal\b", r"\bmuscular\b", r"\batrophy\b", r"\bsmn\b", r"\bspinraza\b", r"\bevrdys\b", r"\bzolgensma\b"],
        "forbidden_context": [r"\bsolar\b", r"\binverter\b", r"\bphotovoltaic\b", r"\bshape\s+memory\b", r"\bclean\s+energy\b"]
    }
}

# Global Retail Stock & Governance Noise Patterns
GLOBAL_NOISE_PATTERNS = [
    r"\boptions\s+(?:market|trading|volume|alert|activity|update|flow)\b",
    r"\b(?:call|put)\s+options?\s+(?:activity|volume|sweep|trades?|flow)\b",
    r"\bunusual\s+options\s+activity\b",
    r"\bshort\s+(?:squeeze|interest\s+report|seller|positions?)\b",
    r"\bclass\s+action\s+(?:lawsuit|investigation|filing|alert)\b",
    r"\bshareholder\s+(?:rights|investigation|alert|lawsuit|litigation)\b",
    r"\b(?:rosen\s+law|pomerantz\s+law|glancy\s+prongay|schall\s+law|bronstein\s+gewirtz|levi\s+korsinsky|hagens\s+berman|faruqi\s+&\s+faruqi|gross\s+law|robbins\s+geller)\b",
    r"\block-?up\s+(?:agreement|expiration|period)\b",
    r"\bstock\s+split\b",
    r"\b(?:13f\s+filing|form\s+4\s+filing|insider\s+(?:buying|selling|sales?|transactions?))\b",
    r"\b(?:looks\s+)?(?:\d+\.?\d*%\s+)?(?:overvalued|undervalued)\s+on\s+gf\s+value\b",
    r"\bgf\s+value\s+(?:score|rank|grade|calculation)\b",
    r"\bzacks\s+(?:rank|investment|consensus|equity\s+research|industry\s+rank)\b",
    r"\b(?:motley\s+fool|simply\s+wall\s+st|seeking\s+alpha|gurufocus|investorplace|tipranks|marketbeat|benzinga|wallstreetzen|investor\s+observer|stock\s+titan|tikr(?:\.com)?)\b",
    r"\bstock\s+(?:rises|falls|jumps|plunges|tumbles|drops|holds\s+above|slides|soars|sinks|dips|climbs|retreats)\b",
    r"\bshares\s+(?:plunge|surge|tumble|drop|soar|rally|gain|slide|sink|slip|fall|jump|rise|advance)\b",
    r"\bwhy\s+(?:is\s+)?[\w\s\.-]+\s+stock\s+(?:is\s+)?(?:falling|rising|dropping|plunging|tumbling|crashing|down|up|moving)\b",
    r"\b(?:is\s+)?[\w\s\.-]+\s+stock\s+a\s+(?:buy|sell|hold)\b",
    r"\bis\s+(?:the\s+)?stock\s+(?:still\s+)?worth\s+(?:owning|buying|holding)\b",
    r"\bat\s+record\s+highs,?\s+is\s+the\s+stock\b",
    r"\bprice\s+target\s+(?:raised|lowered|boosted|cut|hiked|slashed|adjusted|set\s+at)\b",
    r"\b(?:top\s+analyst\s+(?:upgrades|downgrades)|analyst\s+upgrades?|analyst\s+downgrades?|initiates\s+coverage\s+on)\b",
    r"\bveterinary\b",
    r"\bcanine\s+(?:epilepsy|health|trial|medicine|cancer|treatment)\b",
    r"\bfeline\b",
    r"\bpet\s+food\s+recall\b",
    r"\bmed\s*spa\b",
    r"\bcompounded\s+(?:semaglutide|tirzepatide)\b",
    r"\bweight\s+loss\s+clinic\b",
    r"\bcounterfeit\s+(?:ozempic|wegovy|mounjaro|zepbound)\b",
    r"\breal\s+estate\b",
    r"\bshopping\s+mall\b",
    r"\bcloud\s+(?:computing|platform|infrastructure)\b",
    r"\bquarterly\s+(?:revenue|arr|cloud)\b",
    r"\brecord\s+(?:quarterly\s+)?revenue\b",
    r"\barr\s+growth\b"
]

# Global Preclinical & Non-Human Animal Model Rejector Bank
GLOBAL_PRECLINICAL_REJECTORS = [
    r"\bmouse\s+models?\b",
    r"\bmice\b",
    r"\bmurine\b",
    r"\brat\s+models?\b",
    r"\brats\b",
    r"\bin\s+vitro\b",
    r"\bcell\s+culture\b",
    r"\bcell\s+lines?\b",
    r"\bxenograft\b",
    r"\bpatient-derived\s+xenografts?\b",
    r"\bpdxf\b",
    r"\bzebrafish\b",
    r"\bdrosophila\b",
    r"\bfruit\s+flies\b",
    r"\bcaenorhabditis\b",
    r"\bc\.\s*elegans\b",
    r"\bpreclinical\s+only\b",
    r"\bsystematic\s+review\b",
    r"\bmeta-?analysis\b",
    r"\bcase\s+reports?\b",
    r"\bletter\s+to\s+(?:the\s+)?editor\b",
    r"\beditorial\b",
    r"\bstudy\s+protocol\s+only\b",
]

GLOBAL_NOISE_COMPILED = re.compile(r'(?:' + '|'.join(GLOBAL_NOISE_PATTERNS) + r')', re.I)
GLOBAL_PRECLINICAL_COMPILED = re.compile(r'(?:' + '|'.join(GLOBAL_PRECLINICAL_REJECTORS) + r')', re.I)


class KeywordMatcher:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.projects = []
        self.rules = []
        self.signal_rules = []
        self.load_rules_and_projects()

    def load_rules_and_projects(self):
        """Load projects and signal rules from 02_Keywords_and_Rules and 03_Config_and_Settings."""
        if not os.path.exists(self.xlsx_path):
            return

        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)

        # 1. Load Projects from 02_Keywords_and_Rules
        tab_k = next((s for s in wb.sheetnames if "Keywords" in s or "Rules" in s), None)
        if tab_k:
            ws_k = wb[tab_k]
            for r in range(2, ws_k.max_row + 1):
                p_name = str(ws_k.cell(row=r, column=1).value or "").strip()
                p_prim = str(ws_k.cell(row=r, column=2).value or "").strip()
                p_sec = str(ws_k.cell(row=r, column=3).value or "").strip()
                p_neg = str(ws_k.cell(row=r, column=4).value or "").strip()
                desk = str(ws_k.cell(row=r, column=5).value or "Executive Briefing Desk").strip()
                tier = str(ws_k.cell(row=r, column=6).value or "🟡 Tier 2 (Daily)").strip()
                active = str(ws_k.cell(row=r, column=7).value or "Active").strip()

                if not p_name or "Paused" in active:
                    continue

                prim_tokens = [t.strip() for t in p_prim.split(",") if t.strip()]
                sec_tokens = [t.strip() for t in p_sec.split(",") if t.strip()]
                neg_tokens = [t.strip() for t in p_neg.split(",") if t.strip()]

                proj_obj = {
                    "name": p_name,
                    "primary": prim_tokens,
                    "secondary": sec_tokens,
                    "negative": neg_tokens,
                    "desk": desk,
                    "tier": tier,
                    "prim_regex": re.compile(r'\b(?:' + '|'.join(re.escape(k.lower()) for k in prim_tokens) + r')\b', re.I) if prim_tokens else None,
                    "sec_regex": re.compile(r'\b(?:' + '|'.join(re.escape(k.lower()) for k in sec_tokens) + r')\b', re.I) if sec_tokens else None,
                    "neg_regex": re.compile(r'\b(?:' + '|'.join(re.escape(k.lower()) for k in neg_tokens) + r')\b', re.I) if neg_tokens else None,
                }
                self.projects.append(proj_obj)
                self.rules.append(proj_obj)

        # 2. Load Signal Detection Patterns
        self.signal_rules = [
            ("regulatory", [r"\bfda\s+approv(?:al|ed|es)\b", r"\bapproved\s+by\s+fda\b", r"\bema\s+approv(?:al|ed|es)\b", r"\bchmp\s+positive\s+opinion\b", r"\bbla\b", r"\bnda\b", r"\bpriority\s+review\b", r"\bbreakthrough\s+therapy\b", r"\bfast\s+track\b", r"\bpdufa\b", r"\bind\s+clearance\b", r"\bcomplete\s+response\b", r"\bcrl\b"], 40),
            ("clinical_neg", [r"\bclinical\s+hold\b", r"\bhalt(?:ed)?\s+testing\b", r"\bsafety\s+concerns?\b", r"\bdiscontinued\b", r"\bfails?\s+(?:to\s+meet|in\s+phase|phase\s+3)\b", r"\bmisses?\s+primary\b", r"\bunderwhelms?\b", r"\bsafety\s+halt\b", r"\bdeaths?\s+in\s+trial\b", r"\bfutility\b", r"\btoxicity\b"], 35),
            ("clinical_pos", [r"\btopline\b", r"\bmet\s+primary\s+endpoint\b", r"\bpositive\s+(?:results|data|phase)\b", r"\bstatistically\s+significant\b", r"\boverall\s+survival\b", r"\bpfs\s+improvement\b", r"\bphase\s+3\s+win\b", r"\bphase\s+2\s+success\b", r"\bachieves?\s+primary\b"], 35),
            ("corporate", [r"\bacquire(?:s|d)?\b", r"\bacquisition\b", r"\bto\s+buy\b", r"\bmerger\b", r"\bin-licensing\b", r"\blicensing\s+(?:deal|pact|agreement)\b", r"\bstrategic\s+partnership\b", r"\bbuyout\b", r"\basset\s+purchase\b", r"\bappoints?\b", r"\bnamed\s+ceo\b", r"\bnew\s+ceo\b", r"\bchief\s+medical\s+officer\b", r"\bcso\b", r"\bleadership\s+shuffle\b", r"\bexecutive\s+appointment\b", r"\bjob\s+postings\b", r"\bcut\s+staff\b", r"\blayoffs?\b", r"\brestructur\b"], 25),
            ("commercial", [r"\bcommercial\s+launch\b", r"\bmarket\s+availability\b", r"\breimbursement\b", r"\bformulary\b", r"\bnice\s+recommendation\b", r"\bpricing\s+(?:agreement|pressure|cut)\b", r"\blabel\s+expansion\b", r"\bprescriptions?\s+rise\b", r"\bmarket\s+share\b", r"\brollout\b"], 20),
            ("general", [r"\bpublishes\b", r"\bpresents\b", r"\bconference\b", r"\babstract\b", r"\bscientific\s+update\b"], 10)
        ]
        self.load_signals_from_workbook()

    def load_signals_from_workbook(self):
        """Dynamically load customized signal types and bonus points from '03_Config_and_Settings'."""
        try:
            script_dir = os.path.dirname(__file__)
            workspace_dir = os.path.abspath(os.path.join(script_dir, "..", ".."))
            xlsx_path = os.path.join(workspace_dir, "RSSFeedChecker_Master_Guide_and_Data.xlsx")
            if not os.path.exists(xlsx_path):
                return
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            if "03_Config_and_Settings" not in wb.sheetnames:
                return
            ws = wb["03_Config_and_Settings"]

            custom_rules = []
            in_section_2 = False
            for r in range(1, ws.max_row + 1):
                val_a = str(ws.cell(row=r, column=1).value or "").strip()
                if "2. SIGNAL TYPE CLASSIFICATION" in val_a:
                    in_section_2 = True
                    continue
                if in_section_2 and ("3. ALGORITHMIC" in val_a or not val_a):
                    if "3. ALGORITHMIC" in val_a:
                        break
                    continue
                if in_section_2 and val_a and "Signal Code" not in val_a:
                    code = val_a
                    patterns_raw = str(ws.cell(row=r, column=3).value or "").split(",")
                    try:
                        bonus = int(ws.cell(row=r, column=4).value or 10)
                    except Exception:
                        bonus = 10
                    compiled = [r"\b" + re.escape(p.strip().lower()) + r"\b" for p in patterns_raw if p.strip()]
                    if compiled:
                        custom_rules.append((code, compiled, bonus))
            if custom_rules:
                self.signal_rules = custom_rules
        except Exception:
            pass

    def detect_signal_type(self, title: str, text: str) -> tuple[str, int]:
        """Detect granular Signal Type and return type code + relevance bonus."""
        comb = f"{title} {text}".lower()

        for rule in self.signal_rules:
            sig_code = rule[0]
            patterns = rule[1]
            bonus = rule[2] if len(rule) > 2 else 10
            for pat in patterns:
                if re.search(pat, comb, re.IGNORECASE):
                    return sig_code, bonus

        return "general", 10

    def generate_clustering_payload(self, title: str, matched_company: str, signal_type: str, project_name: str, pub_date: str) -> tuple[str, str]:
        """Generate deterministic Cluster ID and Cluster Hint JSON."""
        # Normalize company slug
        if matched_company and matched_company != "Not Identified":
            clean_company = re.sub(r"[^\w\s-]", "", matched_company).strip()
            comp_slug = re.sub(r"\s+", "_", clean_company.lower())[:20]
        else:
            comp_slug = "unidentified"

        # Event slug
        event_slug = signal_type if signal_type != "general" else "event"

        # Semantic title fingerprint
        title_words = [w.lower() for w in re.findall(r"\b[a-zA-Z0-9]{3,}\b", title) if w.lower() not in [
            "the", "and", "for", "with", "from", "that", "this", "after", "into", "over", "under", "about"
        ]]
        hash_seed = "".join(title_words[:6])
        hash4 = hashlib.md5(hash_seed.encode("utf-8")).hexdigest()[:4]

        date_str = pub_date.replace("-", "") if pub_date else datetime.now(timezone.utc).strftime("%Y%m%d")
        cluster_id = f"{date_str}_{comp_slug}_{event_slug}_{hash4}"

        hint_dict = {
            "company": matched_company if matched_company else "Not Identified",
            "project": project_name,
            "signal_type": signal_type,
            "event_fingerprint": hash4
        }
        return cluster_id, json.dumps(hint_dict)

    def match(self, title: str, body: str = "", source_class: str = "", source_name: str = "") -> dict:
        """
        Evaluate article against 3-Gate Matching Engine & Scoring Matrix.
        Returns complete intelligence dictionary.
        """
        # Clean incoming title: HTML unescape and strip publisher suffixes
        import html
        title = html.unescape(title).strip()
        title = re.sub(r'\s*[-–—|]\s*(?:BioSpace|Fierce\s*Biotech|Fierce\s*Pharma|Endpoints\s*News|STAT\s*News|Reuters|PR\s*Newswire|Business\s*Wire|GlobeNewswire|The\s*Pharmaletter|Pharmaphorum|Pink\s*Sheet|Scrip|FirstWord\s*Pharma|MedCity\s*News|BioWorld|GEN|CNBC|Evaluate\s*Vantage)\b.*$', '', title, flags=re.I).strip()

        comb_text = f"{title} {body}"
        comb_lower = comb_text.lower()
        title_lower = title.lower()

        # ---------------------------------------------------------------------
        # GATE 1: GLOBAL NOISE & PRECLINICAL REJECTION
        # ---------------------------------------------------------------------
        if GLOBAL_NOISE_COMPILED.search(title_lower) or GLOBAL_NOISE_COMPILED.search(comb_lower):
            return {"matched": False, "is_noise": True, "noise_reason": "Hit financial noise pattern"}

        if GLOBAL_PRECLINICAL_COMPILED.search(title_lower) or GLOBAL_PRECLINICAL_COMPILED.search(comb_lower):
            return {"matched": False, "is_noise": True, "noise_reason": "Hit preclinical noise pattern"}

        # ---------------------------------------------------------------------
        # GATE 2: CONTEXT-GATING FOR AMBIGUOUS SHORT TOKENS
        # ---------------------------------------------------------------------
        for amb_token, rules in AMBIGUOUS_TERMS_REQUIRING_STRICT_CONTEXT.items():
            if re.search(r"\b" + re.escape(amb_token) + r"\b", comb_text):
                has_context = any(re.search(r"\b" + re.escape(ctx) + r"\b", comb_lower) for ctx in rules["required_context"])
                if not has_context:
                    # Strip ambiguous token if context missing
                    comb_lower = re.sub(r"\b" + re.escape(amb_token.lower()) + r"\b", "", comb_lower)
                    title_lower = re.sub(r"\b" + re.escape(amb_token.lower()) + r"\b", "", title_lower)

        # ---------------------------------------------------------------------
        # GATE 3: INDICATION & PROJECT SCORING
        # ---------------------------------------------------------------------
        best_project = None
        best_score = 0
        best_matches = []

        for proj in self.projects:
            # Check project-specific negative rejectors
            if proj["neg_regex"] and proj["neg_regex"].search(comb_lower):
                continue

            # Check Primary & Secondary Inclusion Keywords with compiled regex
            prim_match = proj["prim_regex"].findall(comb_lower) if proj["prim_regex"] else []
            sec_match = proj["sec_regex"].findall(comb_lower) if proj["sec_regex"] else []

            # Specific disease/indication projects REQUIRE matched_prim!
            is_general_theme = "General Biopharma" in proj["name"]
            if prim_match or (is_general_theme and sec_match):
                score = 50  # Baseline
                if proj["prim_regex"] and proj["prim_regex"].search(title_lower):
                    score += 30  # Primary in title
                elif proj["sec_regex"] and proj["sec_regex"].search(title_lower):
                    score += 20  # Secondary in title

                if sec_match:
                    score += 15  # Whitelist context match
                if any(p in comb_lower for p in ["phase 3", "phase iii", "pivotal", "approval", "topline", "fda", "ema"]):
                    score += 15  # Phase 3 / regulatory milestone

                if score > best_score:
                    best_score = score
                    best_project = proj
                    best_matches = prim_match + sec_match

        # If no specific therapeutic project matched, reject as unassigned / out-of-scope
        if best_project is None:
            return {"matched": False, "is_noise": False, "reason": "No targeted therapeutic indication match"}

        # ---------------------------------------------------------------------
        # GATE 4: SIGNAL TYPE, DESK ROUTING & RELEVANCE SCORING
        # ---------------------------------------------------------------------
        signal_type, sig_bonus = self.detect_signal_type(title, body)
        total_relevance = min(100, best_score + sig_bonus if best_score > 0 else 40 + sig_bonus)

        if any(sc in str(source_class) for sc in ["Corporate", "SEC EDGAR", "Commercial PR", "Regulatory"]):
            total_relevance = min(100, total_relevance + 10)

        # Priority tier
        if total_relevance >= 80:
            top_priority = "🔴 Tier 1 (Urgent)"
        elif total_relevance >= 60:
            top_priority = "🟡 Tier 2 (Daily)"
        else:
            top_priority = "🟢 Tier 3 (Weekly)"

        # Headline-Dominant Standardized Desk Taxonomy Routing
        project_name = best_project["name"]
        
        # 1. Headline-Specific Corporate / M&A / Financing / Workforce Check
        if re.search(r'\b(cut\s+staff|job\s+postings|layoffs?|restructur|reorganiz|\$\d+(\.\d+)?\s*(?:m|million|b|billion)\s+offering|to\s+acquire|merger|acquisition|partnership|deal\s+with|cans\s+its\s+partnership)\b', title_lower) and not any(d in title_lower for d in ["phase 3", "fda", "approv", "trial", "study"]):
            routed_desk = "Corporate Strategy & M&A Desk"
        # 2. Multiple Myeloma
        elif re.search(r'\b(multiple\s+myeloma|myeloma|carvykti|darzalex|cilta-cel)\b', title_lower) or re.search(r'\b(multiple\s+myeloma|myeloma)\b', comb_lower):
            routed_desk = "Multiple Myeloma Desk"
        # 3. Oncology / Immuno-Oncology
        elif re.search(r'\b(lung\s+cancer|nsclc|sclc|breast\s+cancer|solid\s+tumor|melanoma|leukemia|lymphoma|enhertu|trodelvy|keytruda|opdivo|tagrisso|adc\b|car-?t)\b', title_lower) or re.search(r'\b(lung\s+cancer|nsclc|sclc|breast\s+cancer|solid\s+tumor|melanoma|leukemia|lymphoma|enhertu|trodelvy|keytruda|opdivo|tagrisso)\b', comb_lower):
            routed_desk = "Oncology & Immuno-Oncology Desk"
        # 4. Neuroscience & CNS
        elif re.search(r'\b(alzheimer|dementia|parkinson|amyotrophic|als\b|huntington|schizophrenia|cns\b|neurodegeneration|lecanemab|donanemab|kisunla|leqembi)\b', title_lower) or re.search(r'\b(alzheimer|dementia|parkinson|als\b|huntington|schizophrenia)\b', comb_lower):
            routed_desk = "Neuroscience & CNS Desk"
        # 5. Metabolic & Obesity
        elif re.search(r'\b(obesity|weight\s+loss|glp-?1|gip\b|semaglutide|wegovy|ozempic|tirzepatide|zepbound|mounjaro|retatrutide|cagrisema|orforglipron|monlunabant|petrelintide|pemvidutide|mash\b|nash\b|diabetes|t2d)\b', title_lower) or re.search(r'\b(obesity|weight\s+loss|glp-?1|semaglutide|wegovy|ozempic|tirzepatide|zepbound|mounjaro|retatrutide|cagrisema|orforglipron)\b', comb_lower):
            routed_desk = "Metabolic & Obesity Desk"
        # 6. Immunology & Inflammation
        elif re.search(r'\b(rheumatoid|arthritis|lupus|crohn|colitis|psoriasis|atopic\s+dermatitis|inflammation|humira|dupixent|skyrizi|rinvoq)\b', title_lower) or re.search(r'\b(rheumatoid|arthritis|lupus|crohn|colitis|psoriasis|atopic\s+dermatitis)\b', comb_lower):
            routed_desk = "Immunology & Inflammation Desk"
        # 7. Rare Disease / Gene Therapy
        elif re.search(r'\b(rare\s+disease|gene\s+therapy|crispr|duchenne|dmd\b|spinal\s+muscular\s+atrophy|sma\b|hunter\s+syndrome|mucopolysaccharidosis|fabry|elevidys)\b', title_lower) or re.search(r'\b(rare\s+disease|gene\s+therapy|crispr|duchenne|dmd\b|spinal\s+muscular\s+atrophy|hunter\s+syndrome)\b', comb_lower):
            routed_desk = "Rare Diseases & Gene Therapy Desk"
        else:
            routed_desk = best_project.get("desk", "Corporate Strategy & M&A Desk")

        matched_str = ", ".join(list(dict.fromkeys(best_matches))[:5]) if best_matches else "Biopharma Stream"
        extracted_company = extract_biopharma_company(title, body, source_name)

        return {
            "matched": True,
            "is_noise": False,
            "project_name": project_name,
            "routed_desk": routed_desk,
            "assigned_desk": routed_desk,
            "top_priority": top_priority,
            "relevance_score": total_relevance,
            "signal_type": signal_type,
            "matched_keywords_str": matched_str,
            "matched_company": extracted_company
        }


# -----------------------------------------------------------------------------
# ROBUST BIOPHARMA COMPANY EXTRACTOR
# -----------------------------------------------------------------------------
_BIOPHARMA_COMPANIES_CACHE = None
_BIOPHARMA_COMPANIES_PATTERN = None

def _get_biopharma_companies() -> list[str]:
    global _BIOPHARMA_COMPANIES_CACHE
    if _BIOPHARMA_COMPANIES_CACHE is not None:
        return _BIOPHARMA_COMPANIES_CACHE

    comps = set()
    script_dir = os.path.dirname(__file__)
    workspace_dir = os.path.abspath(os.path.join(script_dir, "..", ".."))

    # 1. From companies.tsv
    tsv_path = os.path.join(workspace_dir, "companies.tsv")
    if os.path.exists(tsv_path):
        try:
            with open(tsv_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    parts = line.strip().split("\t")
                    if parts and parts[0] and parts[0] != "Company":
                        c = parts[0].strip()
                        clean_c = re.sub(r"\b(Inc\.?|LLC|Corp\.?|Corporation|AG|SE|S\.A\.|PLC|Co\.|Company|Group|Holdings|Pharma|Pharmaceuticals|Therapeutics|Biosciences|Biotech)\b", "", c, flags=re.I).strip()
                        if len(clean_c) >= 3:
                            comps.add(clean_c)
                        if len(c) >= 3:
                            comps.add(c)
        except Exception:
            pass

    # Top aliases & leading drugmakers
    top_aliases = [
        "Eli Lilly", "Lilly", "Roche", "Genentech", "Hanmi", "Protagonist", "Regenxbio", "Capricor",
        "Gilead", "Pfizer", "AstraZeneca", "Novartis", "Merck", "Bristol Myers Squibb", "BMS", "Sanofi",
        "GSK", "AbbVie", "Amgen", "Biogen", "Moderna", "BioNTech", "Novo Nordisk", "Vertex", "Insmed",
        "Amylyx", "Neurocrine", "Zealand", "Cellares", "Akeso", "Evopoint", "Blossomhill", "Ambrosia",
        "Haisco", "OmicsBank", "Alnylam", "Argenx", "Genmab", "Incyte", "Viatris", "Baxter", "Grifols",
        "Bausch", "Teva", "UCB", "Organon", "Sandoz", "Ipsen", "Jazz", "Exelixis", "United Therapeutics",
        "Daiichi Sankyo", "Eisai", "Takeda", "Otsuka", "Chugai", "Kyowa Kirin", "Astellas", "BioMarin",
        "Sarepta", "Ionis", "Ultragenyx", "Beam Therapeutics", "Intellia", "CRISPR Therapeutics",
        "Kymera", "Blueprint", "Deciphera", "Mirati", "Karyopharm", "Legend Biotech", "Summit Therapeutics"
    ]
    for a in top_aliases:
        comps.add(a)

    _BIOPHARMA_COMPANIES_CACHE = sorted(list(comps), key=lambda x: -len(x))
    return _BIOPHARMA_COMPANIES_CACHE


def _get_comps_pattern():
    global _BIOPHARMA_COMPANIES_PATTERN
    if _BIOPHARMA_COMPANIES_PATTERN is None:
        comps = _get_biopharma_companies()
        valid = [re.escape(c) for c in comps if len(c) >= 3 and c.lower() not in COMPANY_BLACKLIST_TERMS]
        _BIOPHARMA_COMPANIES_PATTERN = re.compile(r'\b(' + '|'.join(valid) + r')\b', re.I)
    return _BIOPHARMA_COMPANIES_PATTERN


COMPANY_ALIASES = {
    "Lilly": "Eli Lilly",
    "Genentech": "Roche",
    "Hanmi": "Hanmi Pharmaceutical",
    "Protagonist": "Protagonist Therapeutics",
    "Regenxbio": "Regenxbio",
    "Capricor": "Capricor Therapeutics",
    "Gilead": "Gilead Sciences",
    "Bristol Myers": "Bristol Myers Squibb",
    "BMS": "Bristol Myers Squibb",
    "Vertex": "Vertex Pharmaceuticals",
    "Amylyx": "Amylyx Pharmaceuticals",
    "Neurocrine": "Neurocrine Biosciences",
    "Zealand": "Zealand Pharma",
    "Evopoint": "Evopoint Biosciences",
    "Blossomhill": "Blossomhill Therapeutics",
    "Ambrosia": "Ambrosia Biosciences",
    "Haisco": "Haisco Pharmaceutical"
}

COMPANY_BLACKLIST_TERMS = {
    "obesity", "cancer", "diabetes", "mash", "nash", "parkinson", "lupus", "myeloma",
    "schizophrenia", "rare diseases", "gene therapy", "cell therapy", "glp-1", "glp1",
    "gip", "adc", "fda", "ema", "mhra", "nmpa", "pmda", "who", "cdc", "nih", "sec",
    "phase 1", "phase 2", "phase 3", "phase i", "phase ii", "phase iii", "ind", "bla",
    "nda", "epar", "orphan designation", "safety roundup", "weekly", "monthly", "daily",
    "press release", "news", "report", "update", "study", "trial", "results", "guidance",
    "biopharma", "biotech", "therapeutics", "pharmaceuticals", "medicines", "drug", "drugs"
}


def extract_biopharma_company(title: str, text: str = "", source_name: str = "") -> str:
    """
    Deterministically extracts authentic biopharma company name from headline, text, or corporate source.
    Returns 'Not Identified' if no drugmaker/biotech entity is matched.
    """
    pat = _get_comps_pattern()

    # 1. Search in title (highest priority)
    m = pat.search(title)
    if m:
        c = m.group(1)
        return COMPANY_ALIASES.get(c, c)

    # 2. Check source name if official corporate newsroom
    clean_src = re.sub(r"-Official|-Newsroom|-PR|Official|Newsroom", "", source_name, flags=re.I).strip()
    if clean_src and clean_src.lower() not in COMPANY_BLACKLIST_TERMS:
        m_src = pat.search(clean_src)
        if m_src:
            c = m_src.group(1)
            return COMPANY_ALIASES.get(c, c)

    # 3. Check leading 300 characters of text
    if text:
        m_txt = pat.search(text[:300])
        if m_txt:
            c = m_txt.group(1)
            return COMPANY_ALIASES.get(c, c)

    return "Not Identified"
