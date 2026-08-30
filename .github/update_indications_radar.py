#!/usr/bin/env python3
"""
Update 05_Indications_Radar (18) with exact CT.gov live trial counts,
recent protocol additions, and live Google News search headlines.
"""

import os
import sys
import re
import json
import time
import urllib.request
import urllib.parse
import ssl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX_PATH = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

def fetch_xml_headline(url):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
        with urllib.request.urlopen(req, timeout=10, context=CTX) as resp:
            body = resp.read().decode("utf-8", "ignore")
            m = re.search(r"<item\b.*?<title[^>]*>(.*?)</title>", body, re.I | re.S)
            items = len(re.findall(r"<item\b", body))
            title = ""
            if m:
                title = re.sub(r"<[^>]+>", "", m.group(1)).replace("&amp;", "&").replace("&quot;", '"').strip()
            return resp.getcode(), items, title[:150]
    except Exception as e:
        return 0, 0, str(e)[:60]

def fetch_ctgov(cond):
    if not cond or cond == "None" or cond == "N/A":
        return "N/A (Industry / Regulatory tracking only)", 0, "N/A"
    try:
        q_cond = urllib.parse.quote(cond)
        url = f"https://clinicaltrials.gov/api/v2/studies?query.cond={q_cond}&sort=LastUpdatePostDate:desc&pageSize=3&countTotal=true"
        req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=15, context=CTX) as resp:
            data = json.loads(resp.read().decode("utf-8", "ignore"))
            total = data.get("totalCount", 0)
            studies = data.get("studies", [])
            latest_info = "N/A"
            if studies:
                s0 = studies[0].get("protocolSection", {})
                nct = s0.get("identificationModule", {}).get("nctId", "")
                title = s0.get("identificationModule", {}).get("briefTitle", "")
                upd_date = s0.get("statusModule", {}).get("lastUpdatePostDateStruct", {}).get("date", "")
                latest_info = f"{nct} ({upd_date}): {title[:120]}"
            return f"200 OK ({total:,} registered trials)", total, latest_info
    except Exception as e:
        return f"Error ({str(e)[:40]})", 0, "N/A"

def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws_name = [n for n in wb.sheetnames if "Indications" in n][0]
    ws = wb[ws_name]
    print(f"Updating '{ws_name}' ({ws.max_row - 1} indications)...")

    # Styles
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    CRIMSON_DARK = "8B0000"
    WHITE = "FFFFFF"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_crimson = PatternFill(start_color=CRIMSON_DARK, end_color=CRIMSON_DARK, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_title = Font(name="Calibri", size=9, italic=True, color="222222")
    font_flag = Font(name="Calibri", size=9, color="111111")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")

    fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # soft green
    fill_note = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")     # soft blue
    fill_failed = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")   # soft red
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Indication / Theme", 22, fill_navy),
        ("Synonyms & Keywords", 38, fill_navy),
        ("CT.gov Condition Term", 22, fill_navy),
        ("Exclude Terms (Negations)", 22, fill_navy),
        ("Broad Feed Query Logic", 45, fill_navy),
        ("Clinical/Reg Feed Query Logic", 45, fill_navy),
        ("CT.gov Protocol Endpoint", 55, fill_navy),
        ("Broad GNews Status", 16, fill_blue),
        ("Broad GNews Headline", 45, fill_blue),
        ("Clinical/Reg GNews Status", 18, fill_blue),
        ("Clinical/Reg GNews Headline", 45, fill_blue),
        ("CT.gov Live Registry Status", 25, fill_blue),
        ("Latest CT.gov Protocol Registered", 55, fill_blue),
        ("Radar Scope & CI Observations", 55, fill_crimson),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    for r in range(2, ws.max_row + 1):
        ind = str(ws.cell(row=r, column=1).value or "")
        syn = str(ws.cell(row=r, column=2).value or "")
        ct_cond = str(ws.cell(row=r, column=3).value or "")
        b_query = str(ws.cell(row=r, column=5).value or "")
        c_query = str(ws.cell(row=r, column=6).value or "")
        
        # 1. Broad GNews
        b_q_clean = b_query.replace(" AND ", " ").replace(" OR ", " OR ").strip()
        b_url = f"https://news.google.com/rss/search?q={urllib.parse.quote(b_q_clean)}&hl=en-US&gl=US&ceid=US:en"
        b_code, b_count, b_title = fetch_xml_headline(b_url)

        # 2. Clinical GNews
        c_q_clean = c_query.replace(" AND ", " ").replace(" OR ", " OR ").strip()
        c_url = f"https://news.google.com/rss/search?q={urllib.parse.quote(c_q_clean)}&hl=en-US&gl=US&ceid=US:en"
        c_code, c_count, c_title = fetch_xml_headline(c_url)

        # 3. CT.gov
        ct_status, ct_total, ct_latest = fetch_ctgov(ct_cond if ct_cond not in ("None", "") else None)
        if ct_cond and ct_cond not in ("None", ""):
            ws.cell(row=r, column=7, value=f"https://clinicaltrials.gov/api/v2/studies?query.cond={urllib.parse.quote(ct_cond)}&sort=LastUpdatePostDate:desc")
        else:
            ws.cell(row=r, column=7, value="N/A")

        # 4. Observations
        obs = []
        if b_count > 0:
            obs.append(f"Broad Radar active ({b_count} articles)")
        if c_count > 0:
            obs.append(f"Clinical Radar active ({c_count} trial readouts)")
        if ct_total > 0:
            obs.append(f"CT.gov active: {ct_total:,} studies registered")
        elif ct_cond in ("None", "", "N/A"):
            obs.append("Industry/Corporate theme tracking (press & M&A intelligence)")

        obs_str = " | ".join(obs)

        # Apply formatting
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.border = thin_border
            if c == 7 and str(cell.value or "").startswith("http"):
                cell.font = font_link
                cell.hyperlink = str(cell.value)

        # Col 8: Broad status
        cell_8 = ws.cell(row=r, column=8, value=f"{b_code} OK ({b_count} items)" if b_code == 200 else f"HTTP {b_code}")
        cell_8.font = font_data
        cell_8.fill = fill_ok if b_code == 200 else fill_failed
        cell_8.border = thin_border
        cell_8.alignment = Alignment(horizontal="center")

        # Col 9: Broad headline
        cell_9 = ws.cell(row=r, column=9, value=b_title)
        cell_9.font = font_title
        cell_9.fill = fill_ok if b_code == 200 else fill_failed
        cell_9.border = thin_border
        cell_9.alignment = Alignment(wrap_text=True)
        if b_url:
            cell_9.hyperlink = b_url

        # Col 10: Clin status
        cell_10 = ws.cell(row=r, column=10, value=f"{c_code} OK ({c_count} items)" if c_code == 200 else f"HTTP {c_code}")
        cell_10.font = font_data
        cell_10.fill = fill_ok if c_code == 200 else fill_failed
        cell_10.border = thin_border
        cell_10.alignment = Alignment(horizontal="center")

        # Col 11: Clin headline
        cell_11 = ws.cell(row=r, column=11, value=c_title)
        cell_11.font = font_title
        cell_11.fill = fill_ok if c_code == 200 else fill_failed
        cell_11.border = thin_border
        cell_11.alignment = Alignment(wrap_text=True)
        if c_url:
            cell_11.hyperlink = c_url

        # Col 12: CT.gov status
        cell_12 = ws.cell(row=r, column=12, value=ct_status)
        cell_12.font = font_data
        cell_12.fill = fill_ok if "200 OK" in ct_status else fill_note if "N/A" in ct_status else fill_failed
        cell_12.border = thin_border
        cell_12.alignment = Alignment(horizontal="center")

        # Col 13: CT.gov latest
        cell_13 = ws.cell(row=r, column=13, value=ct_latest)
        cell_13.font = font_code
        cell_13.fill = fill_ok if "200 OK" in ct_status else fill_note if "N/A" in ct_status else fill_failed
        cell_13.border = thin_border
        cell_13.alignment = Alignment(wrap_text=True)

        # Col 14: Observations
        cell_14 = ws.cell(row=r, column=14, value=obs_str)
        cell_14.font = font_flag
        cell_14.fill = fill_white
        cell_14.border = thin_border
        cell_14.alignment = Alignment(wrap_text=True)

        ws.row_dimensions[r].height = 24
        print(f"  [{r-1:>2}/18] Verified {ind:<22} | CT.gov: {ct_status[:30]} | Broad GNews: {b_count} items")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{ws.max_row}"

    wb.save(XLSX_PATH)
    print(f"Indications Radar successfully verified and saved to {XLSX_PATH}!")

if __name__ == "__main__":
    main()
