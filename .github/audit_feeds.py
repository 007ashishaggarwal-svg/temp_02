#!/usr/bin/env python3
"""
Deep audit and comprehensive testing of all feeds in 03_Feeds_Master (518):
1. Verifies and tests all feeds (including live test for Rows 503-505).
2. Performs intelligent competitive intelligence categorization.
3. Audits all URL links for duplicates, sitemap navigation noise, "Hello world!" empty posts,
   non-pharma companies, search false positives, and foreign language requirements.
4. Corrects false noise flags on valid feeds (Feed_005, Feed_060, Feed_502-504).
5. Updates RSSFeedChecker_Master_Guide_and_Data.xlsx with complete color-coded styling.
"""

import os
import sys
import re
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from check_feeds import fetch, classify, try_fallback_cascade

XLSX = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"
SHEET = "03_Feeds_Master (518)"

# ======================== KNOWN DUPLICATES ========================
DUPLICATES_MAP = {
    "Feed_096": "Duplicate of Feed_095 (same Ionis IR content via ir.ionispharma.com domain)",
    "Feed_350": "Duplicate of Feed_095/096 (third Ionis feed via ionis.com/feed)",
    "Feed_155": "Duplicate of Feed_154 (same RegenXBio IR content via gcs-web mirror)",
    "Feed_157": "Duplicate of Feed_156 (same Tiziana IR content)",
    "Feed_169": "Duplicate of Feed_158 (same Vertex IR content via news.vrtx.com)",
    "Feed_277": "Duplicate of Feed_099 (same AbbVie IR content via /rss/news-releases.xml)",
    "Feed_078": "Duplicate of Feed_257 (Denali company website vs IR - website is empty WP 'Hello world!')",
    "Feed_148": "Duplicate of Feed_146 (same IGC Pharma content)",
    "Feed_332": "Duplicate of Feed_331 (same Moderna content via newsroom vs investor relations)",
    "Feed_355": "Duplicate of Feed_280 (Voyager sitemap RSS vs IR feed - sitemap yields 'Careers')",
    "Feed_357": "Duplicate of Feed_145 (Allogene events RSS vs main feed)",
    "Feed_076": "Duplicate of Feed_075 (same AriBio content via alternative domain)",
    "Feed_246": "Duplicate of Feed_205 (Genethon French vs English site)",
    "Feed_447": "Duplicate of Feed_033/499 (Endpoints News feed via /news/ path)",
    "Feed_497": "Duplicate of Feed_501 (FierceBiotech feed URL mislabeled as Xconomy/Bio)",
    "Feed_500": "Duplicate of Feed_496 (FiercePharma feed URL duplicate)",
    "Feed_503": "Duplicate of Feed_438 (FDA Drugs CDER feed duplicate URL)",
}


def check_noise(feed_id, url, label, cat, title, result):
    """Return noise flag and rationale, or empty string if clean."""
    url_lower = url.lower()
    label_lower = label.lower()
    title_lower = (title or "").lower().strip()
    res_lower = (result or "").lower()
    flags = []

    # 1. Duplicates
    if feed_id in DUPLICATES_MAP:
        flags.append(DUPLICATES_MAP[feed_id])

    # 2. WordPress "Hello world!" default unpopulated feeds
    if title_lower == "hello world!" or (title_lower == "hello world" and "ok" in res_lower):
        flags.append("NOISE: Feed returns default WordPress 'Hello world!' template - CMS unpopulated with news")

    # 3. Sitemap RSS feeds that return navigation/footer items
    if "sitemap.rss" in url_lower or "sitemap" in url_lower:
        nav_words = ["footer", "careers", "about", "about us", "home", "meeting", "contact", "board", "leadership"]
        if any(w in title_lower for w in nav_words):
            flags.append("NOISE: Sitemap RSS returning navigation/page structure items, not news content")

    # 4. Feeds returning generic navigation page titles
    if title_lower in ("home", "about", "about us", "about us (3)", "footer", "contact us"):
        flags.append("NOISE: Feed returns website navigation pages, not news/press releases")

    # 5. Non-pharma/non-CI content
    non_pharma_indicators = [
        ("greenlandmines", "NOT PHARMA: Greenland Mines is a rare earth mining company, not pharma/biotech"),
        ("orix.co.jp", "NOT PHARMA: ORIX Corporation is a Japanese financial services/leasing company"),
        ("sisfirst.com", "NOT PHARMA: Surgical Information Systems is a healthcare IT company for ambulatory surgery centers"),
        ("4btechnology", "NOT PHARMA: 4B Technology is a medtech/engineering company, not pharma drug development"),
        ("longwoodfund", "BORDERLINE: Longwood Fund is a VC firm - useful for deal intelligence but not direct pharma news"),
        ("pmgcholdings", "NOISE: PMGC Holdings appears to be an investment holding company, not pharma"),
    ]
    for pattern, reason in non_pharma_indicators:
        if pattern in url_lower:
            flags.append(reason)

    # 6. Very broad / off-topic feeds that add noise
    if "forbes.com/media/feed" in url_lower:
        flags.append("BROAD: Forbes general media feed covers ALL industries - very high noise for pharma CI")
    if "businesskorea" in url_lower:
        flags.append("BROAD: BusinessKorea covers ALL Korean business sectors - high noise ratio for pharma CI")
    if "connect.medrxiv.org/medrxiv_xml.php?subject=all" in url_lower:
        flags.append("BROAD: medRxiv ALL subjects - will include non-pharma medical research. Consider filtering to pharmacology only")

    # 7. SEC filing feeds (not press releases)
    if "data.sec.gov" in url_lower or "sec.gov" in url_lower:
        flags.append("BORDERLINE: SEC EDGAR filing feed - returns regulatory filings (10-K, 10-Q), not press releases. Useful for M&A/financial intelligence but adds volume")

    # 8. Google News company feeds picking up stock analysis noise
    if "news.google.com" in url_lower:
        if "gn-companies-22" in label_lower and "galapagos" in url_lower:
            flags.append("NOISE RISK: 'Galapagos' in GN query matches Galapagos Islands news, not just Galapagos NV pharma")
        stock_noise_words = ["undervalued", "stock options", "share buy-back", "options market", "fairly valued", "getting attention", "lock-up agreement"]
        if any(w in title_lower for w in stock_noise_words):
            flags.append("NOTE: Google News company feed picks up stock analysis articles alongside real pharma news - expected behavior, some noise is inherent")

    # 9. Non-English feeds requiring CI translation notes
    if title:
        non_english = []
        if re.search(r'[\uac00-\ud7af]', title):
            non_english.append("Korean")
        if re.search(r'[\u3040-\u30ff\u4e00-\u9fff]', title) and not re.search(r'[\uac00-\ud7af]', title):
            non_english.append("Japanese/Chinese")
        if re.search(r'[\u0400-\u04ff]', title):
            non_english.append("Russian")
        if re.search(r'[\u0104\u0105\u0106\u0107\u0118\u0119\u0141\u0142\u0143\u0144\u00d3\u00f3\u015a\u015b\u0179\u017a\u017b\u017c]', title):
            non_english.append("Polish")
        if non_english:
            lang = "/".join(non_english)
            flags.append(f"NOTE: Feed content in {lang} - may need translation for English-speaking CI team")

    return " | ".join(flags) if flags else ""


def fix_category(feed_id, url, label, current_cat, title):
    """Accurately determine CI category with strict domain matching."""
    import urllib.parse
    u = url.lower()
    netloc = urllib.parse.urlparse(u).netloc

    if "news.google.com" in netloc:
        return "Google News Aggregation"

    if netloc in ["connect.medrxiv.org", "medrxiv.org", "biorxiv.org", "connect.biorxiv.org", "arxiv.org"]:
        return "Preprint Server"

    journal_domains = ["nature.com", "nejm.org", "thelancet.com", "jamanetwork.com", "cell.com", "annalsofoncology.org", "ascopubs.org"]
    if any(netloc == d or netloc.endswith("." + d) for d in journal_domains):
        if "brainstorm-cell" not in netloc:
            return "Medical Journal"

    reg_domains = ["fda.gov", "ema.europa.eu", "gov.uk", "cdc.gov", "who.int", "g-ba.de", "sec.gov"]
    if any(netloc == d or netloc.endswith("." + d) for d in reg_domains) or "data.sec.gov" in netloc:
        return "Regulatory & Government"

    if any(netloc == d or netloc.endswith("." + d) for d in ["prnewswire.com", "globenewswire.com", "businesswire.com"]):
        return "Newswire"

    patient_advocacy = [
        "hemophilianewstoday.com", "alsnewstoday.com", "smanewstoday.com", "huntingtonsdiseasenews.com",
        "fsma.pl", "f-sma.ru", "famigliesma.org", "curesma.org", "smauk.org.uk", "smabenimleyuru.org.tr",
        "cureduchenne.org", "parentprojectmd.org", "esperare.org", "hdsa.org", "hdbuzz.net", "ehdn.org",
        "wfh.org", "eahad.org"
    ]
    if any(netloc == d or netloc.endswith("." + d) for d in patient_advocacy):
        return "Patient Advocacy & Disease"

    cdmo = ["criver.com", "theemmesgroup.com", "catalent.com", "lonza.com", "siegfried.ch", "bachem.com", "evonik.com", "recipharm.com"]
    if any(netloc == d or netloc.endswith("." + d) for d in cdmo):
        return "CDMO / CRO / Services"

    if netloc in ["health.ucsd.edu", "ucsd.edu"]:
        return "Academic / Research"

    trade_press = [
        "statnews.com", "endpoints.news", "endpts.com", "bioworld.com", "fiercepharma.com", "fiercebiotech.com",
        "fiercehealthcare.com", "biospace.com", "genengnews.com", "labiotech.eu", "pharmexec.com", "pharmaphorum.com",
        "pharmafile.com", "pharmatimes.com", "pharmavoice.com", "healthcaredive.com", "biopharmadive.com", "medtechdive.com",
        "medcitynews.com", "drugdiscoverytrends.com", "pharmaceuticalprocessingworld.com", "worldpharmanews.com",
        "insideprecisionmedicine.com", "neurologylive.com", "hemostasistoday.com", "citeline.com", "businesskorea.co.kr",
        "thebionews.net"
    ]
    if any(netloc == d or netloc.endswith("." + d) for d in trade_press):
        return "Trade Press"

    aggregators = ["medpagetoday.com", "drugs.com", "sciencedaily.com", "forbes.com", "icer.org"]
    if any(netloc == d or netloc.endswith("." + d) for d in aggregators):
        return "Industry News & Analysis"

    return "Company IR / Press Release"


def main():
    print(f"Loading workbook: {XLSX}")
    wb = openpyxl.load_workbook(XLSX)
    target_sheet = next((s for s in wb.sheetnames if "Feeds" in s), None)
    if not target_sheet:
        print(f"ERROR: No Feeds sheet found. Available: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[target_sheet]
    print(f"Loaded sheet '{target_sheet}' with {ws.max_row} rows.")

    # Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    CRIMSON_DARK = "8B0000"
    WHITE = "FFFFFF"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_crimson = PatternFill(start_color=CRIMSON_DARK, end_color=CRIMSON_DARK, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_title = Font(name="Calibri", size=9, italic=True, color="222222")
    font_flag = Font(name="Calibri", size=9, color="111111")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")

    fill_ice = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # light green
    fill_recovered = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")# light yellow
    fill_blocked = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # light pink
    fill_failed = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")   # light red
    fill_other = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")    # light blue-grey

    fill_noise = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")    # yellow noise
    fill_notpharma = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # red non-pharma
    fill_dup = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")      # grey duplicate
    fill_note = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")     # blue note

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Feed ID", 14, fill_navy),
        ("Feed URL", 75, fill_navy),
        ("Label / Source Name", 36, fill_navy),
        ("Category / Coverage Type", 28, fill_navy),
        ("Protocol", 12, fill_navy),
        ("HTTP Code", 12, fill_blue),
        ("Response Time (s)", 16, fill_blue),
        ("Result", 42, fill_blue),
        ("Last Item Title", 55, fill_blue),
        ("Fallback URL", 60, fill_blue),
        ("Test Timestamp", 20, fill_blue),
        ("Noise / Quality Flag", 68, fill_crimson),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    # Live test any untested rows (specifically Rows 503-505)
    untested = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=6).value
        if not code:
            fid = ws.cell(row=r, column=1).value
            url = ws.cell(row=r, column=2).value
            lbl = ws.cell(row=r, column=3).value
            untested.append((r, fid, url, lbl))

    if untested:
        print(f"Testing {len(untested)} previously untested feeds...")
        for r, fid, url, lbl in untested:
            code, elapsed, body, reason, lax_ok = fetch(url)
            res, title = classify(code, body, reason, lax_ok)
            fb_url = ""
            if not (code == 200 and title) or "BLOCKED" in res or "FAILED" in res:
                fb_res, fb_title, fb_cand = try_fallback_cascade(url, lbl)
                if fb_res and fb_title:
                    res, title, fb_url = fb_res, fb_title, fb_cand
            ws.cell(row=r, column=6, value=code)
            ws.cell(row=r, column=7, value=round(elapsed, 3))
            ws.cell(row=r, column=8, value=res)
            ws.cell(row=r, column=9, value=title[:200] if title else "")
            ws.cell(row=r, column=10, value=fb_url)
            ws.cell(row=r, column=11, value=time.strftime("%Y-%m-%d %H:%M:%S"))
            print(f"  Tested {fid} ({lbl}): {code} | {res} | {title[:50]}")

    cat_changes = 0
    flag_count = 0
    flag_stats = {}

    for r in range(2, ws.max_row + 1):
        fid = str(ws.cell(row=r, column=1).value or "")
        url = str(ws.cell(row=r, column=2).value or "")
        lbl = str(ws.cell(row=r, column=3).value or "")
        old_cat = str(ws.cell(row=r, column=4).value or "")
        proto = "HTTPS" if url.startswith("https") else "HTTP"
        res = str(ws.cell(row=r, column=8).value or "")
        title = str(ws.cell(row=r, column=9).value or "")
        fb_url = str(ws.cell(row=r, column=10).value or "")

        new_cat = fix_category(fid, url, lbl, old_cat, title)
        if new_cat != old_cat:
            cat_changes += 1
            print(f"  [CATEGORY UPDATE] Row {r} ({fid}): '{old_cat}' -> '{new_cat}'")
        ws.cell(row=r, column=4, value=new_cat)
        ws.cell(row=r, column=5, value=proto)

        flag_text = check_noise(fid, url, lbl, new_cat, title, res)
        ws.cell(row=r, column=12, value=flag_text if flag_text else None)

        if flag_text:
            flag_count += 1
            for part in flag_text.split(" | "):
                k = part.split(":")[0]
                flag_stats[k] = flag_stats.get(k, 0) + 1

        # Apply styles
        cell_a = ws.cell(row=r, column=1)
        cell_a.font = font_code
        cell_a.border = thin_border
        cell_a.fill = fill_ice if r % 2 == 0 else fill_white

        cell_b = ws.cell(row=r, column=2)
        cell_b.font = font_link
        cell_b.border = thin_border
        cell_b.hyperlink = url

        cell_c = ws.cell(row=r, column=3)
        cell_c.font = font_data
        cell_c.border = thin_border

        cell_d = ws.cell(row=r, column=4)
        cell_d.font = font_data
        cell_d.border = thin_border

        cell_e = ws.cell(row=r, column=5)
        cell_e.font = font_code
        cell_e.border = thin_border
        cell_e.alignment = Alignment(horizontal="center")

        if "OK" in res and "fetchable" in res:
            res_fill = fill_ok
        elif "RECOVERED" in res:
            res_fill = fill_recovered
        elif "BLOCKED" in res:
            res_fill = fill_blocked
        elif "FAILED" in res:
            res_fill = fill_failed
        else:
            res_fill = fill_other

        cell_f = ws.cell(row=r, column=6)
        cell_f.font = font_data
        cell_f.fill = res_fill
        cell_f.border = thin_border
        cell_f.alignment = Alignment(horizontal="center")

        cell_g = ws.cell(row=r, column=7)
        cell_g.font = font_data
        cell_g.fill = res_fill
        cell_g.border = thin_border
        cell_g.alignment = Alignment(horizontal="right")

        cell_h = ws.cell(row=r, column=8)
        cell_h.font = font_data
        cell_h.fill = res_fill
        cell_h.border = thin_border
        cell_h.alignment = Alignment(wrap_text=True)

        cell_i = ws.cell(row=r, column=9)
        cell_i.font = font_title
        cell_i.fill = res_fill
        cell_i.border = thin_border
        cell_i.alignment = Alignment(wrap_text=True)

        cell_j = ws.cell(row=r, column=10)
        cell_j.font = font_code
        cell_j.fill = res_fill
        cell_j.border = thin_border
        if fb_url:
            cell_j.hyperlink = fb_url

        cell_k = ws.cell(row=r, column=11)
        cell_k.font = font_code
        cell_k.fill = res_fill
        cell_k.border = thin_border
        cell_k.alignment = Alignment(horizontal="center")

        cell_l = ws.cell(row=r, column=12)
        cell_l.font = font_flag
        cell_l.border = thin_border
        cell_l.alignment = Alignment(wrap_text=True)

        if flag_text:
            if "NOT PHARMA" in flag_text:
                cell_l.fill = fill_notpharma
            elif "Duplicate" in flag_text:
                cell_l.fill = fill_dup
            elif "NOISE" in flag_text:
                cell_l.fill = fill_noise
            else:
                cell_l.fill = fill_note
        else:
            cell_l.fill = fill_white

        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"

    print(f"\nSaving Master Workbook to {XLSX}...")
    wb.save(XLSX)
    print("Save successful!")

    print(f"\n{'='*78}")
    print("AUDIT & VERIFICATION COMPLETE")
    print(f"{'='*78}")
    print(f"  Total Feeds in Catalog: {ws.max_row - 1}")
    print(f"  Category Adjustments:   {cat_changes}")
    print(f"  Flagged Noise/Quality:  {flag_count}")
    print("\nFlag Breakdown:")
    for k, v in sorted(flag_stats.items(), key=lambda x: -x[1]):
        print(f"  {k:<30}: {v}")

    print("\nCategory Distribution:")
    cats = {}
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=4).value
        cats[c] = cats.get(c, 0) + 1
    for c, n in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {c:<30}: {n}")


if __name__ == "__main__":
    main()

