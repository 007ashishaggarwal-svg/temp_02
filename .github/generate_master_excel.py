#!/usr/bin/env python3
"""
Generate a comprehensive, beautifully formatted Master Excel Workbook for RSSFeedChecker.
Contains all data across all 3 methods, expanded data dictionaries with fallback strategies,
time-window filtering rules, ClinicalTrials.gov protocol change intelligence, and architecture dossiers.
"""

import csv
import json
import os
import sys

# Ensure UTF-8 stdout
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
OUTPUT_FILE = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

def create_master_workbook():
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    # Color Palette
    NAVY_DARK = "1B365D"
    NAVY_MED = "2D5584"
    ICE_BLUE = "EBF2FA"
    LIGHT_GRAY = "F4F6F8"
    WHITE = "FFFFFF"
    BORDER_COLOR = "D0D7DE"
    
    font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="555555")
    font_sec_header = Font(name="Calibri", size=12, bold=True, color=WHITE)
    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_normal = Font(name="Calibri", size=10, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    
    fill_header = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_sec_header = PatternFill(start_color=NAVY_MED, end_color=NAVY_MED, fill_type="solid")
    fill_ice = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    
    # Status badges
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_green = Font(name="Calibri", size=10, bold=True, color="155724")
    
    fill_yellow = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_yellow = Font(name="Calibri", size=10, bold=True, color="856404")
    
    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_red = Font(name="Calibri", size=10, bold=True, color="721C24")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    header_border = Border(
        left=Side(style='thin', color="FFFFFF"),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='medium', color=NAVY_DARK),
        bottom=Side(style='medium', color=NAVY_DARK)
    )

    # -------------------------------------------------------------------------
    # TAB 1: 01_System_Overview
    # -------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="01_System_Overview")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 36
    ws1.column_dimensions["D"].width = 30
    ws1.column_dimensions["E"].width = 65

    # Title Block
    ws1.merge_cells("B2:E2")
    ws1["B2"] = "RSSFeedChecker — Pharma & Biotech Intelligence System"
    ws1["B2"].font = font_title
    
    ws1.merge_cells("B3:E3")
    ws1["B3"] = "Complete Reference Guide, Operational Architecture & Multi-Pillar Data Directory"
    ws1["B3"].font = font_subtitle

    row = 5
    # Section: Executive Summary
    ws1.merge_cells(f"B{row}:E{row}")
    ws1[f"B{row}"] = "1. EXECUTIVE SUMMARY & PROBLEM STATEMENT"
    ws1[f"B{row}"].font = font_sec_header
    ws1[f"B{row}"].fill = fill_sec_header
    ws1[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[row].height = 24
    row += 1

    summary_texts = [
        ("Core Mission", "RSSFeedChecker is an automated pharma and biotech intelligence engine engineered to solve the coverage, latency, and reliability limitations of relying solely on standard RSS feeds."),
        ("The 59% Blindspot", "Audits across 1,600+ biotech & pharma domains show that only 41% of corporate domains expose a working RSS feed. Relying only on publisher feeds leaves 59% of company corporate releases completely undetected."),
        ("5-Core Schema Guarantee", "Across ALL methods, the system extracts: 1) Publication Date, 2) Title, 3) Snippet/Excerpt, 4) Full Article Text, and 5) Authentic Canonical URL."),
        ("Strict Time Windowing", "Default 3-day (72h) time filter. User can configure 24h, 36h, 3d, 7d, or N days. If a company has no release in that window, it clearly reports 'No releases in last N days · Last release: YYYY-MM-DD'.")
    ]
    for label, desc in summary_texts:
        ws1[f"B{row}"] = label
        ws1[f"B{row}"].font = font_bold
        ws1[f"B{row}"].fill = fill_ice
        ws1[f"B{row}"].border = thin_border
        
        ws1.merge_cells(f"C{row}:E{row}")
        ws1[f"C{row}"] = desc
        ws1[f"C{row}"].font = font_normal
        ws1[f"C{row}"].border = thin_border
        ws1[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws1.row_dimensions[row].height = 36
        row += 1

    row += 1
    # Section: The 3 Pillars Architecture Matrix
    ws1.merge_cells(f"B{row}:E{row}")
    ws1[f"B{row}"] = "2. THE 3-PILLAR ARCHITECTURE MATRIX"
    ws1[f"B{row}"].font = font_sec_header
    ws1[f"B{row}"].fill = fill_sec_header
    ws1[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[row].height = 24
    row += 1

    pillars = [
        ("Pillar 1: Publisher Feed Checker", ".github/scripts/check_feeds.py", "feeds.tsv (518 feeds)", "Maintains breadth across trade press (STAT, Endpoints, Fierce, BioSpace), regulatory agencies (FDA, EMA), and journals (Nature). Verifies SSL, chunked responses, XML parseability, and first item titles."),
        ("Pillar 2: Company Newsroom Watch", ".github/scripts/watch_newsrooms.py", "companies.tsv (616 domains)", "Achieves 100% corporate completeness. Waterfall: RSS autodiscovery -> XML sitemaps (<lastmod>) -> HTML scraping -> Google News 'site:' fallback. Strict time-window filtering (default 3 days). Discovered RSS feeds are promoted to Master Feeds list."),
        ("Pillar 3: Indication & Clinical Trial Radar", ".github/scripts/make_indication_feeds.py & check_clinical_trials.py", "indications.tsv (18 indications)", "Constructs Google News search feeds for new disease areas, and tracks ClinicalTrials.gov protocol changes (New Registrations, Recruitment Status Deltas, Enrollment size changes 50->90).")
    ]

    p_headers = ["Pillar / Method", "Script Path", "Primary Dataset", "Operational Role & Capabilities"]
    for col_idx, h in enumerate(p_headers, start=2):
        col_letter = get_column_letter(col_idx)
        ws1[f"{col_letter}{row}"] = h
        ws1[f"{col_letter}{row}"].font = font_tbl_header
        ws1[f"{col_letter}{row}"].fill = fill_header
        ws1[f"{col_letter}{row}"].border = header_border
        ws1[f"{col_letter}{row}"].alignment = Alignment(vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 22
    row += 1

    for p_name, p_script, p_data, p_role in pillars:
        ws1[f"B{row}"] = p_name
        ws1[f"B{row}"].font = font_bold
        ws1[f"B{row}"].border = thin_border
        ws1[f"B{row}"].fill = fill_ice
        
        ws1[f"C{row}"] = p_script
        ws1[f"C{row}"].font = font_code
        ws1[f"C{row}"].border = thin_border
        
        ws1[f"D{row}"] = p_data
        ws1[f"D{row}"].font = font_normal
        ws1[f"D{row}"].border = thin_border
        
        ws1[f"E{row}"] = p_role
        ws1[f"E{row}"].font = font_normal
        ws1[f"E{row}"].border = thin_border
        ws1[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws1.row_dimensions[row].height = 55
        row += 1

    row += 1
    # Section: Technical Fetching Architecture & Risk Management
    ws1.merge_cells(f"B{row}:E{row}")
    ws1[f"B{row}"] = "3. HOW FETCHING WORKS: HEADERS, NETWORKING, BLOCKS & MITIGATIONS"
    ws1[f"B{row}"].font = font_sec_header
    ws1[f"B{row}"].fill = fill_sec_header
    ws1[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
    ws1.row_dimensions[row].height = 24
    row += 1

    fetch_details = [
        ("HTTP Client & Protocol", "Standard Python urllib.request with HTTPS handlers. Sends direct HTTP/HTTPS GET requests with chunked stream buffers (up to 4MB)."),
        ("Browser Emulation & Headers", "Uses realistic Chrome 126 Windows User-Agent with standard Accept (application/rss+xml, application/xml, */*), Accept-Language, and gzip decompression headers."),
        ("IP Address & Proxy Routing", "Runs via your machine's direct public IP locally (or GitHub runner IP in cloud). Supports standard HTTP_PROXY / HTTPS_PROXY environment variables for rotating residential proxies."),
        ("Why Blocks Occur", "1) Datacenter IP blocking (Cloudflare/Akamai blocking cloud server IPs), 2) Anti-bot JavaScript challenges, 3) Rate limiting (429) from aggressive burst crawling, 4) Missing gov agency contact headers."),
        ("Risk Mitigation Strategy", "1) FEED_HOST_GAP=1.0s prevents 429s, 2) Waterfall fallback switches blocked RSS to XML sitemaps or Google News, 3) Automatic header rotation (Accept: */* retry), 4) SEC_UA / CONTACT_UA headers for regulatory sites.")
    ]

    for f_label, f_text in fetch_details:
        ws1[f"B{row}"] = f_label
        ws1[f"B{row}"].font = font_bold
        ws1[f"B{row}"].border = thin_border
        ws1[f"B{row}"].fill = fill_ice
        
        ws1.merge_cells(f"C{row}:E{row}")
        ws1[f"C{row}"] = f_text
        ws1[f"C{row}"].font = font_normal
        ws1[f"C{row}"].border = thin_border
        ws1[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws1.row_dimensions[row].height = 36
        row += 1

    # -------------------------------------------------------------------------
    # TAB 2: 02_Data_Dictionary (Expanded with 5 Fallback Methods)
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="02_Data_Dictionary")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 28
    ws2.column_dimensions["F"].width = 32
    ws2.column_dimensions["G"].width = 32
    ws2.column_dimensions["H"].width = 32
    ws2.column_dimensions["I"].width = 30

    ws2.merge_cells("B2:I2")
    ws2["B2"] = "RSSFeedChecker — Comprehensive Data Dictionary & 5-Layer Fallback Matrix"
    ws2["B2"].font = font_title
    
    ws2.merge_cells("B3:I3")
    ws2["B3"] = "Every status code, technical root cause, and multi-layer recovery methods (UA Tuning, Sitemaps, GNews, HTML Scraping, Proxy)"
    ws2["B3"].font = font_subtitle

    row = 5
    # Section: Status Codes & Multi-Layer Fallback Matrix
    ws2.merge_cells(f"B{row}:I{row}")
    ws2[f"B{row}"] = "1. FEED CHECK STATUS CLASSIFICATIONS & 5-LAYER FALLBACK RECOVERY STRATEGIES"
    ws2[f"B{row}"].font = font_sec_header
    ws2[f"B{row}"].fill = fill_sec_header
    ws2[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
    ws2.row_dimensions[row].height = 24
    row += 1

    status_headers = [
        "Status Output", "Technical Root Cause", "Health Badge",
        "Fallback 1: Header/UA Tuning", "Fallback 2: Sitemap Discovery",
        "Fallback 3: Google News site: Query", "Fallback 4: Direct HTML / RSS-Bridge",
        "Fallback 5: Proxy / Residential IP"
    ]
    for col_idx, h in enumerate(status_headers, start=2):
        col_letter = get_column_letter(col_idx)
        ws2[f"{col_letter}{row}"] = h
        ws2[f"{col_letter}{row}"].font = font_tbl_header
        ws2[f"{col_letter}{row}"].fill = fill_header
        ws2[f"{col_letter}{row}"].border = header_border
        ws2[f"{col_letter}{row}"].alignment = Alignment(vertical="center", wrap_text=True)
    ws2.row_dimensions[row].height = 28
    row += 1

    statuses_expanded = [
        (
            "OK — fetchable",
            "HTTP 200, valid XML parsed, real article title extracted.",
            "HEALTHY (Pass)",
            "None needed. Working natively.",
            "Active as baseline backup.",
            "Optional supplemental feed.",
            "Not required.",
            "Direct IP is healthy.",
            fill_green, font_green
        ),
        (
            "OK — fetchable (SSL chain incomplete)",
            "Host server missing intermediate certificate authority in SSL bundle (e.g. hutch-med.com).",
            "WARNING",
            "Relax verify via unverified SSL context while logging notice.",
            "Check if HTTP or alternative CDN domain serves valid cert.",
            "Query Google News RSS for mirror articles.",
            "Scrape HTML over standard browser engine.",
            "Not proxy related; certificate issue.",
            fill_yellow, font_yellow
        ),
        (
            "OK (200) — no readable items",
            "HTTP 200 returned but XML payload has 0 items or unparseable custom XML tags.",
            "CRITICAL FAILURE",
            "Inspect raw XML for non-standard item tags (<article>, <entry>, <post>).",
            "Crawl /sitemap.xml to find real article URLs directly.",
            "Use Google News RSS: 'news.google.com/rss/search?q=site:domain.com'.",
            "Scrape company /news/ or /press-releases/ HTML listing directly.",
            "Not proxy related; feed is empty or deprecated.",
            fill_red, font_red
        ),
        (
            "OK (200) — HTML returned, not a feed",
            "URL redirects to standard homepage or HTML newsroom instead of RSS/Atom XML.",
            "CRITICAL FAILURE",
            "Check HTML for <link rel='alternate' type='application/rss+xml'> autodiscovery tag.",
            "Read sitemap.xml / news-sitemap.xml for full URL listing.",
            "Generate Google News search feed for company name.",
            "Extract press release cards from HTML using CSS/regex parsers.",
            "Check if feed URL moved to subfolder.",
            fill_red, font_red
        ),
        (
            "BLOCKED (401/403 Forbidden)",
            "Cloudflare, Akamai, or Imperva WAF bot challenge blocked request.",
            "BLOCKED",
            "Try bare UA (no browser headers) for J&J/Akamai or Chrome UA for AstraZeneca.",
            "Access /sitemap.xml which is often unblocked on CDN level.",
            "Query Google News RSS: 'site:company.com (press OR release OR approval)'.",
            "Use Playwright headless browser or RSS-Bridge generator.",
            "Route request through rotating residential proxy pool (BrightData/SmartProxy).",
            fill_red, font_red
        ),
        (
            "BLOCKED (429 Rate Limited)",
            "Too many requests sent in short burst (common on Squarespace, Wix, PR Newswire).",
            "RATE LIMITED",
            "Increase FEED_HOST_GAP from 1.0s to 3.0s between hits on same host.",
            "Sitemaps are cached on CDN and rarely rate limited.",
            "Google News aggregates without hitting company host repeatedly.",
            "Implement exponential backoff retry (1s, 2s, 4s).",
            "Distribute requests across multiple residential egress IPs.",
            fill_yellow, font_yellow
        ),
        (
            "FAILED (415 Unsupported Media Type)",
            "Server rejects specific Accept header (e.g. application/rss+xml).",
            "CONFIG ERROR",
            "Retry with generic Accept: '*/*' header (already built into checker).",
            "Fetch sitemap via text/xml or application/xml.",
            "Google News bypasses publisher server entirely.",
            "HTML scraper uses text/html Accept header.",
            "Header issue, not proxy related.",
            fill_yellow, font_yellow
        ),
        (
            "FAILED (Timeout > 30s)",
            "Server hung or took longer than FEED_TIMEOUT to respond.",
            "TIMEOUT",
            "Increase timeout to 45s (FEED_TIMEOUT=45) or retry with connection close.",
            "Sitemaps are static XML files and load in <1s.",
            "Google News responds in <500ms from Google servers.",
            "Fetch lightweight static RSS mirror.",
            "Route through low-latency local proxy.",
            fill_red, font_red
        ),
        (
            "FAILED (DNS Error / Connection Refused)",
            "Domain abandoned, DNS host offline, or server refused connection on port 443/80.",
            "DEAD ENDPOINT",
            "Verify if domain migrated (e.g. .com to .io or parent pharma acquisition).",
            "Check parent company sitemap (e.g. Roche for Genentech, Pfizer for Seagen).",
            "Search Google News for company name keywords to find new domain.",
            "Mark feed deprecated in feeds.tsv; update to new domain URL.",
            "Direct network failure; verify internet connection.",
            fill_red, font_red
        )
    ]

    for s_row in statuses_expanded:
        s_name, s_cause, s_badge, fb1, fb2, fb3, fb4, fb5, s_fill, s_font = s_row
        ws2[f"B{row}"] = s_name
        ws2[f"B{row}"].font = font_bold
        ws2[f"B{row}"].border = thin_border
        ws2[f"B{row}"].fill = fill_ice
        
        ws2[f"C{row}"] = s_cause
        ws2[f"C{row}"].font = font_normal
        ws2[f"C{row}"].border = thin_border
        ws2[f"C{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        
        ws2[f"D{row}"] = s_badge
        ws2[f"D{row}"].font = s_font
        ws2[f"D{row}"].fill = s_fill
        ws2[f"D{row}"].border = thin_border
        ws2[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        for c_offset, fb_text in enumerate([fb1, fb2, fb3, fb4, fb5], start=5):
            col_letter = get_column_letter(c_offset)
            ws2[f"{col_letter}{row}"] = fb_text
            ws2[f"{col_letter}{row}"].font = font_normal
            ws2[f"{col_letter}{row}"].border = thin_border
            ws2[f"{col_letter}{row}"].alignment = Alignment(wrap_text=True, vertical="center")
            
        ws2.row_dimensions[row].height = 48
        row += 1

    row += 1
    # Section: Standardized 5-Core Schema & ClinicalTrials.gov Protocol Change Dictionary
    ws2.merge_cells(f"B{row}:I{row}")
    ws2[f"B{row}"] = "2. STANDARDIZED 5-CORE DATA SCHEMA ACROSS ALL WORKFLOWS"
    ws2[f"B{row}"].font = font_sec_header
    ws2[f"B{row}"].fill = fill_sec_header
    ws2[f"B{row}"].alignment = Alignment(vertical="center", indent=1)
    ws2.row_dimensions[row].height = 24
    row += 1

    for col_idx, h in enumerate(["Core Field", "Field Type", "Example Value", "Extraction Mechanism", "Standard Across Methods", "ClinicalTrials.gov Representation", "Missing Value Handling", "Description & Usage"], start=2):
        col_letter = get_column_letter(col_idx)
        ws2[f"{col_letter}{row}"] = h
        ws2[f"{col_letter}{row}"].font = font_tbl_header
        ws2[f"{col_letter}{row}"].fill = fill_header
        ws2[f"{col_letter}{row}"].border = header_border
    ws2.row_dimensions[row].height = 22
    row += 1

    core_schema = [
        ("1. Publication Date", "ISO Date String (YYYY-MM-DD)", "2026-08-25", "Article <time>/meta tags, sitemap <lastmod>, or feed <pubDate>", "M1, M2, M3", "lastUpdatePostDateStruct / startDateStruct", "Falls back to lastmod, then 'Undated'", "Exact date/time when release or protocol change was published."),
        ("2. Title / Headline", "Clean String", "FDA Approves New Obesity Therapy", "Extracted headline with HTML tags stripped and entities decoded", "M1, M2, M3", "briefTitle / officialTitle", "N/A (Title required)", "Primary headline or clinical study title."),
        ("3. Snippet / Excerpt", "Summary Text (1-3 sentences)", "Company announced today that the Phase 3 trial met its primary endpoint...", "Meta description, feed <description>, or first paragraph of article body", "M1, M2, M3", "briefSummary text", "First 250 characters of body text", "Fast contextual summary for scanning."),
        ("4. Full Text Content", "Full Article Body / Diff", "Complete article paragraphs or structured JSON change delta", "Deep article page reader / Playwright full-text extractor", "M1, M2, M3", "Detailed study protocol JSON + History tab delta", "Available via full-text enricher", "Full textual body or exact protocol delta diff."),
        ("5. Authentic URL", "Canonical Web Link", "https://novartis.com/news/media-releases/...", "Canonical <link rel='canonical'> or direct href link", "M1, M2, M3", "https://clinicaltrials.gov/study/NCT01234567", "Required", "Direct authentic link to original company source or NCT page.")
    ]

    for c_row in core_schema:
        f_name, f_type, f_ex, f_ext, f_std, f_ct, f_miss, f_desc = c_row
        ws2[f"B{row}"] = f_name
        ws2[f"B{row}"].font = font_bold
        ws2[f"B{row}"].border = thin_border
        ws2[f"B{row}"].fill = fill_ice
        
        ws2[f"C{row}"] = f_type
        ws2[f"C{row}"].font = font_code
        ws2[f"C{row}"].border = thin_border
        
        ws2[f"D{row}"] = f_ex
        ws2[f"D{row}"].font = font_normal
        ws2[f"D{row}"].border = thin_border
        
        ws2[f"E{row}"] = f_ext
        ws2[f"E{row}"].font = font_normal
        ws2[f"E{row}"].border = thin_border
        
        ws2[f"F{row}"] = f_std
        ws2[f"F{row}"].font = font_bold
        ws2[f"F{row}"].border = thin_border
        ws2[f"F{row}"].alignment = Alignment(horizontal="center")
        
        ws2[f"G{row}"] = f_ct
        ws2[f"G{row}"].font = font_normal
        ws2[f"G{row}"].border = thin_border
        
        ws2[f"H{row}"] = f_miss
        ws2[f"H{row}"].font = font_normal
        ws2[f"H{row}"].border = thin_border
        
        ws2[f"I{row}"] = f_desc
        ws2[f"I{row}"].font = font_normal
        ws2[f"I{row}"].border = thin_border
        ws2[f"I{row}"].alignment = Alignment(wrap_text=True, vertical="center")
        
        ws2.row_dimensions[row].height = 36
        row += 1

    # -------------------------------------------------------------------------
    # TAB 3: 03_Feeds_Master (518 feeds)
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="03_Feeds_Master (518)")
    ws3.views.sheetView[0].showGridLines = True
    
    with open(os.path.join(WORKSPACE, "feeds.tsv"), encoding="utf-8") as f:
        feed_rows = list(csv.reader(f, delimiter="\t"))

    feed_headers = ["Feed ID", "Feed URL", "Label / Source Name", "Category / Coverage Type", "Protocol"]
    for col_idx, h in enumerate(feed_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws3[f"{col_letter}1"] = h
        ws3[f"{col_letter}1"].font = font_tbl_header
        ws3[f"{col_letter}1"].fill = fill_header
        ws3[f"{col_letter}1"].border = header_border
        ws3[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws3.row_dimensions[1].height = 24

    for r_idx, r in enumerate(feed_rows[1:], start=2):
        fid = r[0] if len(r) > 0 else f"Feed_{r_idx-1:03d}"
        url = r[1] if len(r) > 1 else ""
        lbl = r[2] if len(r) > 2 else ""
        
        cat = "Industry News"
        if "fda.gov" in url or "ema.europa" in url or "sec.gov" in url or "nih.gov" in url or "regulatory" in lbl.lower():
            cat = "Regulatory & Government"
        elif "nature.com" in url or "thelancet.com" in url or "nejm.org" in url or "cell.com" in url or "jamanetwork" in url or "journal" in lbl.lower():
            cat = "Medical Journal"
        elif "endpoints" in url or "statnews" in url or "fierce" in url or "biospace" in url or "bioworld" in url or "genengnews" in url:
            cat = "Trade Press"
        elif "news.google.com" in url:
            cat = "Google News Aggregation"
        elif "prnewswire" in url or "businesswire" in url or "globenewswire" in url:
            cat = "Newswire"

        proto = "HTTPS" if url.startswith("https") else "HTTP"

        ws3[f"A{r_idx}"] = fid
        ws3[f"A{r_idx}"].font = font_code
        ws3[f"A{r_idx}"].border = thin_border
        ws3[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws3[f"B{r_idx}"] = url
        ws3[f"B{r_idx}"].font = font_link
        ws3[f"B{r_idx}"].border = thin_border
        ws3[f"B{r_idx}"].hyperlink = url
        
        ws3[f"C{r_idx}"] = lbl
        ws3[f"C{r_idx}"].font = font_bold
        ws3[f"C{r_idx}"].border = thin_border
        
        ws3[f"D{r_idx}"] = cat
        ws3[f"D{r_idx}"].font = font_normal
        ws3[f"D{r_idx}"].border = thin_border
        
        ws3[f"E{r_idx}"] = proto
        ws3[f"E{r_idx}"].font = font_code
        ws3[f"E{r_idx}"].border = thin_border
        ws3[f"E{r_idx}"].alignment = Alignment(horizontal="center")
        
        ws3.row_dimensions[r_idx].height = 20

    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 75
    ws3.column_dimensions["C"].width = 38
    ws3.column_dimensions["D"].width = 25
    ws3.column_dimensions["E"].width = 12
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:E{len(feed_rows)}"

    # -------------------------------------------------------------------------
    # TAB 4: 04_Companies_Master (616 companies)
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet(title="04_Companies_Master (616)")
    ws4.views.sheetView[0].showGridLines = True

    with open(os.path.join(WORKSPACE, "companies.tsv"), encoding="utf-8") as f:
        comp_rows = list(csv.reader(f, delimiter="\t"))

    comp_headers = ["Company Name", "Domain", "Newsroom Path / Custom Override", "Tracking Strategy", "Notes / Feed Promotion"]
    for col_idx, h in enumerate(comp_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws4[f"{col_letter}1"] = h
        ws4[f"{col_letter}1"].font = font_tbl_header
        ws4[f"{col_letter}1"].fill = fill_header
        ws4[f"{col_letter}1"].border = header_border
        ws4[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws4.row_dimensions[1].height = 24

    for r_idx, r in enumerate(comp_rows[1:], start=2):
        c_name = r[0] if len(r) > 0 else ""
        c_dom = r[1] if len(r) > 1 else ""
        c_path = r[2] if len(r) > 2 else ""
        
        strat = "Sitemap & RSS Discovery"
        notes = "Automated discovery via robots.txt -> sitemaps. Discovered working feeds get promoted to Master Feeds list."
        if c_path.strip():
            strat = "Tuned Newsroom Path + Sitemap"
            notes = f"Explicit newsroom path override ({c_path}). Prioritizes direct targeted scraping."
            
        ws4[f"A{r_idx}"] = c_name
        ws4[f"A{r_idx}"].font = font_bold
        ws4[f"A{r_idx}"].border = thin_border
        ws4[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws4[f"B{r_idx}"] = c_dom
        ws4[f"B{r_idx}"].font = font_code
        ws4[f"B{r_idx}"].border = thin_border
        
        ws4[f"C{r_idx}"] = c_path
        ws4[f"C{r_idx}"].font = font_code if c_path else font_normal
        ws4[f"C{r_idx}"].border = thin_border
        
        ws4[f"D{r_idx}"] = strat
        ws4[f"D{r_idx}"].font = font_normal
        ws4[f"D{r_idx}"].border = thin_border
        
        ws4[f"E{r_idx}"] = notes
        ws4[f"E{r_idx}"].font = font_normal
        ws4[f"E{r_idx}"].border = thin_border
        
        ws4.row_dimensions[r_idx].height = 20

    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 28
    ws4.column_dimensions["C"].width = 35
    ws4.column_dimensions["D"].width = 30
    ws4.column_dimensions["E"].width = 65
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:E{len(comp_rows)}"

    # -------------------------------------------------------------------------
    # TAB 5: 05_Sample_Test_Set (22 companies)
    # -------------------------------------------------------------------------
    ws5 = wb.create_sheet(title="05_Sample_Test_Set (22)")
    ws5.views.sheetView[0].showGridLines = True

    with open(os.path.join(WORKSPACE, "companies_sample.tsv"), encoding="utf-8") as f:
        sample_rows = list(csv.reader(f, delimiter="\t"))

    s_headers = ["Company Name", "Domain", "Newsroom Path", "Test Purpose / Edge Case Benchmarked"]
    for col_idx, h in enumerate(s_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws5[f"{col_letter}1"] = h
        ws5[f"{col_letter}1"].font = font_tbl_header
        ws5[f"{col_letter}1"].fill = fill_header
        ws5[f"{col_letter}1"].border = header_border
        ws5[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws5.row_dimensions[1].height = 24

    sample_edge_cases = {
        "merck.com": "Verifies custom /news/ path and heavy sitemap index traversal.",
        "novartis.com": "Verifies enterprise media-releases path and multi-language filtering.",
        "gsk.com": "Verifies UK /en-gb/ locale prefix routing.",
        "amgen.com": "Tests standard enterprise PR layout.",
        "jnj.com": "Akamai bot challenge edge case and Google News job-posting filtering.",
        "biomarin.com": "Clean news path verification.",
        "acadia.com": "Multi-language PDF deduplication (de-de, fr-fr, it-it filtering).",
        "beamtx.com": "Fast-moving gene editing biotech sitemap check.",
        "apellis.com": "Filters out board member profile pages and analyst coverage from news feed.",
        "blueprintmedicines.com": "Checks precision oncology PR sitemap.",
        "theravance.com": "Filters corporate bio pages and analyst commentary from items.",
        "mallinckrodt.com": "Enterprise legacy PR archive traversal.",
        "neurocrine.com": "CNS neurology press release sitemap indexing.",
        "ionispharma.com": "Filters out non-news corporate executive additions.",
        "bavarian-nordic.com": "European vaccine biotech sitemap verification.",
        "zealandpharma.com": "Danish obesity peptide biotech tracking.",
        "faesfarma.com": "Spanish pharmaceutical domain routing.",
        "basilea.com": "Swiss anti-infectives biotech newsroom check.",
        "calicolabs.com": "Alphabet life sciences subsidiary news listing test.",
        "glpg.com": "Belgian biotech sitemap parser.",
        "morphictx.com": "Integrin biotech news release extraction.",
        "vandapharma.com": "Commercial-stage biopharma PR verification."
    }

    for r_idx, r in enumerate(sample_rows[1:], start=2):
        c_name = r[0] if len(r) > 0 else ""
        c_dom = r[1] if len(r) > 1 else ""
        c_path = r[2] if len(r) > 2 else ""
        c_edge = sample_edge_cases.get(c_dom.lower(), "Benchmark sample domain for fast validation.")

        ws5[f"A{r_idx}"] = c_name
        ws5[f"A{r_idx}"].font = font_bold
        ws5[f"A{r_idx}"].border = thin_border
        ws5[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws5[f"B{r_idx}"] = c_dom
        ws5[f"B{r_idx}"].font = font_code
        ws5[f"B{r_idx}"].border = thin_border
        
        ws5[f"C{r_idx}"] = c_path
        ws5[f"C{r_idx}"].font = font_code if c_path else font_normal
        ws5[f"C{r_idx}"].border = thin_border
        
        ws5[f"D{r_idx}"] = c_edge
        ws5[f"D{r_idx}"].font = font_normal
        ws5[f"D{r_idx}"].border = thin_border
        
        ws5.row_dimensions[r_idx].height = 22

    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 28
    ws5.column_dimensions["C"].width = 35
    ws5.column_dimensions["D"].width = 75
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = f"A1:D{len(sample_rows)}"

    # -------------------------------------------------------------------------
    # TAB 6: 06_Indications_Radar (18 indications)
    # -------------------------------------------------------------------------
    ws6 = wb.create_sheet(title="06_Indications_Radar (18)")
    ws6.views.sheetView[0].showGridLines = True

    with open(os.path.join(WORKSPACE, "indications.tsv"), encoding="utf-8") as f:
        ind_rows = list(csv.reader(f, delimiter="\t"))

    ind_headers = ["Indication / Theme", "Synonyms & Keywords", "ClinicalTrials.gov Condition Term", "Exclude Terms (Negations)", "Broad Feed Query Logic", "Clinical/Reg Feed Query Logic", "CT.gov Protocol Change Endpoint"]
    for col_idx, h in enumerate(ind_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws6[f"{col_letter}1"] = h
        ws6[f"{col_letter}1"].font = font_tbl_header
        ws6[f"{col_letter}1"].fill = fill_header
        ws6[f"{col_letter}1"].border = header_border
        ws6[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws6.row_dimensions[1].height = 24

    for r_idx, r in enumerate(ind_rows[1:], start=2):
        ind = r[0] if len(r) > 0 else ""
        syn = r[1] if len(r) > 1 else ""
        ct_term = r[2] if len(r) > 2 else ""
        exc = r[3] if len(r) > 3 else ""
        
        all_terms = [ind] + [s.strip() for s in syn.split(",") if s.strip()]
        quoted_terms = [f'"{t}"' if " " in t else t for t in all_terms]
        broad_q = f"({' OR '.join(quoted_terms[:4])}) AND (drug OR therapy OR trial OR FDA)"
        clin_q = f"({' OR '.join(quoted_terms[:3])}) AND (Phase OR trial OR FDA OR approval OR efficacy)"
        ct_q = f"https://clinicaltrials.gov/api/v2/studies?query.cond={ct_term.replace(' ', '+')}&sort=LastUpdatePostDate:desc" if ct_term else "N/A"

        ws6[f"A{r_idx}"] = ind
        ws6[f"A{r_idx}"].font = font_bold
        ws6[f"A{r_idx}"].border = thin_border
        ws6[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws6[f"B{r_idx}"] = syn
        ws6[f"B{r_idx}"].font = font_normal
        ws6[f"B{r_idx}"].border = thin_border
        
        ws6[f"C{r_idx}"] = ct_term
        ws6[f"C{r_idx}"].font = font_bold if ct_term else font_normal
        ws6[f"C{r_idx}"].border = thin_border
        
        ws6[f"D{r_idx}"] = exc
        ws6[f"D{r_idx}"].font = font_normal
        ws6[f"D{r_idx}"].border = thin_border
        
        ws6[f"E{r_idx}"] = broad_q
        ws6[f"E{r_idx}"].font = font_code
        ws6[f"E{r_idx}"].border = thin_border
        
        ws6[f"F{r_idx}"] = clin_q
        ws6[f"F{r_idx}"].font = font_code
        ws6[f"F{r_idx}"].border = thin_border
        
        ws6[f"G{r_idx}"] = ct_q
        ws6[f"G{r_idx}"].font = font_code
        ws6[f"G{r_idx}"].border = thin_border
        
        ws6.row_dimensions[r_idx].height = 24

    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 45
    ws6.column_dimensions["C"].width = 30
    ws6.column_dimensions["D"].width = 25
    ws6.column_dimensions["E"].width = 45
    ws6.column_dimensions["F"].width = 45
    ws6.column_dimensions["G"].width = 45
    ws6.freeze_panes = "A2"
    ws6.auto_filter.ref = f"A1:G{len(ind_rows)}"

    # -------------------------------------------------------------------------
    # TAB 7: 07_Seen_State_Tracking
    # -------------------------------------------------------------------------
    ws7 = wb.create_sheet(title="07_Seen_State_Tracking")
    ws7.views.sheetView[0].showGridLines = True

    with open(os.path.join(WORKSPACE, "state", "seen.json"), encoding="utf-8") as f:
        seen_data = json.load(f)

    seen_headers = ["Tracked Domain", "Total Press Releases Cached", "Most Recent Tracked URL (Sample)", "State Lifecycle & Retention"]
    for col_idx, h in enumerate(seen_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws7[f"{col_letter}1"] = h
        ws7[f"{col_letter}1"].font = font_tbl_header
        ws7[f"{col_letter}1"].fill = fill_header
        ws7[f"{col_letter}1"].border = header_border
        ws7[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws7.row_dimensions[1].height = 24

    r_idx = 2
    for dom, urls in sorted(seen_data.items(), key=lambda x: len(x[1]), reverse=True):
        count = len(urls)
        sample_url = urls[0] if urls else "None"
        retention = f"Capped at {len(urls)} URLs. Recency preserved in insertion order."

        ws7[f"A{r_idx}"] = dom
        ws7[f"A{r_idx}"].font = font_bold
        ws7[f"A{r_idx}"].border = thin_border
        ws7[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws7[f"B{r_idx}"] = count
        ws7[f"B{r_idx}"].font = font_normal
        ws7[f"B{r_idx}"].border = thin_border
        ws7[f"B{r_idx}"].alignment = Alignment(horizontal="center")
        
        ws7[f"C{r_idx}"] = sample_url
        ws7[f"C{r_idx}"].font = font_link
        ws7[f"C{r_idx}"].border = thin_border
        if sample_url.startswith("http"):
            ws7[f"C{r_idx}"].hyperlink = sample_url
            
        ws7[f"D{r_idx}"] = retention
        ws7[f"D{r_idx}"].font = font_normal
        ws7[f"D{r_idx}"].border = thin_border
        
        ws7.row_dimensions[r_idx].height = 20
        r_idx += 1

    ws7.column_dimensions["A"].width = 28
    ws7.column_dimensions["B"].width = 28
    ws7.column_dimensions["C"].width = 80
    ws7.column_dimensions["D"].width = 50
    ws7.freeze_panes = "A2"
    ws7.auto_filter.ref = f"A1:D{r_idx-1}"

    # -------------------------------------------------------------------------
    # TAB 8: 08_Config_and_Settings
    # -------------------------------------------------------------------------
    ws8 = wb.create_sheet(title="08_Config_and_Settings")
    ws8.views.sheetView[0].showGridLines = True

    cfg_headers = ["Workflow / Script", "Setting / Parameter", "Default Value", "Valid Range / Options", "Operational Role & Impact"]
    for col_idx, h in enumerate(cfg_headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws8[f"{col_letter}1"] = h
        ws8[f"{col_letter}1"].font = font_tbl_header
        ws8[f"{col_letter}1"].fill = fill_header
        ws8[f"{col_letter}1"].border = header_border
        ws8[f"{col_letter}1"].alignment = Alignment(vertical="center")
    ws8.row_dimensions[1].height = 24

    configs = [
        ("Method 1: Feed Check", "FEED_TIMEOUT", "30", "Seconds (int/float)", "HTTP timeout per feed attempt. Prevents slow servers from hanging the job."),
        ("Method 1: Feed Check", "FEED_WORKERS", "8", "Integer (1-32)", "Number of parallel threads fetching feeds simultaneously."),
        ("Method 1: Feed Check", "FEED_HOST_GAP", "1.0", "Seconds (float)", "Minimum delay between consecutive requests to the same host domain (prevents 429 rate limits)."),
        ("Method 1: Feed Check", "FEED_RETRIES", "1", "Integer (0-3)", "Extra retry attempts for transient errors (HTTP 0, 401, 403, 429, 5xx)."),
        ("Method 1: Feed Check", "FEED_OUT_DIR", "results", "Path string", "Directory where results.tsv, results.csv, and results.xlsx are written."),
        
        ("Method 2: Newsroom Watch", "NEWS_DAYS_BACK", "3", "Integer (e.g. 1, 2, 3, 7, 30)", "Strict time-window filter. Only reports releases published in the last N days (default: 3 days / 72 hours)."),
        ("Method 2: Newsroom Watch", "NEWS_TIMEOUT", "25", "Seconds (int/float)", "Per-request timeout for company homepages, robots.txt, sitemaps, and PR pages."),
        ("Method 2: Newsroom Watch", "NEWS_WORKERS", "6", "Integer (1-16)", "Parallel thread pool size for company crawling."),
        ("Method 2: Newsroom Watch", "NEWS_MAX_SUBSITEMAPS", "20", "Integer", "Max child sitemaps inspected per company (prevents runaway indexing)."),
        ("Method 2: Newsroom Watch", "NEWS_STATE_CAP", "6000", "Integer", "Maximum historical release URLs remembered per company domain in seen.json."),
        ("Method 2: Newsroom Watch", "NEWS_SEED", "false", "Boolean (true/false)", "If true, silently logs all discovered URLs into seen.json without generating 'NEW' alerts."),
        ("Method 2: Newsroom Watch", "SEC_UA", "None", "String ('Name email@domain.com')", "Custom user-agent header required by SEC.gov EDGAR feeds to prevent 403 blocks."),
        
        ("Method 3: Indication Feeds", "IND_TIMEOUT", "30", "Seconds (int/float)", "Per-request verification timeout for generated Google News and CT.gov feeds."),
        ("Method 3: Indication Feeds", "IND_WORKERS", "4", "Integer (1-8)", "Worker threads for parallel feed verification."),
        ("Method 3: Indication Feeds", "CT_DIFF_WINDOW_DAYS", "3", "Integer (e.g. 1, 3, 7, 14)", "ClinicalTrials.gov protocol delta timeframe. Detects new trial registrations & protocol changes."),
        ("Method 3: Indication Feeds", "CONTACT_UA", "None", "String ('Name email@domain.com')", "Contact header for government clinical registries.")
    ]

    for r_idx, (w_name, var_name, def_val, v_type, desc) in enumerate(configs, start=2):
        ws8[f"A{r_idx}"] = w_name
        ws8[f"A{r_idx}"].font = font_bold
        ws8[f"A{r_idx}"].border = thin_border
        ws8[f"A{r_idx}"].fill = fill_ice if r_idx % 2 == 0 else fill_white
        
        ws8[f"B{r_idx}"] = var_name
        ws8[f"B{r_idx}"].font = font_code
        ws8[f"B{r_idx}"].border = thin_border
        
        ws8[f"C{r_idx}"] = def_val
        ws8[f"C{r_idx}"].font = font_code
        ws8[f"C{r_idx}"].border = thin_border
        ws8[f"C{r_idx}"].alignment = Alignment(horizontal="center")
        
        ws8[f"D{r_idx}"] = v_type
        ws8[f"D{r_idx}"].font = font_normal
        ws8[f"D{r_idx}"].border = thin_border
        
        ws8[f"E{r_idx}"] = desc
        ws8[f"E{r_idx}"].font = font_normal
        ws8[f"E{r_idx}"].border = thin_border
        ws8[f"E{r_idx}"].alignment = Alignment(wrap_text=True, vertical="center")
        
        ws8.row_dimensions[r_idx].height = 24

    ws8.column_dimensions["A"].width = 28
    ws8.column_dimensions["B"].width = 30
    ws8.column_dimensions["C"].width = 16
    ws8.column_dimensions["D"].width = 25
    ws8.column_dimensions["E"].width = 75
    ws8.freeze_panes = "A2"
    ws8.auto_filter.ref = f"A1:E{len(configs)+1}"

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    wb.save(OUTPUT_FILE)
    print(f"Master Excel successfully updated at: {OUTPUT_FILE}")

if __name__ == "__main__":
    create_master_workbook()
