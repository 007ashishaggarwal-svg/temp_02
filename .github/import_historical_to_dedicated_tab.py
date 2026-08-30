import os
import sys
import json
import re
import html
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
MASTER_XLSX = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")
SOURCE_XLSX = os.path.join(WORKSPACE, "Alternative", "Results to add.xlsx")

# 1. Load Master Catalogs
sys.path.insert(0, os.path.join(WORKSPACE, ".github", "scripts"))
from match_keywords import KeywordMatcher
from clustering_engine import MultiVectorClusterEngine

matcher = KeywordMatcher(xlsx_path=MASTER_XLSX)
cluster_engine = MultiVectorClusterEngine(xlsx_path=MASTER_XLSX)

print(f"▶ Loading source workbook: {SOURCE_XLSX}...")
wb_src = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
ws_src = wb_src["RESULTS"]

# Therapeutic Desk Mapping
DESK_MAP = {
    "obesity": "Metabolic & Obesity Desk",
    "rheumatoid arthritis": "Immunology & Inflammation Desk",
    "small cell lung cancer": "Oncology & Immuno-Oncology Desk",
    "nsclc": "Oncology & Immuno-Oncology Desk",
    "breast cancer": "Oncology & Immuno-Oncology Desk",
    "myeloma": "Multiple Myeloma Desk",
    "alzheimer's": "Neuroscience & CNS Desk",
    "schizophrenia": "Neuroscience & CNS Desk",
    "companies": "Corporate Strategy & M&A Desk",
    "regulatory": "Regulatory & Health Authority Desk"
}

def clean_text(val):
    if val is None:
        return ""
    txt = str(val).strip()
    txt = html.unescape(txt)
    txt = re.sub(r'<figure[^>]*>.*?</figure>', ' ', txt, flags=re.DOTALL | re.IGNORECASE)
    txt = re.sub(r'<img[^>]*>', ' ', txt, flags=re.IGNORECASE)
    txt = re.sub(r'<[^>]+>', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt

items_to_import = []

for r in range(2, ws_src.max_row + 1):
    val_date = ws_src.cell(r, 1).value
    val_title = ws_src.cell(r, 5).value
    if not val_date and not val_title:
        continue

    # Extract Date / Time
    date_obj = ws_src.cell(r, 12).value or ws_src.cell(r, 1).value
    pub_date = ""
    pub_time = "00:00"
    if isinstance(date_obj, datetime):
        pub_date = date_obj.strftime("%Y-%m-%d")
        pub_time = date_obj.strftime("%H:%M")
    elif date_obj:
        date_str = str(date_obj).strip()
        pub_date = date_str[:10]
        if len(date_str) > 11:
            pub_time = date_str[11:16]

    project_raw = clean_text(ws_src.cell(r, 2).value) or "General Biopharma"
    source_name = clean_text(ws_src.cell(r, 3).value) or "Biopharma Feed"
    source_type = clean_text(ws_src.cell(r, 4).value) or "Direct RSS"
    headline = clean_text(ws_src.cell(r, 5).value)
    excerpt = clean_text(ws_src.cell(r, 6).value)
    raw_url = str(ws_src.cell(r, 7).value or "").strip()
    ai_summary = clean_text(ws_src.cell(r, 8).value)
    ai_provider = clean_text(ws_src.cell(r, 9).value)
    matched_kw = clean_text(ws_src.cell(r, 10).value)
    signal_type_raw = clean_text(ws_src.cell(r, 11).value)
    discovered_at = str(ws_src.cell(r, 13).value or f"{pub_date} 00:00:00 UTC")

    # Signal Type Standardizer
    sig_lower = signal_type_raw.lower()
    if "approv" in sig_lower or "regulatory" in sig_lower:
        signal_type = "regulatory"
    elif "clinical_pos" in sig_lower or "win" in sig_lower:
        signal_type = "clinical_pos"
    elif "clinical_neg" in sig_lower or "fail" in sig_lower or "hold" in sig_lower:
        signal_type = "clinical_neg"
    elif "corporate" in sig_lower or "deal" in sig_lower or "licens" in sig_lower:
        signal_type = "corporate"
    elif "commercial" in sig_lower or "launch" in sig_lower:
        signal_type = "commercial"
    else:
        signal_type = "general"

    # Desk Routing
    desk = DESK_MAP.get(project_raw.lower(), "Therapeutic CI Desk")

    # Match scoring via Engine
    res = matcher.match(headline, excerpt, source_class=source_type, source_name=source_name)
    item_score = res.get("relevance_score", 65)
    if item_score < 40:
        item_score = 60

    if item_score >= 80:
        tier = "🔴 Tier 1 (Urgent)"
    elif item_score >= 60:
        tier = "🟡 Tier 2 (Daily)"
    else:
        tier = "🟢 Tier 3 (Weekly)"

    # Source classification
    if "Google" in source_name or "Alert" in source_name:
        s_class = "1. News Aggregator & Trade Press"
    elif "Company" in source_name or "Official" in source_name or "IR" in source_name:
        s_class = "4. Public Company Investor Relations & SEC EDGAR"
    elif "Regulatory" in source_name or "EMA" in source_name or "FDA" in source_name:
        s_class = "3. Health Authority & Regulatory Body"
    elif "PubMed" in source_name:
        s_class = "5. Biomedical Preprint / Literature"
    else:
        s_class = "1. News Aggregator & Trade Press"

    matched_comp = res.get("matched_company", "Not Identified")
    if matched_comp == "Not Identified":
        # Check title for known biopharma companies
        for alias, c_name in cluster_engine.company_alias_map.items():
            if re.search(r'\b' + re.escape(alias) + r'\b', headline.lower()):
                matched_comp = c_name
                break

    item = {
        "published_date": pub_date,
        "published_time": pub_time,
        "headline": headline,
        "project_name": project_raw,
        "signal_type": signal_type,
        "relevance_score": f"{item_score}/100",
        "priority_tier": tier,
        "routed_desk": desk,
        "matched_keywords": matched_kw or res.get("matched_keywords_str", "--"),
        "source_name": source_name,
        "source_class": s_class,
        "raw_url": raw_url,
        "snippet": excerpt, # Pure excerpt (never mixed with AI summary)
        "cluster_id": "",
        "cluster_hint": "{}",
        "discovery_pillar": "Pillar 1: Publisher Feeds",
        "extraction_vector": "1. Native RSS Feed",
        "discovered_at": discovered_at,
        "full_text": excerpt, # Pure excerpt (never mixed with AI summary)
        "event_id": f"HIST_{len(items_to_import)+1:04d}",
        "ai_summary": ai_summary, # Dedicated Column U
        "matched_company": matched_comp
    }
    items_to_import.append(item)

print(f"✓ Parsed {len(items_to_import)} historical intelligence records.")

# Run 5-Vector Event Clustering across historical records
print("▶ Running 5-Vector Event Clustering across historical events...")
clustered_feed = cluster_engine.cluster_feed_items(items_to_import)

# Write to Master Workbook in tab: 'IMPORTED HISTORICAL'
print(f"▶ Writing to tab 'IMPORTED HISTORICAL' in {MASTER_XLSX}...")
wb_master = openpyxl.load_workbook(MASTER_XLSX)

tab_name = "IMPORTED HISTORICAL"
if tab_name in wb_master.sheetnames:
    del wb_master[tab_name]

# Insert after Results tab (index 2)
ws_hist = wb_master.create_sheet(title=tab_name, index=2)
ws_hist.views.sheetView[0].showGridLines = True

# Palettes & Styling
NAVY_DARK = "1B365D"
BLUE_DARK = "1F4E79"
PURPLE_DARK = "382D5C"
TEAL_DARK = "0E5A5E"
GREEN_DARK = "155724"
WHITE = "FFFFFF"
BORDER_COLOR = "D0D7DE"
ICE_BLUE = "F0F5FA"

fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
fill_teal = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
fill_purple = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
fill_green_dark = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
fill_ice_blue = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")

font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_data = Font(name="Calibri", size=10, color="000000")
font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
font_code = Font(name="Consolas", size=9, color="555555")

thin_border = Border(
    left=Side(style='thin', color=BORDER_COLOR),
    right=Side(style='thin', color=BORDER_COLOR),
    top=Side(style='thin', color=BORDER_COLOR),
    bottom=Side(style='thin', color=BORDER_COLOR)
)

headers_21 = [
    ("Published Date", 16, fill_navy),
    ("Published Time (UTC)", 20, fill_navy),
    ("Event Headline", 48, fill_navy),
    ("Project / Indication Theme", 26, fill_blue),
    ("Signal Type", 20, fill_teal),
    ("Relevance Score", 18, fill_teal),
    ("Priority Tier", 18, fill_green_dark),
    ("Routed Desk / CI Workstream", 28, fill_purple),
    ("Matched Biopharma Catalyst", 30, fill_navy),
    ("Source Entity Name", 30, fill_navy),
    ("Source Classification", 28, fill_blue),
    ("Direct Publisher / SEC URL", 55, fill_navy),
    ("Editorial Snippet / Summary", 50, fill_navy),
    ("Cluster ID", 35, fill_teal),
    ("Cluster Hint JSON", 38, fill_teal),
    ("Discovery Pillar", 24, fill_purple),
    ("Extraction Vector / Method", 28, fill_purple),
    ("Discovered At (UTC)", 24, fill_navy),
    ("Full Body Excerpt", 50, fill_navy),
    ("Event UUID", 16, fill_navy),
    ("AI Summary", 55, fill_purple),  # Column U
]

for col_idx, (h_name, width, h_fill) in enumerate(headers_21, start=1):
    col_letter = get_column_letter(col_idx)
    cell = ws_hist.cell(row=1, column=col_idx, value=h_name)
    cell.font = font_tbl_header
    cell.fill = h_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_hist.column_dimensions[col_letter].width = width

ws_hist.row_dimensions[1].height = 32

# Populate Rows
for r_idx, item in enumerate(clustered_feed, start=2):
    row_values = [
        item["published_date"],
        item["published_time"],
        item["headline"],
        item["project_name"],
        item["signal_type"],
        item["relevance_score"],
        item["priority_tier"],
        item["routed_desk"],
        item["matched_keywords"],
        item["source_name"],
        item["source_class"],
        item["raw_url"],
        item["snippet"],
        item.get("cluster_id", ""),
        item.get("cluster_hint", "{}"),
        item["discovery_pillar"],
        item["extraction_vector"],
        item["discovered_at"],
        item["full_text"],
        item["event_id"],
        item.get("ai_summary", "")  # Column U
    ]

    fill_to_use = fill_ice_blue if r_idx % 2 == 0 else PatternFill(fill_type=None)

    for c_idx, val in enumerate(row_values, start=1):
        cell = ws_hist.cell(row=r_idx, column=c_idx, value=val)
        cell.font = font_data
        cell.border = thin_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use

        if c_idx == 3: # Headline
            cell.font = font_bold
        elif c_idx == 12: # URL
            cell.font = font_link
        elif c_idx in (14, 15, 20): # Code/JSON/UUID
            cell.font = font_code

        cell.alignment = Alignment(
            horizontal="left" if c_idx in (3, 9, 10, 12, 13, 15, 19, 21) else "center",
            vertical="center"
        )
    ws_hist.row_dimensions[r_idx].height = 22

# Safe save
temp_save = MASTER_XLSX + ".tmp.xlsx"
wb_master.save(temp_save)
if os.path.exists(MASTER_XLSX):
    os.replace(temp_save, MASTER_XLSX)
else:
    os.rename(temp_save, MASTER_XLSX)

print(f"🎉 Successfully created tab '{tab_name}' with {len(clustered_feed)} structured historical events in {MASTER_XLSX}!")
