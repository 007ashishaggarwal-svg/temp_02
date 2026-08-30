#!/usr/bin/env python3
"""
Add two combined source tabs to RSSFeedChecker_Master_Guide_and_Data.xlsx:
1. Preview_01_Combined_Wide: Wide schema (1 row per entity, multi-column fallback cascade)
2. Preview_02_Combined_Long: Long schema (each approach/vector on separate sequential rows per entity)
"""

import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

def build_combined_previews():
    print(f"Loading workbook from {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Styles
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    PURPLE_DARK = "382D5C"
    WHITE = "FFFFFF"
    BORDER_COLOR = "D0D7DE"
    ICE_BLUE = "EBF2FA"
    ICE_PURPLE = "F3EBF7"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_purple = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
    fill_ice_blue = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_ice_purple = PatternFill(start_color=ICE_PURPLE, end_color=ICE_PURPLE, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    font_code = Font(name="Consolas", size=9, color="111111")

    # Status Badges
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_green = Font(name="Calibri", size=10, bold=True, color="155724")

    fill_amber = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_amber = Font(name="Calibri", size=10, bold=True, color="856404")

    fill_blue_badge = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    font_blue_badge = Font(name="Calibri", size=10, bold=True, color="0C5460")

    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_red = Font(name="Calibri", size=10, bold=True, color="721C24")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # -------------------------------------------------------------------------
    # 1. VERSION 1: Preview_01_Combined_Wide
    # -------------------------------------------------------------------------
    TAB_WIDE = "Preview_01_Combined_Wide"
    if TAB_WIDE in wb.sheetnames:
        del wb[TAB_WIDE]

    ws_wide = wb.create_sheet(title=TAB_WIDE)
    ws_wide.views.sheetView[0].showGridLines = True

    wide_headers = [
        ("Source ID", 12, fill_navy),
        ("Organization / Entity Name", 30, fill_navy),
        ("Source Classification", 22, fill_navy),
        ("Domain / Ticker", 24, fill_navy),
        ("Approach 1: Primary Endpoint (RSS / API)", 45, fill_blue),
        ("Approach 2: XML Sitemap (<lastmod>)", 40, fill_blue),
        ("Approach 3: Google News 'site:' Fallback", 48, fill_blue),
        ("Approach 4: Direct HTML Newsroom URL", 42, fill_blue),
        ("Approach 5: Regulatory / SEC / Backup", 42, fill_blue),
        ("Active Working Route", 26, fill_purple),
        ("Health & Verification Status", 24, fill_purple),
        ("Therapy Focus / Coverage Scope", 30, fill_navy),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(wide_headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_wide.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_wide.column_dimensions[col_letter].width = width
    ws_wide.row_dimensions[1].height = 28

    # Extract sample entities representing media, regulatory, companies with RSS, companies with sitemaps, companies with GNews
    sample_wide_data = [
        # Media / Trade Press
        ("SRC_001", "STAT News", "Industry Trade Media", "statnews.com", "https://www.statnews.com/feed/", "https://www.statnews.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Astatnews.com&hl=en-US", "https://www.statnews.com/latest/", "None (Direct Media)", "Approach 1: Native RSS", "● 200 OK (Healthy)", "General Biopharma & Biotech"),
        ("SRC_002", "Endpoints News", "Industry Trade Media", "endpts.com", "https://endpts.com/feed/", "https://endpts.com/sitemap_index.xml", "https://news.google.com/rss/search?q=site%3Aendpts.com&hl=en-US", "https://endpts.com/news/", "None (Direct Media)", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Biopharma R&D, Deals & VC"),
        ("SRC_003", "Fierce Pharma", "Industry Trade Media", "fiercepharma.com", "https://www.fiercepharma.com/rss/xml", "https://www.fiercepharma.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Afiercepharma.com&hl=en-US", "https://www.fiercepharma.com/pharma", "None (Direct Media)", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Commercial Pharma & Manufacturing"),
        ("SRC_004", "BioWorld", "Industry Trade Media", "bioworld.com", "https://www.bioworld.com/rss/all", "https://www.bioworld.com/sitemaps.xml", "https://news.google.com/rss/search?q=site%3Abioworld.com&hl=en-US", "https://www.bioworld.com/articles", "None (Direct Media)", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Global Drug Development"),
        
        # Regulatory
        ("REG_001", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "https://www.fda.gov/newsroom/press-announcements/rss.xml", "https://www.fda.gov/sitemap.xml", "https://news.google.com/rss/search?q=site%3Afda.gov+approval&hl=en-US", "https://www.fda.gov/news-events/press-announcements", "https://api.fda.gov/drug/event.json", "Approach 1: Native RSS", "● 200 OK (Healthy)", "US Drug & Biologic Approvals"),
        ("REG_002", "EMA News & Press Releases", "Regulatory Agency", "ema.europa.eu", "https://www.ema.europa.eu/en/rss.xml", "https://www.ema.europa.eu/en/sitemap.xml", "https://news.google.com/rss/search?q=site%3Aema.europa.eu&hl=en-US", "https://www.ema.europa.eu/en/news", "None (Direct Agency)", "Approach 1: Native RSS", "● 200 OK (Healthy)", "EU CHMP Opinions & Approvals"),
        ("REG_003", "ClinicalTrials.gov Protocol Engine", "Clinical Registry", "clinicaltrials.gov", "https://clinicaltrials.gov/api/v2/studies", "https://clinicaltrials.gov/sitemap.xml", "https://news.google.com/rss/search?q=site%3Aclinicaltrials.gov&hl=en-US", "https://clinicaltrials.gov/study/", "https://clinicaltrials.gov/api/v2/stats/size", "Approach 1: REST API v2", "● 200 OK (Healthy)", "Global Interventional Trials"),

        # Companies with Native RSS
        ("COMP_001", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "https://www.novartis.com/news/media-releases/feed", "https://www.novartis.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Anovartis.com+press&hl=en-US", "https://www.novartis.com/news/media-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Cardiovascular, Oncology, Immunology"),
        ("COMP_002", "Novo Nordisk", "Corporate Drugmaker", "novonordisk.com (NVO)", "https://www.novonordisk.com/content/nncorp/global/en/news-and-media/news-and-ir-materials/news-details.rss", "https://www.novonordisk.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Anovonordisk.com+press&hl=en-US", "https://www.novonordisk.com/news-and-media.html", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Diabetes, Obesity, Rare Blood"),
        ("COMP_003", "AstraZeneca", "Corporate Drugmaker", "astrazeneca.com (AZN)", "https://www.astrazeneca.com/media-centre/press-releases.rss", "https://www.astrazeneca.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Aastrazeneca.com+press&hl=en-US", "https://www.astrazeneca.com/media-centre/press-releases.html", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 1: Native RSS", "● 200 OK (Healthy)", "Oncology, Rare Diseases, Vaccines"),
        ("COMP_004", "Anavex Life Sciences", "Corporate Drugmaker", "anavex.com (AVXL)", "https://anavex.com/feed/", "https://anavex.com/sitemap_index.xml", "https://news.google.com/rss/search?q=site%3Aanavex.com+press&hl=en-US", "https://anavex.com/press-releases/", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 1: Native RSS", "● 200 OK (Healthy)", "CNS, Alzheimer's, Rett Syndrome"),

        # Companies with XML Sitemap / Google News Fallbacks (No Public RSS)
        ("COMP_005", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "https://www.pfizer.com/rss.xml (Dead 404)", "https://www.pfizer.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Apfizer.com+press&hl=en-US", "https://www.pfizer.com/news/press-release/press-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 2: XML Sitemap", "● 200 OK (Sitemap Active)", "Oncology, Vaccines, Rare Disease"),
        ("COMP_006", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "https://investor.lilly.com/rss/news-releases.xml (Akamai 403)", "https://www.lilly.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Alilly.com+press&hl=en-US", "https://investor.lilly.com/news-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 3: Google News", "● Fallback Active (WAF Blocked)", "Incretins, Obesity, Alzheimer's"),
        ("COMP_007", "Alnylam Pharmaceuticals", "Corporate Drugmaker", "alnylam.com (ALNY)", "None (No RSS Exposed)", "https://www.alnylam.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Aalnylam.com+press&hl=en-US", "https://www.alnylam.com/media/press-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 3: Google News", "● Fallback Active (Zero RSS)", "RNAi Therapeutics, TTR Amyloidosis"),
        ("COMP_008", "Biogen", "Corporate Drugmaker", "biogen.com (BIIB)", "https://investors.biogen.com/rss/news-releases.xml (WAF Blocked)", "https://www.biogen.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Abiogen.com+press&hl=en-US", "https://investors.biogen.com/news-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 3: Google News", "● Fallback Active (WAF Blocked)", "Neurology, ALS, Alzheimer's"),
        ("COMP_009", "Vertex Pharmaceuticals", "Corporate Drugmaker", "vrtx.com (VRTX)", "None (No RSS Exposed)", "https://www.vrtx.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Avrtx.com+press&hl=en-US", "https://news.vrtx.com/press-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 3: Google News", "● Fallback Active (Zero RSS)", "Cystic Fibrosis, Pain, Sickle Cell"),
        ("COMP_010", "Regeneron Pharmaceuticals", "Corporate Drugmaker", "regeneron.com (REGN)", "https://investor.regeneron.com/rss/news-releases.xml (Akamai 403)", "https://www.regeneron.com/sitemap.xml", "https://news.google.com/rss/search?q=site%3Aregeneron.com+press&hl=en-US", "https://investor.regeneron.com/press-releases", "https://www.sec.gov/edgar/searchedgar/companysearch", "Approach 3: Google News", "● Fallback Active (WAF Blocked)", "Ophthalmology, Immunology, Oncology")
    ]

    for r_idx, row in enumerate(sample_wide_data, start=2):
        ws_wide.cell(row=r_idx, column=1, value=row[0]).font = font_bold
        ws_wide.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
        ws_wide.cell(row=r_idx, column=2, value=row[1]).font = font_bold
        ws_wide.cell(row=r_idx, column=3, value=row[2]).font = font_data
        ws_wide.cell(row=r_idx, column=4, value=row[3]).font = font_bold

        # URLs
        for c_i, url_val in enumerate(row[4:9], start=5):
            cell = ws_wide.cell(row=r_idx, column=c_i, value=url_val)
            if url_val.startswith("http"):
                cell.font = font_link
            else:
                cell.font = font_code

        # Active Route
        ws_wide.cell(row=r_idx, column=10, value=row[9]).font = font_bold

        # Status badge
        cell_stat = ws_wide.cell(row=r_idx, column=11, value=row[10])
        cell_stat.alignment = Alignment(horizontal="center")
        if "Healthy" in row[10] or "Sitemap Active" in row[10]:
            cell_stat.font = font_green
            cell_stat.fill = fill_green
        elif "Fallback" in row[10]:
            cell_stat.font = font_amber
            cell_stat.fill = fill_amber
        else:
            cell_stat.font = font_red
            cell_stat.fill = fill_red

        ws_wide.cell(row=r_idx, column=12, value=row[11]).font = font_data

        for col_c in range(1, 13):
            ws_wide.cell(row=r_idx, column=col_c).border = thin_border
        ws_wide.row_dimensions[r_idx].height = 24

    ws_wide.freeze_panes = "A2"
    ws_wide.auto_filter.ref = f"A1:L{len(sample_wide_data)+1}"

    # -------------------------------------------------------------------------
    # 2. VERSION 2: Preview_02_Combined_Long (Each approach in separate rows)
    # -------------------------------------------------------------------------
    TAB_LONG = "Preview_02_Combined_Long"
    if TAB_LONG in wb.sheetnames:
        del wb[TAB_LONG]

    ws_long = wb.create_sheet(title=TAB_LONG)
    ws_long.views.sheetView[0].showGridLines = True

    long_headers = [
        ("Entity ID", 14, fill_navy),
        ("Organization / Entity Name", 30, fill_navy),
        ("Classification", 22, fill_navy),
        ("Domain & Ticker", 24, fill_navy),
        ("Cascade Priority", 18, fill_blue),
        ("Approach / Ingestion Vector", 32, fill_blue),
        ("Endpoint URL / Query Definition", 55, fill_blue),
        ("Execution Status", 22, fill_purple),
        ("Last Verified Response", 28, fill_purple),
        ("Technical Notes & Bypass Configuration", 42, fill_navy),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(long_headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_long.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_long.column_dimensions[col_letter].width = width
    ws_long.row_dimensions[1].height = 28

    sample_long_data = [
        # NOVARTIS (5 Rows)
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "Level 1 (Primary)", "1. Native RSS Feed", "https://www.novartis.com/news/media-releases/feed", "ACTIVE (Primary)", "HTTP 200 OK (22 releases)", "Native RSS discovered & verified. Primary stream."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "Level 2 (Secondary)", "2. XML Sitemap Index", "https://www.novartis.com/sitemap.xml", "STANDBY (Backup)", "HTTP 200 OK (<lastmod> active)", "Capped at 20 sub-sitemaps. Standby fallback."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "Level 3 (Fallback)", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Anovartis.com+press&hl=en-US", "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Bypasses corporate CDN. Active backup."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "Level 4 (Direct HTML)", "4. HTML Newsroom Scraper", "https://www.novartis.com/news/media-releases", "STANDBY (Backup)", "HTTP 200 OK (HTML DOM)", "Deep lead paragraph scraping for Tier 1 alerts."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "novartis.com (NVS)", "Level 5 (Regulatory)", "5. SEC EDGAR / Filings", "https://www.sec.gov/edgar/searchedgar/companysearch", "STANDBY (Backup)", "HTTP 200 OK (EDGAR API)", "Requires SEC_UA header to prevent 403."),

        # PFIZER (5 Rows - Demonstrating Sitemap Primary Fallback)
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "Level 1 (Primary)", "1. Native RSS Feed", "https://www.pfizer.com/rss.xml", "DEGRADED (404 Dead)", "HTTP 404 Not Found (0 items)", "Feed removed during CMS migration. Auto-failed."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "Level 2 (Secondary)", "2. XML Sitemap Index", "https://www.pfizer.com/sitemap.xml", "ACTIVE (Fallback)", "HTTP 200 OK (8 fresh PRs)", "Active primary route. Parses /news/press-release/<lastmod>."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "Level 3 (Fallback)", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Apfizer.com+press&hl=en-US", "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Standby mirror if sitemap fails."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "Level 4 (Direct HTML)", "4. HTML Newsroom Scraper", "https://www.pfizer.com/news/press-release/press-releases", "STANDBY (Backup)", "HTTP 200 OK (HTML DOM)", "Deep full-text extraction."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "pfizer.com (PFE)", "Level 5 (Regulatory)", "5. SEC EDGAR / Filings", "https://www.sec.gov/edgar/searchedgar/companysearch", "STANDBY (Backup)", "HTTP 200 OK (EDGAR API)", "SEC Form 8-K regulatory stream."),

        # ELI LILLY (5 Rows - Demonstrating WAF Block & GNews Fallback)
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "Level 1 (Primary)", "1. Native RSS Feed", "https://investor.lilly.com/rss/news-releases.xml", "BLOCKED (Akamai WAF)", "HTTP 403 Forbidden", "Akamai bot challenge blocks cloud runner IP."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "Level 2 (Secondary)", "2. XML Sitemap Index", "https://www.lilly.com/sitemap.xml", "BLOCKED (WAF Challenge)", "HTTP 403 Forbidden", "Sitemap endpoint behind same WAF."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "Level 3 (Fallback)", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Alilly.com+press&hl=en-US", "ACTIVE (Fallback)", "HTTP 200 OK (14 releases)", "Active primary route. 100% reliable bypass."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "Level 4 (Direct HTML)", "4. HTML Newsroom Scraper", "https://investor.lilly.com/news-releases", "STANDBY (Backup)", "HTTP 403 (Requires Playwright)", "Headless browser fallback for deep body text."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "lilly.com (LLY)", "Level 5 (Regulatory)", "5. SEC EDGAR / Filings", "https://www.sec.gov/edgar/searchedgar/companysearch", "STANDBY (Backup)", "HTTP 200 OK (EDGAR API)", "SEC 8-K material event stream."),

        # STAT NEWS (5 Rows - Trade Media Example)
        ("PUB_STAT", "STAT News", "Industry Trade Media", "statnews.com", "Level 1 (Primary)", "1. Native RSS Feed", "https://www.statnews.com/feed/", "ACTIVE (Primary)", "HTTP 200 OK (30 stories)", "High-velocity media stream. 30 worker threads."),
        ("PUB_STAT", "STAT News", "Industry Trade Media", "statnews.com", "Level 2 (Secondary)", "2. XML Sitemap Index", "https://www.statnews.com/sitemap.xml", "STANDBY (Backup)", "HTTP 200 OK (News Sitemap)", "Standby news sitemap."),
        ("PUB_STAT", "STAT News", "Industry Trade Media", "statnews.com", "Level 3 (Fallback)", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Astatnews.com&hl=en-US", "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Standby mirror."),
        ("PUB_STAT", "STAT News", "Industry Trade Media", "statnews.com", "Level 4 (Direct HTML)", "4. HTML Newsroom Scraper", "https://www.statnews.com/latest/", "STANDBY (Backup)", "HTTP 200 OK (HTML DOM)", "Paywalled excerpts extracted via meta tags."),
        ("PUB_STAT", "STAT News", "Industry Trade Media", "statnews.com", "Level 5 (Regulatory)", "5. Editorial Archive", "https://www.statnews.com/pharmalot/", "STANDBY (Backup)", "HTTP 200 OK (Pharmalot)", "Specialized FDA/Pharmalot column."),

        # FDA (5 Rows - Regulatory Agency Example)
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "Level 1 (Primary)", "1. Native RSS Feed", "https://www.fda.gov/newsroom/press-announcements/rss.xml", "ACTIVE (Primary)", "HTTP 200 OK (12 approvals)", "Official FDA regulatory press stream."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "Level 2 (Secondary)", "2. XML Sitemap Index", "https://www.fda.gov/sitemap.xml", "STANDBY (Backup)", "HTTP 200 OK (Gov Sitemap)", "Standby gov sitemap."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "Level 3 (Fallback)", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Afda.gov+approval&hl=en-US", "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Standby broad approval query."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "Level 4 (Direct HTML)", "4. HTML Newsroom Scraper", "https://www.fda.gov/news-events/press-announcements", "STANDBY (Backup)", "HTTP 200 OK (HTML DOM)", "Scrapes complete approval decision letters."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "fda.gov", "Level 5 (Regulatory)", "5. openFDA Drug API", "https://api.fda.gov/drug/event.json", "STANDBY (Backup)", "HTTP 200 OK (REST API)", "Structured adverse event & label JSON API.")
    ]

    current_entity = ""
    is_alt = False

    for r_idx, row in enumerate(sample_long_data, start=2):
        # Alternate background fill per entity block
        if row[1] != current_entity:
            current_entity = row[1]
            is_alt = not is_alt

        row_fill = fill_ice_blue if is_alt else PatternFill(fill_type=None)

        ws_long.cell(row=r_idx, column=1, value=row[0]).font = font_bold
        ws_long.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
        ws_long.cell(row=r_idx, column=2, value=row[1]).font = font_bold
        ws_long.cell(row=r_idx, column=3, value=row[2]).font = font_data
        ws_long.cell(row=r_idx, column=4, value=row[3]).font = font_bold

        ws_long.cell(row=r_idx, column=5, value=row[4]).font = font_bold
        ws_long.cell(row=r_idx, column=5).alignment = Alignment(horizontal="center")
        ws_long.cell(row=r_idx, column=6, value=row[5]).font = font_bold

        cell_url = ws_long.cell(row=r_idx, column=7, value=row[6])
        if str(row[6]).startswith("http"):
            cell_url.font = font_link
        else:
            cell_url.font = font_code

        # Execution status badge
        cell_st = ws_long.cell(row=r_idx, column=8, value=row[7])
        cell_st.alignment = Alignment(horizontal="center")
        if "ACTIVE (Primary)" in row[7]:
            cell_st.font = font_green
            cell_st.fill = fill_green
        elif "ACTIVE (Fallback)" in row[7]:
            cell_st.font = font_blue_badge
            cell_st.fill = fill_blue_badge
        elif "STANDBY" in row[7]:
            cell_st.font = font_amber
            cell_st.fill = fill_amber
        else:
            cell_st.font = font_red
            cell_st.fill = fill_red

        ws_long.cell(row=r_idx, column=9, value=row[8]).font = font_code
        ws_long.cell(row=r_idx, column=10, value=row[9]).font = font_data

        for col_c in range(1, 11):
            cell_c = ws_long.cell(row=r_idx, column=col_c)
            cell_c.border = thin_border
            if cell_c.fill.fill_type is None and row_fill.fill_type is not None:
                cell_c.fill = row_fill

        ws_long.row_dimensions[r_idx].height = 22

    ws_long.freeze_panes = "A2"
    ws_long.auto_filter.ref = f"A1:J{len(sample_long_data)+1}"

    # Save
    wb.save(XLSX_PATH)
    print(f"✅ Successfully added '{TAB_WIDE}' and '{TAB_LONG}' to {XLSX_PATH}!")

if __name__ == "__main__":
    build_combined_previews()
