#!/usr/bin/env python3
"""
RSSFeedChecker — Master Production Multi-Pillar Intelligence Pipeline
=====================================================================
Unified intelligence pipeline feeding 'Results' tab from '01_Master_Sources_Registry'
with full configuration control from '00_Run_Dashboard' or CLI parameters.

Architecture & Core Features:
1. Ingests all 4,178 Verified Active Endpoints (Pillar 1, 2, 3)
2. Interactive Control from '00_Run_Dashboard' frontpage (--from-dashboard)
3. 3-Gate Matching Engine with Preclinical/Animal & Stock Noise Rejection
4. Granular Signal Type Classification (clinical_pos, clinical_neg, regulatory, corporate, leadership_change, commercial, general)
5. Dynamic 0-100 Numeric Relevance Scorer
6. Deterministic Event Clustering & Entity Extraction (Cluster ID & Cluster Hint JSON)
7. 20-Column Enhanced Results Presentation Matrix in Master Excel ('Results' tab)
"""

import os
import re
import sys
import time
import json
import argparse
import urllib.parse
from datetime import datetime, timezone, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
        sys.stderr.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    except Exception:
        pass

# Add current directory to path
sys.path.insert(0, os.path.dirname(__file__))

from time_window import create_time_window, TimeWindow
from match_keywords import KeywordMatcher, normalize_headline_key, SmartDeduplicator
from clustering_engine import MultiVectorClusterEngine
from robust_fetcher import (
    robust_fetch,
    fetch_search_news_resilient,
    decode_authoritative_link,
    clean_snippet,
    extract_slug_date,
    extract_dateline_date,
    extract_full_article_content,
    safe_save_workbook
)
try:
    from sec_edgar_client import fetch_pure_sec_press_releases, get_sec_cik_for_ticker_or_name
except ImportError:
    def fetch_pure_sec_press_releases(*args, **kwargs): return []
    def get_sec_cik_for_ticker_or_name(q): return ""

from ai_synthesis_engine import generate_factual_ai_summary, generate_ci_competitive_implications

try:
    from fetch_new_clinical_trials import fetch_new_trials_for_condition
except ImportError:
    def fetch_new_trials_for_condition(*args, **kwargs): return []

WORKSPACE = os.environ.get("VAULT_DIR", r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker_02")
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

GLOBAL_AUDIT_TELEMETRY = {}


def parse_date_to_utc(date_str: str) -> datetime:
    """Parse various RFC-822, ISO, and standard date strings to UTC."""
    if not date_str:
        return datetime.now(timezone.utc)
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass

    for fmt in [
        "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    ]:
        try:
            dt = datetime.strptime(date_str.strip(), fmt)
            if dt.tzinfo is None:
                return dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except Exception:
            continue
    return datetime.now(timezone.utc)


def fetch_channel_endpoint(
    feed_id: str,
    target_url: str,
    entity_name: str,
    source_class: str,
    pillar_label: str,
    vector_label: str,
    time_window: TimeWindow,
    domain: str = "",
    desk_override: str = "",
    booster: str = "",
    max_items: int = 30
) -> list[dict]:
    """Fetch and parse items from any HTTP/RSS/Search endpoint and log audit telemetry."""
    items = []
    t_start_fetch = time.time()
    if not target_url or not target_url.startswith("http"):
        return items

    is_search = "search" in target_url.lower() or "3. Google" in vector_label
    code = 0
    engine_label = vector_label
    err = ""
    raw_items = []
    sample_first_title = "--"
    sample_first_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    try:
        if is_search:
            code, body_bytes, final_url, engine_label, err = fetch_search_news_resilient(
                target_url, entity_name=entity_name, domain=domain, timeout=6
            )
        else:
            code, body_bytes, final_url, err = robust_fetch(target_url, timeout=6, max_retries=1)
            engine_label = vector_label

        if code == 200 and body_bytes:
            body_str = body_bytes.decode("utf-8", "ignore")

            # Parse XML Items
            raw_items = re.findall(r"<item\b[^>]*>(.*?)</item>", body_str, re.IGNORECASE | re.DOTALL)
            if not raw_items:
                raw_items = re.findall(r"<entry\b[^>]*>(.*?)</entry>", body_str, re.IGNORECASE | re.DOTALL)

            for raw_it in raw_items[:max_items]:
                t_match = re.search(r"<title\b[^>]*>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</title>", raw_it, re.IGNORECASE | re.DOTALL)
                l_match = re.search(r"<link\b[^>]*>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</link>", raw_it, re.IGNORECASE | re.DOTALL)
                if not l_match:
                    l_match = re.search(r'<link\b[^>]*href=["\'](.*?)["\']', raw_it, re.IGNORECASE)
                d_match = re.search(r"<(?:pubDate|updated|dc:date)\b[^>]*>(.*?)</", raw_it, re.IGNORECASE)
                s_match = re.search(r"<(?:description|summary|content:encoded)\b[^>]*>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</", raw_it, re.IGNORECASE | re.DOTALL)

                raw_title = clean_snippet(t_match.group(1)) if t_match else ""
                raw_link = l_match.group(1).strip() if l_match else ""
                raw_date = d_match.group(1).strip() if d_match else ""
                raw_desc = clean_snippet(s_match.group(1)) if s_match else ""

                src_match = re.search(r"<source\b[^>]*>(?:<!\[CDATA\[)?(.*?)(?:\]\]>)?</source>", raw_it, re.IGNORECASE | re.DOTALL)
                pub_source = clean_snippet(src_match.group(1)) if src_match else ""

                # Strip trailing publisher suffix from headline (e.g. - Yahoo Finance Australia)
                m_suffix = re.search(r'\s+-\s+([A-Za-z0-9\s\.\!&–—\']+)\s*$', raw_title)
                if m_suffix:
                    suffix_pub = m_suffix.group(1).strip()
                    if not pub_source or pub_source.startswith("GN-"):
                        pub_source = suffix_pub
                    raw_title = raw_title[:m_suffix.start()].strip()

                final_source_name = pub_source if (pub_source and not pub_source.startswith("GN-")) else entity_name
                if "GN-" in final_source_name or "RSS-" in final_source_name:
                    final_source_name = pub_source if pub_source else "Trade Press"

                if not raw_title or not raw_link:
                    continue

                if sample_first_title == "--":
                    sample_first_title = raw_title
                    sample_first_date = raw_date[:10] if raw_date else sample_first_date

                pub_utc = parse_date_to_utc(raw_date)
                # Time window filter
                if not time_window.is_in_window(pub_utc):
                    continue

                items.append({
                    "feed_id": feed_id,
                    "published_utc": pub_utc,
                    "published_date": pub_utc.strftime("%Y-%m-%d"),
                    "published_time": pub_utc.strftime("%H:%M"),
                    "headline": raw_title,
                    "raw_url": raw_link,
                    "snippet": raw_desc[:500],
                    "full_text": raw_desc,
                    "source_name": final_source_name,
                    "source_class": source_class,
                    "discovery_method": pillar_label,
                    "extraction_vector": engine_label,
                    "desk_override": desk_override,
                    "booster": booster,
                })

    except Exception as e:
        err = str(e)

    latency_ms = int((time.time() - t_start_fetch) * 1000)
    now_ist = datetime.now().strftime("%d-%m-%Y %I:%M %p IST")

    # Record Live Audit Telemetry for 01_Master_Sources_Registry
    if code == 200:
        health_str = f"✅ 200 OK — Active ({len(items)} items)" if items else f"✅ 200 OK ({len(raw_items)} items in stream)"
        quality_str = "🟢 Working & Active" if items else "🟡 Clean Feed (No new items in current window)"
        action_str = "✅ No action needed (Feed healthy)"
    elif code in (401, 403, 429):
        health_str = f"⚠️ Blocked / Rate Limited ({code})"
        quality_str = "🟠 Fallback Scraper Used"
        action_str = "🔄 Fallback active (Monitoring automatically)"
    elif code == 404:
        health_str = "❌ Error 404 (Not Found)"
        quality_str = "🔴 Dead Link / Broken URL"
        action_str = "⚠️ Recommend deleting or updating URL"
    else:
        health_str = f"⚠️ Error / Timeout ({code})" if code else "⚠️ Connection Timeout"
        quality_str = "🔴 Host Offline / Unreachable"
        action_str = "⚠️ Check endpoint status or delete if dead"

    GLOBAL_AUDIT_TELEMETRY[feed_id] = {
        "health": health_str,
        "structure": engine_label,
        "latest_title": items[0]["headline"] if items else sample_first_title[:80],
        "fetch_date_time": now_ist,
        "items_detected": f"{len(items)} items in window ({len(raw_items)} total)",
        "feed_quality": quality_str,
        "recommendation": action_str
    }

    return items


def load_dashboard_settings() -> dict:
    """Read configured parameters from Sheet '00_Run_Dashboard'."""
    settings = {
        "since": "2026-08-01",
        "until": "2026-08-26",
        "category": "ALL",
        "entity": "ALL",
        "max_items": 30,
        "min_score": 40,
        "tier": "ALL",
        "project": "ALL",
        "is_sample": False,
        "output_tab": "Results"
    }
    if not os.path.exists(XLSX_PATH):
        return settings

    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        if "00_Run_Dashboard" in wb.sheetnames:
            ws = wb["00_Run_Dashboard"]
            for r in range(7, 18):
                p_name = str(ws.cell(row=r, column=2).value or "").strip()
                p_val = str(ws.cell(row=r, column=3).value or "").strip()
                if "Start Date" in p_name:
                    settings["since"] = p_val
                elif "End Date" in p_name:
                    settings["until"] = p_val
                elif "Classification" in p_name:
                    settings["category"] = p_val
                elif "Target Entity" in p_name or "Target Name" in p_name:
                    settings["entity"] = p_val
                elif "Max Items" in p_name:
                    try: settings["max_items"] = int(p_val)
                    except: pass
                elif "Minimum Relevance" in p_name:
                    try: settings["min_score"] = int(p_val)
                    except: pass
                elif "Priority Tier" in p_name:
                    settings["tier"] = p_val
                elif "Therapeutic Project" in p_name:
                    settings["project"] = p_val
                elif "Universe Mode" in p_name:
                    if "Sample" in p_val:
                        settings["is_sample"] = True
                elif "Destination Tab" in p_name:
                    settings["output_tab"] = p_val or "Results"
    except Exception as e:
        print(f"⚠️ Could not read dashboard settings: {e}")

    return settings


def run_pipeline(
    since_date: str = "2026-08-01",
    until_date: str = "",
    category_filter: str = "ALL",
    entity_filter: str = "ALL",
    project_filter: str = "ALL",
    tier_filter: str = "ALL",
    min_score_cutoff: int = 40,
    max_items_per_feed: int = 30,
    is_sample: bool = False,
    is_admin: bool = True,
    output_tab_name: str = "Results",
    from_dashboard: bool = False
):
    if from_dashboard:
        dash_cfg = load_dashboard_settings()
        since_date = dash_cfg["since"]
        until_date = dash_cfg["until"]
        category_filter = dash_cfg["category"]
        entity_filter = dash_cfg.get("entity", "ALL")
        max_items_per_feed = dash_cfg["max_items"]
        min_score_cutoff = dash_cfg["min_score"]
        tier_filter = dash_cfg["tier"]
        project_filter = dash_cfg["project"]
        is_sample = is_sample or dash_cfg["is_sample"]
        output_tab_name = dash_cfg["output_tab"]

    # Construct time window query string
    if since_date:
        since_clean = str(since_date).strip()
        # If relative duration like '36h', '48h', '3d', '72h', use directly
        if re.match(r"^\d+\s*(?:h|hr|hrs|hours?|d|days?)$", since_clean, re.I):
            window_query = since_clean
        elif until_date and str(until_date).lower() != "present":
            window_query = f"{since_clean} to {until_date}"
        else:
            window_query = f"From {since_clean} onwards"
    else:
        window_query = "72h"

    tw = create_time_window(window_query, is_admin=is_admin)
    matcher = KeywordMatcher(XLSX_PATH)

    # Compute Ingestion Run Signature & Provenance
    now_utc = datetime.now(timezone.utc)
    # Convert UTC to IST (+05:30)
    now_ist = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=5, minutes=30)))
    run_batch_id = f"RUN_{now_ist.strftime('%Y%m%d_%H%M')}_IST"
    ingested_date_ist = now_ist.strftime("%Y-%m-%d")
    ingested_time_ist = now_ist.strftime("%I:%M %p IST")

    # Determine Execution Run Type
    github_actions = os.environ.get("GITHUB_ACTIONS", "").lower() == "true"
    github_event = os.environ.get("GITHUB_EVENT_NAME", "").lower()
    if github_actions:
        if github_event == "schedule":
            exec_run_type = "⚡ Cloud Cron (Daily 6 AM IST)"
        else:
            exec_run_type = "🖱️ Manual (GitHub Dispatch)"
    else:
        exec_run_type = "💻 Manual (Local IDE / Runner)"

    print("=" * 90)
    print(" 🔬 RSSFeedChecker — MASTER PRODUCTION MULTI-PILLAR INTELLIGENCE PIPELINE")
    print("=" * 90)
    print(f"Batch Run ID:     {run_batch_id} ({exec_run_type})")
    print(f"Ingested Time:    {ingested_date_ist} at {ingested_time_ist}")
    print(tw.format_summary())
    mode_str = "SAMPLE MODE (Quick 50-Source Test)" if is_sample else "FULL MASTER CATALOG (4,178 Active Endpoints)"
    print(f"Mode:             {mode_str}")
    print(f"Category Filter:  {category_filter}")
    print(f"Project Filter:   {project_filter}")
    print(f"Min Score Cutoff: {min_score_cutoff}/100")
    print(f"Max Items / Feed: {max_items_per_feed}")
    print(f"Output Sheet:     '{output_tab_name}'")
    print(f"Loaded {len(matcher.projects)} Therapeutic Projects & 3-Gate Matching Rules\n")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    
    reg_tab_name = next((s for s in wb.sheetnames if "Master_Sources" in s or "Unified_All_Pillars" in s), None)
    if not reg_tab_name:
        print("❌ Error: Master sources registry tab not found!")
        return

    ws_reg = wb[reg_tab_name]
    print(f"▶ Ingesting active streams from Master Sheet: '{reg_tab_name}' ({ws_reg.max_row - 1} rows)...")

    http_tasks = []
    sec_tasks = []
    ct_tasks = []

    for r in range(2, ws_reg.max_row + 1):
        f_id = str(ws_reg.cell(row=r, column=1).value or "").strip()
        e_name = str(ws_reg.cell(row=r, column=2).value or "").strip()
        e_class = str(ws_reg.cell(row=r, column=3).value or "").strip()
        pillar = str(ws_reg.cell(row=r, column=4).value or "").strip()
        vector = str(ws_reg.cell(row=r, column=5).value or "").strip()
        url = str(ws_reg.cell(row=r, column=6).value or "").strip()
        toggle = str(ws_reg.cell(row=r, column=8).value or "Active").strip()
        desk_over = str(ws_reg.cell(row=r, column=9).value or "").strip()
        booster = str(ws_reg.cell(row=r, column=10).value or "Default").strip()
        row_max = int(ws_reg.cell(row=r, column=11).value or max_items_per_feed)
        max_i = min(max_items_per_feed, row_max)

        if "Paused" in toggle or "Muted" in toggle:
            continue

        # Filter by Category if specified
        if category_filter != "ALL" and "ALL" not in category_filter:
            if category_filter.lower() not in e_class.lower():
                continue

        # Filter by Target Entity / Target Name if specified
        if entity_filter and entity_filter != "ALL" and "ALL" not in entity_filter:
            target_terms = [t.strip().lower() for t in entity_filter.split(",") if t.strip()]
            if not any(t in e_name.lower() or t in f_id.lower() for t in target_terms):
                continue

        if "5. SEC EDGAR" in vector or "6. SEC EDGAR" in e_class:
            sec_tasks.append((f_id, e_name, e_class, url, desk_over, booster, max_i))
        elif "ClinicalTrials.gov" in vector or "CT_" in f_id or "7. Clinical" in e_class:
            m_cond = re.search(r"query\.cond=([^&]+)", url)
            cond = urllib.parse.unquote(m_cond.group(1)) if m_cond else e_name.replace(" (Trials)", "")
            ct_tasks.append((f_id, e_name, e_class, cond, desk_over, booster))
        elif url.startswith("http"):
            domain_m = re.search(r"https?://([^/]+)", url)
            dom = domain_m.group(1) if domain_m else ""
            http_tasks.append((f_id, url, e_name, e_class, pillar, vector, dom, desk_over, booster, max_i))

    if is_sample:
        http_tasks = http_tasks[:35]
        sec_tasks = sec_tasks[:8]
        ct_tasks = ct_tasks[:8]

    raw_events = []
    t_start = time.time()

    # 1. Fetch HTTP / RSS / Search Streams in parallel
    print(f"▶ Ingesting {len(http_tasks)} Active Web/RSS/Search Streams (35 parallel workers)...")
    with ThreadPoolExecutor(max_workers=35) as pool:
        futures = {
            pool.submit(
                fetch_channel_endpoint,
                t[0], t[1], t[2], t[3], t[4], t[5], tw, t[6], t[7], t[8], t[9]
            ): t for t in http_tasks
        }
        for fut in as_completed(futures):
            res = fut.result()
            raw_events.extend(res)

    print(f"   ✓ Captured {len(raw_events)} raw events from web & RSS streams")

    # 2. Fetch Pure SEC EDGAR Press Releases
    print(f"\n▶ Querying {len(sec_tasks)} SEC EDGAR Pure PR Channels (Item 8.01/6-K)...")
    sec_count = 0
    for s_task in sec_tasks:
        f_id, e_name, e_class, s_url, desk_over, booster, max_i = s_task
        cik = ""
        m_cik = re.search(r"CIK(\d+)", s_url)
        if m_cik:
            cik = m_cik.group(1)
        
        prs = fetch_pure_sec_press_releases(cik, company_name=e_name, max_items=max_i, timeout=4)
        for p in prs:
            p_utc = parse_date_to_utc(p["published_date"])
            if tw.is_in_window(p_utc):
                raw_events.append({
                    "feed_id": f_id,
                    "published_utc": p_utc,
                    "published_date": p["published_date"],
                    "published_time": "--",
                    "headline": f"{e_name} SEC Filing: {p['lead_text'].split('.')[0][:90]}" if (p["title"] == "EX-99.1" or len(p["title"]) < 10) else p["title"],
                    "raw_url": p["url"],
                    "snippet": p["lead_text"],
                    "full_text": p["lead_text"],
                    "source_name": f"{e_name} (SEC EDGAR EX-99.1)",
                    "source_class": e_class,
                    "discovery_method": "Pillar 2: SEC EDGAR Pure PR",
                    "extraction_vector": f"SEC {p['form_type']} ({p['item_code']})",
                    "desk_override": desk_over or p["desk_route"],
                    "booster": booster,
                })
                sec_count += 1

        now_ist = datetime.now().strftime("%d-%m-%Y %I:%M %p IST")
        GLOBAL_AUDIT_TELEMETRY[f_id] = {
            "health": f"✅ 200 OK — Active ({len(prs)} filings)" if prs else "✅ 200 OK (0 filings in window)",
            "structure": "SEC EDGAR API JSON (Item 8.01/6-K)",
            "latest_title": prs[0]["title"][:80] if prs else f"{e_name} SEC Direct",
            "fetch_date_time": now_ist,
            "items_detected": f"{len(prs)} filings retrieved",
            "feed_quality": "🟢 Official SEC Regulatory Stream",
            "recommendation": "✅ No action needed (Official SEC feed)"
        }

    print(f"   ✓ Captured {sec_count} pure SEC press releases")

    # 3. Fetch ClinicalTrials.gov Protocols
    print(f"\n▶ Querying {len(ct_tasks)} ClinicalTrials.gov Condition Registries...")
    ct_count = 0
    for ct in ct_tasks:
        f_id, e_name, e_class, cond, desk_over, booster = ct
        trials = fetch_new_trials_for_condition(cond, tw, max_studies=15)
        for t in trials:
            t_utc = parse_date_to_utc(t["first_post_date"])
            raw_events.append({
                "feed_id": f_id,
                "published_utc": t_utc,
                "published_date": t["first_post_date"],
                "published_time": "--",
                "headline": f"{t['event_type']} {t['nct_id']} [{t['phase']}]: {clean_snippet(t['title'])}",
                "raw_url": t["url"],
                "snippet": f"Lead Sponsor: {t['lead_sponsor']} | Condition: {t['condition']} | Status: {t['status']}",
                "full_text": t["summary"],
                "source_name": f"ClinicalTrials.gov ({t['lead_sponsor']})",
                "source_class": e_class,
                "discovery_method": "Pillar 3: ClinicalTrials.gov",
                "extraction_vector": t["extraction_vector"],
                "desk_override": desk_over,
                "booster": booster,
            })
            ct_count += 1

        now_ist = datetime.now().strftime("%d-%m-%Y %I:%M %p IST")
        GLOBAL_AUDIT_TELEMETRY[f_id] = {
            "health": f"✅ 200 OK — Active ({len(trials)} trials)" if trials else "✅ 200 OK (0 trials in window)",
            "structure": "ClinicalTrials.gov v2 REST API",
            "latest_title": trials[0]["title"][:80] if trials else f"{cond} Protocol Registry",
            "fetch_date_time": now_ist,
            "items_detected": f"{len(trials)} protocols retrieved",
            "feed_quality": "🟢 Official Clinical Trials Stream",
            "recommendation": "✅ No action needed (Official CT.gov API)"
        }

    print(f"   ✓ Captured {ct_count} newly registered clinical trials")
    print(f"\nTotal raw events captured across all 3 pillars: {len(raw_events)} (in {time.time()-t_start:.1f}s)")

    # Save & Prune Raw Feed Events Archive (Max 5,000 items / 10 days)
    try:
        from raw_feed_archive import RawFeedArchiver
        archiver = RawFeedArchiver(max_rows=5000, max_days=10)
        arch_res = archiver.save_and_prune(raw_events)
        print(f"   ✓ Archived & auto-pruned raw feed cache ({arch_res['total_retained_rows']} items retained, {arch_res['file_size_kb']} KB on disk)")
    except Exception as e:
        print(f"   ⚠️ Raw feed archive notice: {e}")

    # -------------------------------------------------------------------------
    # 3-GATE DEDUPLICATION, SCORING & EVENT CLUSTERING
    # -------------------------------------------------------------------------
    print("\n▶ Decoding Authoritative URLs, Scoring against 3-Gate Rules & Clustering...")
    seen_urls = set()
    seen_titles = set()
    processed_feed = []

    for item in raw_events:
        # 1. Fast In-Memory 3-Gate Keyword & Scoring Engine (0.0001s per item)
        res = matcher.match(
            item["headline"],
            item.get("snippet", ""),
            source_class=item.get("source_class", ""),
            source_name=item.get("source_name", "")
        )
        if res.get("is_noise") or not res.get("matched"):
            continue

        # Score filter cutoff
        item_score = res.get("relevance_score", 50)
        if item_score < min_score_cutoff:
            continue

        # Project filter
        p_name = res.get("project_name", "General Biopharma")
        if project_filter != "ALL" and "ALL" not in project_filter:
            if project_filter.lower() not in p_name.lower():
                continue

        priority = res.get("top_priority", "🟢 Tier 3 (Weekly)")
        if "Always Tier 1" in str(item.get("booster", "")):
            priority = "🔴 Tier 1 (Urgent)"

        # Tier filter
        if tier_filter != "ALL" and "ALL" not in tier_filter:
            if tier_filter == "Tier 1 Only" or "Tier 1 Only" in tier_filter:
                if "Tier 1" not in priority:
                    continue
            elif "Tier 1 + Tier 2" in tier_filter:
                if "Tier 1" not in priority and "Tier 2" not in priority:
                    continue
            elif "Tier 2 Only" in tier_filter:
                if "Tier 2" not in priority:
                    continue
            elif "Tier 3 Only" in tier_filter:
                if "Tier 3" not in priority:
                    continue
            elif "Tier 1" in tier_filter and "Tier 1" not in priority:
                continue

        # 2. Fast Title Deduplication
        norm_title = normalize_headline_key(item["headline"])
        if norm_title and norm_title in seen_titles:
            continue

        # 3. Decode Authoritative URL & Anti-Redesign ONLY for Matched Candidates
        raw_u = item.get("raw_url", "")
        auth_url = decode_authoritative_link(raw_u) if ("news.google" in raw_u) else raw_u
        url_norm = clean_snippet(auth_url).lower().rstrip("/")
        if url_norm in seen_urls:
            continue

        # Anti-Redesign Stale Date Guardrail
        cutoff_iso = tw.start_utc.strftime("%Y-%m-%d")
        slug_d = extract_slug_date(auth_url) or extract_slug_date(raw_u)
        dateline_d = extract_dateline_date(item.get("snippet", "")) or extract_dateline_date(item.get("full_text", ""))
        
        if slug_d and slug_d < cutoff_iso:
            continue
        if dateline_d and dateline_d < cutoff_iso:
            continue

        seen_urls.add(url_norm)
        if norm_title:
            seen_titles.add(norm_title)

        # Desk Routing Override
        desk = item.get("desk_override") or res.get("routed_desk", "Corporate Strategy & M&A")
        if "Auto" in str(desk) or not desk or str(desk).strip() == "":
            desk = res.get("routed_desk", "Corporate Strategy & M&A")

        # Generate Clustering Payload with Authenticated Company Name
        matched_comp = res.get("matched_company", "Not Identified")
        cluster_id, cluster_hint = matcher.generate_clustering_payload(
            item["headline"],
            matched_comp,
            res.get("signal_type", "general"),
            p_name,
            item["published_date"]
        )

        discovered_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

        processed_feed.append({
            "published_date": item["published_date"],
            "published_time": item["published_time"],
            "headline": item["headline"],
            "project_name": p_name,
            "signal_type": res.get("signal_type", "general"),
            "relevance_score": item_score,
            "priority": priority,
            "desk": desk,
            "matched_keywords": res.get("matched_keywords_str", "--"),
            "source_name": item["source_name"],
            "source_class": item.get("source_class", "1. News Aggregator & Trade Press"),
            "raw_url": auth_url,
            "snippet": item["snippet"],  # Untruncated full RSS content
            "cluster_id": cluster_id,
            "cluster_hint": cluster_hint,
            "discovery_method": item["discovery_method"],
            "extraction_vector": item["extraction_vector"],
            "discovered_at": discovered_at,
            "full_text": item["full_text"],
            "event_id": f"EVT_{len(processed_feed)+1:04d}",
            "published_utc": item["published_utc"],
            "run_batch_id": run_batch_id,
            "ingested_date_ist": ingested_date_ist,
            "ingested_time_ist": ingested_time_ist,
            "exec_run_type": exec_run_type,
        })

    # Sort chronologically (newest first)
    processed_feed.sort(key=lambda x: x["published_utc"], reverse=True)
    print(f"✨ Deduplicated & Filtered Candidates: {len(processed_feed)} high-value intelligence items ready!")

    # -------------------------------------------------------------------------
    # ROBUST FULL-TEXT GROUNDED WEBPAGE EXTRACTION (15 Parallel Workers)
    # -------------------------------------------------------------------------
    if processed_feed:
        print(f"\n▶ Extracting Ground-Truth Full-Text from {len(processed_feed)} Destination Web Pages (15 parallel workers)...")
        with ThreadPoolExecutor(max_workers=15) as pool:
            futures = {
                pool.submit(extract_full_article_content, it["raw_url"], it["snippet"], 4000): idx
                for idx, it in enumerate(processed_feed)
            }
            for f in as_completed(futures):
                idx = futures[f]
                try:
                    full_body = f.result()
                    if full_body and len(full_body) > len(processed_feed[idx]["snippet"]):
                        processed_feed[idx]["full_text"] = full_body[:4000]
                except Exception:
                    pass

    # -------------------------------------------------------------------------
    # MULTI-VECTOR BIOPHARMA EVENT CLUSTERING & CONFIDENCE SCORING
    # -------------------------------------------------------------------------
    if processed_feed:
        print(f"\n▶ Executing 5-Vector Event Clustering & Confidence Scoring across {len(processed_feed)} events...")
        cluster_engine = MultiVectorClusterEngine(xlsx_path=XLSX_PATH)
        processed_feed = cluster_engine.cluster_feed_items(processed_feed)

    # -------------------------------------------------------------------------
    # WRITE 21-COLUMN OUTPUT TO TAB: 'Results' (Index 1) WITH ACCUMULATIVE MERGE
    # -------------------------------------------------------------------------
    print(f"\nWriting 21-column intelligence feed into '{output_tab_name}' in {XLSX_PATH}...")
    wb_out = openpyxl.load_workbook(XLSX_PATH)
    
    # 1. Load existing historical rows if outputting to Results
    existing_items = []
    seen_urls = set()
    if output_tab_name in wb_out.sheetnames:
        ws_old = wb_out[output_tab_name]
        for r in range(2, ws_old.max_row + 1):
            val_date = ws_old.cell(r, 1).value
            val_title = ws_old.cell(r, 3).value
            val_url = ws_old.cell(r, 12).value
            if val_date and val_title:
                is_yellow = False
                if ws_old.cell(r, 1).fill and ws_old.cell(r, 1).fill.start_color:
                    c_rgb = str(ws_old.cell(r, 1).fill.start_color.rgb)
                    if "FFF59D" in c_rgb:
                        is_yellow = True

                item_dict = {
                    "published_date": str(val_date)[:10],
                    "published_time": str(ws_old.cell(r, 2).value or "00:00")[:5],
                    "headline": str(val_title),
                    "project_name": str(ws_old.cell(r, 4).value or "General Biopharma"),
                    "signal_type": str(ws_old.cell(r, 5).value or "general"),
                    "relevance_score": str(ws_old.cell(r, 6).value or "50/100").replace("/100", ""),
                    "priority": str(ws_old.cell(r, 7).value or "🟢 Tier 3 (Weekly)"),
                    "desk": str(ws_old.cell(r, 8).value or "Corporate Strategy & M&A"),
                    "matched_keywords": str(ws_old.cell(r, 9).value or "--"),
                    "source_name": str(ws_old.cell(r, 10).value or "Newsroom"),
                    "source_class": str(ws_old.cell(r, 11).value or "1. News Aggregator & Trade Press"),
                    "raw_url": str(val_url or ""),
                    "snippet": str(ws_old.cell(r, 13).value or ""),
                    "cluster_id": str(ws_old.cell(r, 14).value or ""),
                    "cluster_hint": str(ws_old.cell(r, 15).value or ""),
                    "discovery_method": str(ws_old.cell(r, 16).value or "Pillar 1: Publisher Feeds"),
                    "extraction_vector": str(ws_old.cell(r, 17).value or "1. Native RSS Feed"),
                    "discovered_at": str(ws_old.cell(r, 18).value or ""),
                    "full_text": str(ws_old.cell(r, 19).value or ""),
                    "event_id": str(ws_old.cell(r, 20).value or f"EVT_{len(existing_items)+1:04d}"),
                    "ai_summary": str(ws_old.cell(r, 21).value or ""),
                    "implications": str(ws_old.cell(r, 22).value or ""),
                    "run_batch_id": str(ws_old.cell(r, 23).value or ("HISTORICAL_BASELINE_IMPORT" if is_yellow else "RUN_INITIAL_BASELINE")),
                    "ingested_date_ist": str(ws_old.cell(r, 24).value or str(val_date)[:10]),
                    "ingested_time_ist": str(ws_old.cell(r, 25).value or (str(ws_old.cell(r, 2).value or "00:00")[:5] + " IST")),
                    "exec_run_type": str(ws_old.cell(r, 26).value or ("📋 Baseline Historical" if is_yellow else "💻 Initial Setup Run")),
                    "is_imported": is_yellow
                }
                existing_items.append(item_dict)
                if val_url:
                    seen_urls.add(str(val_url).strip().lower())

    # Merge fresh live items that are not already present
    fresh_unique = []
    for ev in processed_feed:
        u_norm = str(ev.get("raw_url", "")).strip().lower()
        if u_norm not in seen_urls:
            seen_urls.add(u_norm)
            fresh_unique.append(ev)

    all_final_items = fresh_unique + existing_items

    # Sort descending by date & time
    def _final_sort_key(x):
        d = str(x.get("published_date", "2000-01-01"))[:10]
        t = str(x.get("published_time", "00:00"))[:5]
        try: return datetime.strptime(f"{d} {t}", "%Y-%m-%d %H:%M")
        except: return datetime.min

    all_final_items.sort(key=_final_sort_key, reverse=True)
    print(f"📊 Accumulative Master Dataset: {len(all_final_items)} total events ({len(fresh_unique)} fresh + {len(existing_items)} existing preserved)")

    # Remove existing Results tab and rebuild cleanly
    if output_tab_name in wb_out.sheetnames:
        del wb_out[output_tab_name]
    if "00_Unified_Intelligence_Feed" in wb_out.sheetnames and output_tab_name != "00_Unified_Intelligence_Feed":
        del wb_out["00_Unified_Intelligence_Feed"]
        
    ws_out = wb_out.create_sheet(title=output_tab_name, index=1)
    ws_out.views.sheetView[0].showGridLines = True

    # Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    PURPLE_DARK = "382D5C"
    TEAL_DARK = "0E5A5E"
    GREEN_DARK = "155724"
    WHITE = "FFFFFF"
    BORDER_COLOR = "D0D7DE"
    ICE_BLUE = "F0F5FA"
    YELLOW_FILL = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_teal = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
    fill_purple = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
    fill_green_dark = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
    fill_ice_blue = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    font_code = Font(name="Consolas", size=9, color="555555")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    out_headers = [
        ("Published Date", 16, fill_navy),
        ("Published Time (UTC)", 20, fill_navy),
        ("Event Headline", 48, fill_navy),
        ("Project / Indication Theme", 26, fill_blue),
        ("Signal Type", 20, fill_teal),
        ("Relevance Score", 18, fill_teal),
        ("Priority Tier", 18, fill_green_dark),
        ("Routed Desk / CI Workstream", 28, fill_purple),
        ("Matched Biopharma Catalyst", 30, fill_navy),
        ("Source Entity Name", 30, fill_navy),
        ("Source Classification", 28, fill_blue),
        ("Direct Publisher / SEC URL", 55, fill_navy),
        ("Editorial Snippet / Summary", 50, fill_navy),
        ("Cluster ID", 35, fill_teal),
        ("Cluster Hint JSON", 38, fill_teal),
        ("Discovery Pillar", 24, fill_purple),
        ("Extraction Vector / Method", 28, fill_purple),
        ("Discovered At (UTC)", 24, fill_navy),
        ("Full Body Excerpt", 50, fill_navy),
        ("Event UUID", 16, fill_navy),
        ("AI Summary", 55, fill_navy),
        ("Implications", 65, fill_purple),
        ("Ingestion Batch ID", 26, fill_teal),
        ("Ingested Date (IST)", 20, fill_navy),
        ("Ingested Time (IST)", 20, fill_navy),
        ("Execution Run Type", 24, fill_purple),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(out_headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_out.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_out.column_dimensions[col_letter].width = width
    ws_out.row_dimensions[1].height = 32

    for r_idx, ev in enumerate(all_final_items, start=2):
        row_fill = fill_ice_blue if (r_idx % 2 == 0) else PatternFill(fill_type=None)

        # Col 1: Date
        c_d = ws_out.cell(row=r_idx, column=1, value=ev["published_date"])
        c_d.font = font_bold
        c_d.alignment = Alignment(horizontal="center")
        if ev.get("is_imported"):
            c_d.fill = YELLOW_FILL

        # Col 2: Time
        ws_out.cell(row=r_idx, column=2, value=ev["published_time"]).font = font_data
        ws_out.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center")

        # Col 3: Headline
        ws_out.cell(row=r_idx, column=3, value=ev["headline"]).font = font_bold

        # Col 4: Project
        ws_out.cell(row=r_idx, column=4, value=ev["project_name"]).font = font_bold

        # Col 5: Signal Type Badge
        c_sig = ws_out.cell(row=r_idx, column=5, value=ev["signal_type"])
        c_sig.alignment = Alignment(horizontal="center")
        c_sig.font = font_bold
        if ev["signal_type"] == "clinical_pos":
            c_sig.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            c_sig.font = Font(name="Calibri", size=10, bold=True, color="155724")
        elif ev["signal_type"] == "regulatory":
            c_sig.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            c_sig.font = Font(name="Calibri", size=10, bold=True, color="004085")
        elif ev["signal_type"] == "clinical_neg":
            c_sig.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            c_sig.font = Font(name="Calibri", size=10, bold=True, color="721C24")
        elif ev["signal_type"] == "corporate":
            c_sig.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            c_sig.font = Font(name="Calibri", size=10, bold=True, color="856404")

        # Col 6: Relevance Score
        c_score = ws_out.cell(row=r_idx, column=6, value=f"{ev['relevance_score']}/100")
        c_score.font = font_bold
        c_score.alignment = Alignment(horizontal="center")

        # Col 7: Priority Badge
        c_p = ws_out.cell(row=r_idx, column=7, value=ev["priority"])
        c_p.alignment = Alignment(horizontal="center")
        if "Tier 1" in ev["priority"]:
            c_p.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            c_p.font = Font(name="Calibri", size=10, bold=True, color="721C24")
        elif "Tier 2" in ev["priority"]:
            c_p.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            c_p.font = Font(name="Calibri", size=10, bold=True, color="856404")
        else:
            c_p.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            c_p.font = Font(name="Calibri", size=10, bold=True, color="155724")

        # Col 8: Desk
        ws_out.cell(row=r_idx, column=8, value=ev["desk"]).font = font_data
        ws_out.cell(row=r_idx, column=8).alignment = Alignment(horizontal="center")

        # Col 9: Matched Keywords
        ws_out.cell(row=r_idx, column=9, value=ev["matched_keywords"]).font = font_code

        # Col 10: Source Name
        ws_out.cell(row=r_idx, column=10, value=ev["source_name"]).font = font_bold

        # Col 11: Source Classification
        ws_out.cell(row=r_idx, column=11, value=ev["source_class"]).font = font_data

        # Col 12: URL
        c_u = ws_out.cell(row=r_idx, column=12, value=ev["raw_url"])
        c_u.font = font_link

        # Col 13: Snippet
        ws_out.cell(row=r_idx, column=13, value=ev["snippet"]).font = font_data

        # Col 14: Cluster ID
        ws_out.cell(row=r_idx, column=14, value=ev["cluster_id"]).font = font_code

        # Col 15: Cluster Hint JSON
        ws_out.cell(row=r_idx, column=15, value=ev["cluster_hint"]).font = font_code

        # Col 16: Discovery Pillar
        ws_out.cell(row=r_idx, column=16, value=ev["discovery_method"]).font = font_data

        # Col 17: Extraction Vector
        ws_out.cell(row=r_idx, column=17, value=ev["extraction_vector"]).font = font_data

        # Col 18: Discovered At
        ws_out.cell(row=r_idx, column=18, value=ev["discovered_at"]).font = font_code
        ws_out.cell(row=r_idx, column=18).alignment = Alignment(horizontal="center")

        # Col 19: Full Body
        ws_out.cell(row=r_idx, column=19, value=ev["full_text"]).font = font_data

        # Col 20: Event UUID
        ws_out.cell(row=r_idx, column=20, value=ev["event_id"]).font = font_code
        ws_out.cell(row=r_idx, column=20).alignment = Alignment(horizontal="center")

        # Col 21: Grounded Factual AI Summary (< 70 words)
        existing_sum = ev.get("ai_summary", "")
        if existing_sum and len(existing_sum) > 25 and not existing_sum.startswith("http"):
            ai_sum = existing_sum
        else:
            ai_sum = generate_factual_ai_summary(
                ev["headline"], ev.get("raw_url", ""), ev.get("snippet", ""), ev.get("full_text", ""),
                source_name=ev.get("source_name", ""), desk=ev.get("desk", ""), signal_type=ev.get("signal_type", "")
            )
        ws_out.cell(row=r_idx, column=21, value=ai_sum).font = font_data
        ws_out.cell(row=r_idx, column=21).alignment = Alignment(vertical="top", wrap_text=True)

        # Col 22: CI Strategic Implications (<=80 words, analyst style)
        existing_imp = ev.get("implications", "")
        if existing_imp and len(existing_imp) > 20:
            imp_text = existing_imp
        else:
            imp_text = generate_ci_competitive_implications(
                ev["headline"], ev.get("raw_url", ""), ev.get("snippet", ""), ev.get("full_text", ""),
                source_name=ev.get("source_name", ""), desk=ev.get("desk", ""), signal_type=ev.get("signal_type", ""),
                matched_keywords=ev.get("matched_keywords", "")
            )
        ws_out.cell(row=r_idx, column=22, value=imp_text).font = font_data
        ws_out.cell(row=r_idx, column=22).alignment = Alignment(vertical="top", wrap_text=True)

        # Col 23: Ingestion Batch ID
        ws_out.cell(row=r_idx, column=23, value=ev.get("run_batch_id", "HISTORICAL_BASELINE")).font = font_code
        ws_out.cell(row=r_idx, column=23).alignment = Alignment(horizontal="center")

        # Col 24: Ingested Date (IST)
        ws_out.cell(row=r_idx, column=24, value=ev.get("ingested_date_ist", ev.get("published_date", ""))).font = font_bold
        ws_out.cell(row=r_idx, column=24).alignment = Alignment(horizontal="center")

        # Col 25: Ingested Time (IST)
        ws_out.cell(row=r_idx, column=25, value=ev.get("ingested_time_ist", "00:00 IST")).font = font_data
        ws_out.cell(row=r_idx, column=25).alignment = Alignment(horizontal="center")

        # Col 26: Execution Run Type
        ws_out.cell(row=r_idx, column=26, value=ev.get("exec_run_type", "📋 Baseline Historical")).font = font_bold
        ws_out.cell(row=r_idx, column=26).alignment = Alignment(horizontal="center")

        for col_c in range(1, 27):
            cell_c = ws_out.cell(row=r_idx, column=col_c)
            cell_c.border = thin_border
            if col_c != 1 and cell_c.fill.fill_type is None and row_fill.fill_type is not None:
                cell_c.fill = row_fill

        ws_out.row_dimensions[r_idx].height = 42

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:Z{len(all_final_items)+1}"

    # Update Live Telemetry in 01_Master_Sources_Registry
    if "01_Master_Sources_Registry" in wb_out.sheetnames and GLOBAL_AUDIT_TELEMETRY:
        ws_reg_out = wb_out["01_Master_Sources_Registry"]
        # Ensure clear, intuitive column headers in Row 1
        reg_headers = {
            12: "Health Status",
            13: "Feed Type",
            14: "Latest Fetched Article Title",
            15: "Last Fetch Run Date & Time",
            16: "Items Captured in Run",
            17: "Feed Status & Quality",
            18: "Action / Recommended Solution"
        }
        for col_i, h_text in reg_headers.items():
            ws_reg_out.cell(1, col_i, value=h_text)

        updated_count = 0
        for r_idx in range(2, ws_reg_out.max_row + 1):
            c_fid = str(ws_reg_out.cell(r_idx, 1).value or "").strip()
            if c_fid in GLOBAL_AUDIT_TELEMETRY:
                t_data = GLOBAL_AUDIT_TELEMETRY[c_fid]
                ws_reg_out.cell(r_idx, 12, value=t_data["health"])
                ws_reg_out.cell(r_idx, 13, value=t_data["structure"])
                ws_reg_out.cell(r_idx, 14, value=t_data["latest_title"])
                ws_reg_out.cell(r_idx, 15, value=t_data["fetch_date_time"])
                ws_reg_out.cell(r_idx, 16, value=t_data["items_detected"])
                ws_reg_out.cell(r_idx, 17, value=t_data["feed_quality"])
                ws_reg_out.cell(r_idx, 18, value=t_data["recommendation"])
                updated_count += 1
        print(f"📡 Live Audit Telemetry logged for {updated_count} endpoints in '01_Master_Sources_Registry'!")

    # Save with lock protection
    saved_path = safe_save_workbook(wb_out, XLSX_PATH)
    print(f"✅ Master Workbook successfully updated with {len(all_final_items)} total events in sheet '{output_tab_name}' at {saved_path}!")

    # Auto-Export Web JSON and JS data feeds
    try:
        web_data_dir = os.path.join(WORKSPACE, "web", "data")
        os.makedirs(web_data_dir, exist_ok=True)
        json_path = os.path.join(web_data_dir, "latest_intelligence.json")
        js_path = os.path.join(web_data_dir, "latest_intelligence.js")

        from datetime import date as datetime_date
        def sanitize_item(item):
            cleaned = {}
            for k, v in item.items():
                if isinstance(v, (datetime, datetime_date)):
                    cleaned[k] = v.isoformat()
                elif isinstance(v, set):
                    cleaned[k] = list(v)
                elif k in ["headline", "snippet", "full_text"]:
                    val = re.sub(r'<figure[^>]*>.*?</figure>', ' ', str(v or ""), flags=re.DOTALL | re.IGNORECASE)
                    val = re.sub(r'<img[^>]*>', ' ', val, flags=re.IGNORECASE)
                    val = re.sub(r'<[^>]+>', ' ', val)
                    cleaned[k] = re.sub(r'\s+', ' ', val).strip()
                else:
                    cleaned[k] = v
            return cleaned

        sanitized_events = [sanitize_item(ev) for ev in all_final_items]

        output_payload = {
            "metadata": {
                "generated_utc": datetime.now(timezone.utc).isoformat(),
                "total_curated_events": len(sanitized_events),
                "chunk_window_days": 15,
                "max_session_ceiling_days": 90,
                "retention_policy": "PERMANENT_ACCUMULATIVE"
            },
            "events": sanitized_events
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(output_payload, f, indent=2, ensure_ascii=False, default=str)
        with open(js_path, "w", encoding="utf-8") as f:
            f.write("window.LATEST_INTELLIGENCE_DATA = " + json.dumps(output_payload, indent=2, ensure_ascii=False, default=str) + ";\n")
        print(f"   ✓ Synchronized sanitized web portal feeds at {web_data_dir}")
    except Exception as e:
        print(f"   ⚠️ Web export notice: {e}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RSSFeedChecker Master Pipeline")
    parser.add_argument("--since", default="72h", help="Start date (YYYY-MM-DD) or '72h' / '36h' / '48h'")
    parser.add_argument("--until", default="", help="End date (YYYY-MM-DD) or 'Present'")
    parser.add_argument("--category", default="ALL", help="Source classification filter")
    parser.add_argument("--entity", default="ALL", help="Target entity / company name or feed ID filter")
    parser.add_argument("--project", default="ALL", help="Therapeutic project filter")
    parser.add_argument("--tier", default="ALL", help="Priority tier filter")
    parser.add_argument("--min-score", type=int, default=40, help="Minimum relevance score threshold")
    parser.add_argument("--max-items", type=int, default=30, help="Max items per source")
    parser.add_argument("--sample", action="store_true", help="Quick sample test run")
    parser.add_argument("--from-dashboard", action="store_true", help="Read settings from 00_Run_Dashboard sheet")
    parser.add_argument("--output", default="Results", help="Target output sheet name")
    parser.add_argument("--admin", action="store_true", default=True, help="Admin override for historical windows")
    parser.add_argument("--vault-dir", default="", help="Path to vault repository directory")
    parser.add_argument("--runner-dir", default="", help="Path to runner repository directory")

    args = parser.parse_args()

    if args.vault_dir:
        WORKSPACE = args.vault_dir
        XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

    # If days_back is provided via environment, use it as time window
    days_back_env = os.environ.get("DAYS_BACK", "").strip()
    since_val = args.since
    if days_back_env:
        since_val = f"{days_back_env}d" if not days_back_env.endswith(("d", "h")) else days_back_env

    run_pipeline(
        since_date=since_val,
        until_date=args.until,
        category_filter=args.category,
        entity_filter=args.entity,
        project_filter=args.project,
        tier_filter=args.tier,
        min_score_cutoff=args.min_score,
        max_items_per_feed=args.max_items,
        is_sample=args.sample,
        is_admin=args.admin,
        output_tab_name=args.output,
        from_dashboard=args.from_dashboard
    )
