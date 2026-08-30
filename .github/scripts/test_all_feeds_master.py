#!/usr/bin/env python3
"""
Test all 518 RSS feeds from '03_Feeds_Master (518)' tab in the master Excel
and write results back into the same tab with new columns colored differently.

Uses the same fetch/classify/fallback logic from check_feeds.py.
"""

import os
import sys
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

# Fix Windows console encoding
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

# --- Import the core fetch/classify/fallback from check_feeds.py ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from check_feeds import fetch, classify, try_fallback_cascade, _first_item_title

# ----------------------------------------------------------------------------- config
WORKERS = 10          # parallel threads
XLSX_PATH = os.path.join(SCRIPT_DIR, "..", "..", "RSSFeedChecker_Master_Guide_and_Data.xlsx")
XLSX_PATH = os.path.normpath(XLSX_PATH)
SHEET_NAME = "03_Feeds_Master (518)"

# Result column headers (will be added starting at column 6)
RESULT_HEADERS = [
    "HTTP Code",           # Col F (6)
    "Response Time (s)",   # Col G (7)
    "Result",              # Col H (8)
    "Last Item Title",     # Col I (9)
    "Fallback URL",        # Col J (10)
    "Test Timestamp",      # Col K (11)
]


def check_feed(feed_id, url, label):
    """Check one feed using full cascade logic. Returns dict of results."""
    code, elapsed, body, reason, lax_ok = fetch(url)
    result, title = classify(code, body, reason, lax_ok)
    fallback_url = ""

    # If primary failed, trigger fallback cascade
    if not (code == 200 and title) or "BLOCKED" in result or "FAILED" in result:
        fb_result, fb_title, fb_url = try_fallback_cascade(url, label)
        if fb_result and fb_title:
            result = fb_result
            title = fb_title
            fallback_url = fb_url

    return {
        "code": code,
        "time": round(elapsed, 3),
        "result": result,
        "title": title[:200] if title else "",
        "fallback_url": fallback_url,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
    }


def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    print(f"Opening: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Dynamic resilient sheet name resolver
    target_sheet = next((s for s in wb.sheetnames if "Feeds" in s), None)
    if not target_sheet:
        print(f"ERROR: No Feeds sheet found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[target_sheet]
    total_rows = ws.max_row
    existing_cols = ws.max_column
    print(f"Found sheet '{target_sheet}' with {total_rows} rows (header + {total_rows - 1} feeds), {existing_cols} columns")

    # Read all feeds
    feeds = []
    for row in range(2, total_rows + 1):
        feed_id = str(ws.cell(row=row, column=1).value or "")
        url = str(ws.cell(row=row, column=2).value or "")
        label = str(ws.cell(row=row, column=3).value or "")
        if url.startswith("http"):
            feeds.append((row, feed_id, url, label))
        else:
            feeds.append((row, feed_id, url, label))  # keep even bad ones

    print(f"\nFound {len(feeds)} feeds to test")
    print(f"Using {WORKERS} parallel threads")
    print("=" * 78)

    # --- Write result headers (Col 6-11) with distinct color ---
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # dark blue
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, header in enumerate(RESULT_HEADERS):
        col = 6 + i
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # --- Color fills for results ---
    OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # light green
    RECOVERED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
    BLOCKED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")   # light pink
    FAILED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")    # red-ish
    OTHER_FILL = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")     # light blue-grey
    DATA_FONT = Font(size=10, name="Calibri")
    TITLE_FONT = Font(size=9, name="Calibri", italic=True)

    def get_fill(result_str):
        if "OK" in result_str and "fetchable" in result_str:
            return OK_FILL
        elif "RECOVERED" in result_str:
            return RECOVERED_FILL
        elif "BLOCKED" in result_str:
            return BLOCKED_FILL
        elif "FAILED" in result_str:
            return FAILED_FILL
        else:
            return OTHER_FILL

    # --- Run all feeds in parallel ---
    results_map = {}
    done_count = 0
    lock = threading.Lock()

    def process_feed(item):
        row_num, feed_id, url, label = item
        if not url.startswith("http"):
            return row_num, {
                "code": 0, "time": 0, "result": "SKIPPED — no valid URL",
                "title": "", "fallback_url": "",
                "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            }
        return row_num, check_feed(feed_id, url, label)

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = {pool.submit(process_feed, f): f for f in feeds}
        for fut in as_completed(futures):
            row_num, res = fut.result()
            results_map[row_num] = res
            with lock:
                done_count += 1
                status_icon = "OK" if "OK" in res["result"] and "fetchable" in res["result"] else \
                              "RECOVERED" if "RECOVERED" in res["result"] else \
                              "BLOCKED" if "BLOCKED" in res["result"] else \
                              "FAIL" if "FAILED" in res["result"] else "other"
                elapsed_total = time.time() - start_time
                print(f"[{done_count:>4}/{len(feeds)}] {futures[fut][1]:<12} "
                      f"{res['code']:<4} {res['time']:>6.2f}s  {status_icon:<10} "
                      f"{res['result'][:50]}  [{elapsed_total:.0f}s elapsed]",
                      flush=True)

    total_elapsed = time.time() - start_time

    # --- Write results to Excel ---
    print(f"\nWriting results to Excel...")

    for row_num, res in results_map.items():
        fill = get_fill(res["result"])

        values = [
            res["code"],
            res["time"],
            res["result"],
            res["title"],
            res["fallback_url"],
            res["timestamp"],
        ]

        for i, val in enumerate(values):
            col = 6 + i
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            if i == 3:  # title column
                cell.font = TITLE_FONT
            else:
                cell.font = DATA_FONT
            cell.alignment = Alignment(wrap_text=(i in (2, 3, 4)))

    # --- Set column widths ---
    col_widths = {
        "F": 12,   # HTTP Code
        "G": 14,   # Response Time
        "H": 42,   # Result
        "I": 55,   # Last Item Title
        "J": 60,   # Fallback URL
        "K": 20,   # Timestamp
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Summary stats ---
    ok = sum(1 for r in results_map.values() if "OK" in r["result"] and "fetchable" in r["result"])
    recovered = sum(1 for r in results_map.values() if "RECOVERED" in r["result"])
    blocked = sum(1 for r in results_map.values() if "BLOCKED" in r["result"])
    failed = sum(1 for r in results_map.values() if "FAILED" in r["result"])
    other = len(results_map) - ok - recovered - blocked - failed

    print("\n" + "=" * 78)
    print(f"RESULTS SUMMARY — {len(results_map)} feeds tested in {total_elapsed:.1f}s")
    print("=" * 78)
    print(f"  OK (fetchable):     {ok}")
    print(f"  RECOVERED:          {recovered}")
    print(f"  BLOCKED:            {blocked}")
    print(f"  FAILED:             {failed}")
    print(f"  Other:              {other}")
    print("=" * 78)

    # --- Save ---
    try:
        wb.save(XLSX_PATH)
        print(f"\nSaved to: {XLSX_PATH}")
    except PermissionError:
        alt_path = XLSX_PATH.replace(".xlsx", "_results.xlsx")
        print(f"\nExcel file is open! Saving to: {alt_path}")
        wb.save(alt_path)
        print(f"Saved to: {alt_path}")


if __name__ == "__main__":
    main()
