import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl
import json
import os
import re
from datetime import datetime, timezone

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")
WEB_DATA_DIR = os.path.join(WORKSPACE, "web", "data")
os.makedirs(WEB_DATA_DIR, exist_ok=True)

JSON_OUTPUT_PATH = os.path.join(WEB_DATA_DIR, "latest_intelligence.json")
JS_OUTPUT_PATH = os.path.join(WEB_DATA_DIR, "latest_intelligence.js")

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb["Results"]

def clean_html_tags(text):
    if not text:
        return ""
    # Strip HTML tags like <figure>, <img>, <div>
    clean = re.sub(r'<figure[^>]*>.*?</figure>', ' ', text, flags=re.DOTALL | re.IGNORECASE)
    clean = re.sub(r'<img[^>]*>', ' ', clean, flags=re.IGNORECASE)
    clean = re.sub(r'<[^>]+>', ' ', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean

items = []

for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 21)]
    if not any(row_vals):
        continue

    score_raw = str(row_vals[5] or "50").split("/")[0].strip()
    score = int(score_raw) if score_raw.isdigit() else 50

    hint_str = str(row_vals[14] or "{}").strip()
    lead_company = "Not Identified"
    partner_company = ""
    try:
        hint_obj = json.loads(hint_str)
        lead_company = hint_obj.get("lead_company", "Not Identified")
        partner_company = hint_obj.get("partner_company", "")
    except Exception:
        pass

    raw_snippet = str(row_vals[12] or "").strip()
    raw_full_text = str(row_vals[18] or "").strip()

    item = {
        "published_date": str(row_vals[0] or "").strip(),
        "published_time": str(row_vals[1] or "").strip(),
        "headline": clean_html_tags(str(row_vals[2] or "").strip()),
        "project_name": str(row_vals[3] or "").strip(),
        "signal_type": str(row_vals[4] or "general").strip(),
        "relevance_score": score,
        "priority_tier": str(row_vals[6] or "🟡 Tier 2 (Daily)").strip(),
        "assigned_desk": str(row_vals[7] or "").strip(),
        "matched_keywords": str(row_vals[8] or "").strip(),
        "source_name": str(row_vals[9] or "").strip(),
        "source_class": str(row_vals[10] or "").strip(),
        "raw_url": str(row_vals[11] or "").strip(),
        "snippet": clean_html_tags(raw_snippet),
        "cluster_id": str(row_vals[13] or "").strip(),
        "cluster_hint": hint_str,
        "discovery_method": str(row_vals[15] or "").strip(),
        "extraction_vector": str(row_vals[16] or "").strip(),
        "discovered_at": str(row_vals[17] or "").strip(),
        "full_text": clean_html_tags(raw_full_text) if raw_full_text else clean_html_tags(raw_snippet),
        "event_id": str(row_vals[19] or f"EVT_{r:04d}").strip(),
        "matched_company": lead_company,
        "partner_company": partner_company
    }
    items.append(item)

# Sort by published_date descending
items.sort(key=lambda x: (x.get("published_date", ""), x.get("published_time", "")), reverse=True)

output_payload = {
    "metadata": {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "total_curated_events": len(items),
        "chunk_window_days": 15,
        "max_session_ceiling_days": 90,
        "retention_policy": "PERMANENT_ACCUMULATIVE"
    },
    "events": items
}

# 1. Write JSON for HTTP / Cloudflare Pages
with open(JSON_OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(output_payload, f, indent=2, ensure_ascii=False)

# 2. Write JS for direct double-click file:// local opening
with open(JS_OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write("window.LATEST_INTELLIGENCE_DATA = " + json.dumps(output_payload, indent=2, ensure_ascii=False) + ";\n")

print(f"✅ Cleaned HTML tags & updated payloads with {len(items)} curated events!")
