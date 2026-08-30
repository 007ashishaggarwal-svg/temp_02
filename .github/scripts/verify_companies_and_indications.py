#!/usr/bin/env python3
"""
Comprehensive Live Verification & Multi-Strategy Intelligence Engine for:
1. 04_Companies_Master (616 Companies)
2. 05_Indications_Radar (18 Indications)

Performs parallel HTTP requests, sitemap probes, RSS discovery, Google News fallbacks,
and ClinicalTrials.gov API validations, updating the Master Excel workbook with
professional color-coded columns, status badges, and actionable CI observations.
"""

import os
import sys
import re
import time
import json
import ssl
import urllib.request
import urllib.parse
import urllib.error
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX_PATH = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

_LAX_CTX = ssl.create_default_context()
_LAX_CTX.check_hostname = False
_LAX_CTX.verify_mode = ssl.CERT_NONE

# -----------------------------------------------------------------------------
# Helper: Fetch URL with timeout & lax SSL
# -----------------------------------------------------------------------------
def fetch_url(url, timeout=8, max_bytes=512*1024):
    """Fetch URL and return (status_code, elapsed_sec, body_bytes, final_url, error_msg)."""
    start = time.perf_counter()
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    try:
        with urllib.request.urlopen(req, timeout=timeout, context=_LAX_CTX) as resp:
            code = resp.getcode()
            final_url = resp.geturl()
            body = resp.read(max_bytes)
            elapsed = time.perf_counter() - start
            return code, elapsed, body, final_url, ""
    except urllib.error.HTTPError as e:
        elapsed = time.perf_counter() - start
        try:
            body = e.read(64*1024)
        except Exception:
            body = b""
        return e.code, elapsed, body, url, str(e.reason)
    except urllib.error.URLError as e:
        elapsed = time.perf_counter() - start
        return 0, elapsed, b"", url, str(e.reason)[:80]
    except Exception as e:
        elapsed = time.perf_counter() - start
        return 0, elapsed, b"", url, str(e)[:80]

def extract_headline_from_xml(body):
    """Extract first headline from RSS/Atom/XML body."""
    if not body:
        return ""
    try:
        root = ET.fromstring(body)
        for item in root.iter("item"):
            t = item.find("title")
            if t is not None and t.text:
                return " ".join(t.text.split())[:180]
        for entry in root.iter("{http://www.w3.org/2005/Atom}entry"):
            t = entry.find("{http://www.w3.org/2005/Atom}title")
            if t is not None and t.text:
                return " ".join(t.text.split())[:180]
    except Exception:
        # Regex fallback
        m = re.search(r"<(?:item|entry)\b.*?<title[^>]*>(.*?)</title>", body.decode("utf-8", "ignore"), re.I | re.S)
        if m:
            clean = re.sub(r"<[^>]+>", "", m.group(1)).replace("&amp;", "&").replace("&quot;", '"').strip()
            return clean[:180]
    return ""

def extract_title_from_html(body):
    """Extract <title> or <h1> from HTML page."""
    if not body:
        return ""
    text = body.decode("utf-8", "ignore")
    m = re.search(r"<title[^>]*>(.*?)</title>", text, re.I | re.S)
    if m:
        t = re.sub(r"<[^>]+>", "", m.group(1)).replace("&amp;", "&").replace("&quot;", '"').strip()
        t = " ".join(t.split())
        return t[:150]
    return ""

# -----------------------------------------------------------------------------
# COMPANY VERIFICATION WORKER
# -----------------------------------------------------------------------------
def verify_company(row_idx, company_name, domain, path, strategy, notes):
    """Verify a single company domain and newsroom path."""
    domain_clean = domain.strip().lower()
    path_clean = (path or "").strip()
    if path_clean and not path_clean.startswith("/"):
        path_clean = "/" + path_clean

    # 1. Test Primary Newsroom URL
    if path_clean:
        primary_url = f"https://{domain_clean}{path_clean}"
    else:
        primary_url = f"https://{domain_clean}"

    code, elapsed, body, final_url, err = fetch_url(primary_url, timeout=8)

    # 2. Check Native RSS autodiscovery / common paths
    has_rss = False
    rss_url = ""
    rss_headline = ""
    for r_path in ["/feed/", "/rss.xml", "/rss", "/news/rss.xml", "/investors/rss.xml"]:
        cand_rss = f"https://{domain_clean}{r_path}"
        r_code, _, r_body, _, _ = fetch_url(cand_rss, timeout=5)
        if r_code == 200 and (b"<rss" in r_body or b"<feed" in r_body or b"<channel" in r_body):
            has_rss = True
            rss_url = cand_rss
            rss_headline = extract_headline_from_xml(r_body)
            break

    # 3. Check XML Sitemap
    has_sitemap = False
    sitemap_url = f"https://{domain_clean}/sitemap.xml"
    s_code, _, s_body, _, _ = fetch_url(sitemap_url, timeout=5)
    if s_code == 200 and (b"<urlset" in s_body or b"<sitemapindex" in s_body):
        has_sitemap = True

    # 4. Check Google News Search Fallback
    q = urllib.parse.quote(f"site:{domain_clean} (press OR release OR approval OR clinical)")
    gnews_url = f"https://news.google.com/rss/search?q={q}&hl=en-US&gl=US&ceid=US:en"
    g_code, _, g_body, _, _ = fetch_url(gnews_url, timeout=6)
    gnews_headline = extract_headline_from_xml(g_body) if g_code == 200 else ""

    # 5. Extract latest headline
    latest_headline = rss_headline or gnews_headline or extract_title_from_html(body)

    # 6. Determine Recommended Strategy
    if has_rss:
        rec_strategy = "Native RSS (Direct)"
    elif has_sitemap and code == 200:
        rec_strategy = "XML Sitemap (Deep Timestamps)"
    elif code == 200:
        rec_strategy = "HTML Scraping (Custom Parser)"
    elif code in (401, 403):
        rec_strategy = "Google News Search Fallback (WAF Protected)"
    else:
        rec_strategy = "Google News Search Fallback (Domain Fallback)"

    # 7. Formulate Observations & CI Notes
    observations = []
    if code == 200:
        if has_rss:
            observations.append(f"Native RSS active at {rss_url}")
        if has_sitemap:
            observations.append("XML Sitemap available")
        if not has_rss and not has_sitemap:
            observations.append("Standard HTML newsroom; scraping active")
    elif code in (401, 403):
        observations.append("WAF bot challenge (Akamai/Cloudflare/Imperva); Google News fallback recommended")
    elif code == 404:
        observations.append("Newsroom path 404; Google News fallback or homepage search recommended")
    elif code == 0:
        observations.append(f"Connection timeout / DNS error ({err or 'unreachable'}); Google News fallback active")
    else:
        observations.append(f"HTTP {code} returned; fallback active")

    if final_url and final_url.rstrip("/") != primary_url.rstrip("/"):
        observations.append(f"Redirected to: {final_url}")

    obs_str = " | ".join(observations)

    return {
        "row_idx": row_idx,
        "company": company_name,
        "domain": domain_clean,
        "http_code": code,
        "elapsed": round(elapsed, 3),
        "verified_url": final_url if code == 200 else primary_url,
        "rec_strategy": rec_strategy,
        "has_rss": f"Yes ({rss_url})" if has_rss else "No",
        "has_sitemap": f"Yes ({sitemap_url})" if has_sitemap else "No",
        "gnews_query": gnews_url,
        "latest_headline": latest_headline[:180],
        "observations": obs_str
    }

# -----------------------------------------------------------------------------
# INDICATION VERIFICATION WORKER
# -----------------------------------------------------------------------------
def verify_indication(row_idx, indication, synonyms, ct_cond, exclude_terms, broad_query, clin_query, ct_endpoint):
    """Verify broad GNews, clinical GNews, and CT.gov endpoints for an indication."""
    # 1. Broad Google News RSS Query
    # Convert query logic to valid Google News search URL
    b_q_clean = broad_query.replace(" AND ", " ").replace(" OR ", " OR ").strip()
    b_url = f"https://news.google.com/rss/search?q={urllib.parse.quote(b_q_clean)}&hl=en-US&gl=US&ceid=US:en"
    b_code, b_elapsed, b_body, _, _ = fetch_url(b_url, timeout=8)
    b_title = extract_headline_from_xml(b_body) if b_code == 200 else ""
    b_count = len(re.findall(r"<item\b", b_body.decode("utf-8", "ignore"))) if b_code == 200 else 0

    # 2. Clinical/Reg Google News RSS Query
    c_q_clean = clin_query.replace(" AND ", " ").replace(" OR ", " OR ").strip()
    c_url = f"https://news.google.com/rss/search?q={urllib.parse.quote(c_q_clean)}&hl=en-US&gl=US&ceid=US:en"
    c_code, c_elapsed, c_body, _, _ = fetch_url(c_url, timeout=8)
    c_title = extract_headline_from_xml(c_body) if c_code == 200 else ""
    c_count = len(re.findall(r"<item\b", c_body.decode("utf-8", "ignore"))) if c_code == 200 else 0

    # 3. ClinicalTrials.gov Protocol Change Endpoint
    ct_status = "N/A"
    ct_count = 0
    ct_latest = "N/A"
    if ct_endpoint and ct_endpoint.startswith("http"):
        ct_code, ct_elapsed, ct_body, _, ct_err = fetch_url(ct_endpoint, timeout=10)
        if ct_code == 200:
            try:
                data = json.loads(ct_body.decode("utf-8", "ignore"))
                ct_count = data.get("totalCount", 0)
                studies = data.get("studies", [])
                if studies:
                    s0 = studies[0].get("protocolSection", {})
                    nct = s0.get("identificationModule", {}).get("nctId", "")
                    brief = s0.get("identificationModule", {}).get("briefTitle", "")
                    upd_date = s0.get("statusModule", {}).get("lastUpdatePostDateStruct", {}).get("date", "")
                    ct_latest = f"{nct} ({upd_date}): {brief[:100]}"
                ct_status = f"200 OK ({ct_count:,} active trials)"
            except Exception as e:
                ct_status = f"Parse Error ({str(e)[:40]})"
        else:
            ct_status = f"HTTP {ct_code} ({ct_err[:30]})"
    else:
        ct_status = "N/A (Industry / Regulatory tracking only)"

    # 4. Observations & Specificity
    obs = []
    if b_count > 0:
        obs.append(f"Broad Radar: {b_count} recent articles")
    if c_count > 0:
        obs.append(f"Clinical Radar: {c_count} high-specificity trial articles")
    if ct_count > 0:
        obs.append(f"CT.gov: {ct_count:,} registered protocols")
    
    obs_str = " | ".join(obs) if obs else "Active query pipeline"

    return {
        "row_idx": row_idx,
        "indication": indication,
        "b_code": b_code,
        "b_url": b_url,
        "b_title": b_title,
        "b_count": b_count,
        "c_code": c_code,
        "c_url": c_url,
        "c_title": c_title,
        "c_count": c_count,
        "ct_status": ct_status,
        "ct_count": ct_count,
        "ct_latest": ct_latest,
        "observations": obs_str
    }

# -----------------------------------------------------------------------------
# MAIN RUNNER & WORKBOOK FORMATTER
# -----------------------------------------------------------------------------
def run():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    
    # -------------------------------------------------------------------------
    # STEP 1: Rename Tabs to Standardized Clean Scheme
    # -------------------------------------------------------------------------
    print("\n--- STEP 1: Standardizing Tab Names ---")
    sheet_renames = {}
    for name in wb.sheetnames:
        if "01_System" in name:
            sheet_renames[name] = "01_System_Overview"
        elif "02_Data" in name:
            sheet_renames[name] = "02_Data_Dictionary"
        elif "03_Feeds" in name:
            ws_feeds = wb[name]
            count = ws_feeds.max_row - 1
            sheet_renames[name] = f"03_Feeds_Master ({count})"
        elif "04_Companies" in name:
            ws_comp = wb[name]
            count = ws_comp.max_row - 1
            sheet_renames[name] = f"04_Companies_Master ({count})"
        elif "Indications" in name:
            ws_ind = wb[name]
            count = ws_ind.max_row - 1
            sheet_renames[name] = f"05_Indications_Radar ({count})"
        elif "Config" in name or "08_Config" in name or "06_Config" in name:
            sheet_renames[name] = "06_Config_and_Settings"

    for old_n, new_n in sheet_renames.items():
        if old_n != new_n:
            print(f"  Renaming tab: '{old_n}' -> '{new_n}'")
            wb[old_n].title = new_n

    print(f"Standardized Sheets: {wb.sheetnames}")

    # Shared Styles
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    CRIMSON_DARK = "8B0000"
    WHITE = "FFFFFF"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_crimson = PatternFill(start_color=CRIMSON_DARK, end_color=CRIMSON_DARK, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_title = Font(name="Calibri", size=9, italic=True, color="222222")
    font_flag = Font(name="Calibri", size=9, color="111111")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")

    fill_ice = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")       # soft green
    fill_recovered = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")# soft yellow
    fill_blocked = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # soft pink
    fill_failed = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")   # soft red
    fill_note = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")     # soft blue

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    # -------------------------------------------------------------------------
    # STEP 2: Live Verification for 04_Companies_Master
    # -------------------------------------------------------------------------
    comp_tab_name = [n for n in wb.sheetnames if "Companies" in n][0]
    ws_comp = wb[comp_tab_name]
    print(f"\n--- STEP 2: Verifying {ws_comp.max_row - 1} Companies in '{comp_tab_name}' ---")

    companies_to_test = []
    for r in range(2, ws_comp.max_row + 1):
        c_name = str(ws_comp.cell(row=r, column=1).value or "")
        dom = str(ws_comp.cell(row=r, column=2).value or "")
        path = str(ws_comp.cell(row=r, column=3).value or "")
        strat = str(ws_comp.cell(row=r, column=4).value or "")
        notes = str(ws_comp.cell(row=r, column=5).value or "")
        if dom:
            companies_to_test.append((r, c_name, dom, path, strat, notes))

    print(f"Launching parallel verification across {len(companies_to_test)} companies (20 worker threads)...")
    comp_results = {}
    done_c = 0
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=20) as pool:
        futures = {pool.submit(verify_company, *item): item for item in companies_to_test}
        for fut in as_completed(futures):
            res = fut.result()
            comp_results[res["row_idx"]] = res
            done_c += 1
            if done_c % 50 == 0 or done_c == len(companies_to_test):
                print(f"  [{done_c:>3}/{len(companies_to_test)}] Tested {res['company'][:25]:<25} | HTTP {res['http_code']} | {res['rec_strategy'][:30]}")

    print(f"Company verification completed in {time.time() - t0:.1f}s")

    # Write Headers for Companies Tab
    comp_headers = [
        ("Company Name", 28, fill_navy),
        ("Domain", 22, fill_navy),
        ("Newsroom Path / Custom Override", 28, fill_navy),
        ("Baseline Strategy", 26, fill_navy),
        ("Baseline Notes", 32, fill_navy),
        ("Live HTTP Code", 14, fill_blue),
        ("Response Time (s)", 16, fill_blue),
        ("Verified Newsroom URL", 55, fill_blue),
        ("Recommended Primary Strategy", 32, fill_blue),
        ("Native RSS Detected", 30, fill_blue),
        ("XML Sitemap Available", 30, fill_blue),
        ("Google News Fallback Query", 55, fill_blue),
        ("Last Detected Headline / Release", 50, fill_blue),
        ("Operational Observations & CI Notes", 65, fill_crimson),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(comp_headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_comp.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_comp.column_dimensions[col_letter].width = width
    ws_comp.row_dimensions[1].height = 28

    # Write Data for Companies Tab
    for r in range(2, ws_comp.max_row + 1):
        c_res = comp_results.get(r)
        if not c_res:
            continue

        code = c_res["http_code"]
        if code == 200:
            res_fill = fill_ok
        elif code in (401, 403):
            res_fill = fill_blocked
        elif code == 404:
            res_fill = fill_failed
        else:
            res_fill = fill_recovered

        # Col A-E: Existing
        for col_i in range(1, 6):
            cell = ws_comp.cell(row=r, column=col_i)
            cell.font = font_data
            cell.border = thin_border
            if col_i == 2:
                cell.font = font_code

        # Col F: Live HTTP Code
        cell_f = ws_comp.cell(row=r, column=6, value=c_res["http_code"])
        cell_f.font = font_data
        cell_f.fill = res_fill
        cell_f.border = thin_border
        cell_f.alignment = Alignment(horizontal="center")

        # Col G: Response Time
        cell_g = ws_comp.cell(row=r, column=7, value=c_res["elapsed"])
        cell_g.font = font_data
        cell_g.fill = res_fill
        cell_g.border = thin_border
        cell_g.alignment = Alignment(horizontal="right")

        # Col H: Verified Newsroom URL
        cell_h = ws_comp.cell(row=r, column=8, value=c_res["verified_url"])
        cell_h.font = font_link
        cell_h.fill = res_fill
        cell_h.border = thin_border
        cell_h.hyperlink = c_res["verified_url"]

        # Col I: Recommended Primary Strategy
        cell_i = ws_comp.cell(row=r, column=9, value=c_res["rec_strategy"])
        cell_i.font = font_data
        cell_i.fill = res_fill
        cell_i.border = thin_border

        # Col J: Native RSS Detected
        cell_j = ws_comp.cell(row=r, column=10, value=c_res["has_rss"])
        cell_j.font = font_code
        cell_j.fill = res_fill
        cell_j.border = thin_border

        # Col K: XML Sitemap Available
        cell_k = ws_comp.cell(row=r, column=11, value=c_res["has_sitemap"])
        cell_k.font = font_code
        cell_k.fill = res_fill
        cell_k.border = thin_border

        # Col L: Google News Fallback Query
        cell_l = ws_comp.cell(row=r, column=12, value=c_res["gnews_query"])
        cell_l.font = font_link
        cell_l.fill = res_fill
        cell_l.border = thin_border
        cell_l.hyperlink = c_res["gnews_query"]

        # Col M: Last Detected Headline
        cell_m = ws_comp.cell(row=r, column=13, value=c_res["latest_headline"])
        cell_m.font = font_title
        cell_m.fill = res_fill
        cell_m.border = thin_border
        cell_m.alignment = Alignment(wrap_text=True)

        # Col N: Operational Observations & CI Notes
        cell_n = ws_comp.cell(row=r, column=14, value=c_res["observations"])
        cell_n.font = font_flag
        cell_n.border = thin_border
        cell_n.alignment = Alignment(wrap_text=True)
        if "WAF" in c_res["observations"]:
            cell_n.fill = fill_blocked
        elif "Native RSS" in c_res["observations"]:
            cell_n.fill = fill_ok
        elif "Sitemap" in c_res["observations"]:
            cell_n.fill = fill_note
        else:
            cell_n.fill = fill_white

        ws_comp.row_dimensions[r].height = 22

    ws_comp.freeze_panes = "A2"
    ws_comp.auto_filter.ref = f"A1:N{ws_comp.max_row}"

    # -------------------------------------------------------------------------
    # STEP 3: Live Verification for 05_Indications_Radar
    # -------------------------------------------------------------------------
    ind_tab_name = [n for n in wb.sheetnames if "Indications" in n][0]
    ws_ind = wb[ind_tab_name]
    print(f"\n--- STEP 3: Verifying {ws_ind.max_row - 1} Indications in '{ind_tab_name}' ---")

    indications_to_test = []
    for r in range(2, ws_ind.max_row + 1):
        ind = str(ws_ind.cell(row=r, column=1).value or "")
        syn = str(ws_ind.cell(row=r, column=2).value or "")
        ct_c = str(ws_ind.cell(row=r, column=3).value or "")
        excl = str(ws_ind.cell(row=r, column=4).value or "")
        b_q = str(ws_ind.cell(row=r, column=5).value or "")
        c_q = str(ws_ind.cell(row=r, column=6).value or "")
        ct_e = str(ws_ind.cell(row=r, column=7).value or "")
        if ind:
            indications_to_test.append((r, ind, syn, ct_c, excl, b_q, c_q, ct_e))

    print(f"Launching parallel verification across {len(indications_to_test)} indications...")
    ind_results = {}
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(verify_indication, *item): item for item in indications_to_test}
        for fut in as_completed(futures):
            res = fut.result()
            ind_results[res["row_idx"]] = res
            print(f"  Tested Indication: {res['indication'][:22]:<22} | Broad GNews: {res['b_code']} ({res['b_count']} items) | CT.gov: {res['ct_status'][:30]}")

    # Write Headers for Indications Tab
    ind_headers = [
        ("Indication / Theme", 22, fill_navy),
        ("Synonyms & Keywords", 35, fill_navy),
        ("CT.gov Condition Term", 22, fill_navy),
        ("Exclude Terms (Negations)", 22, fill_navy),
        ("Broad Feed Query Logic", 45, fill_navy),
        ("Clinical/Reg Feed Query Logic", 45, fill_navy),
        ("CT.gov Protocol Endpoint", 55, fill_navy),
        ("Broad GNews Status", 16, fill_blue),
        ("Broad GNews Headline", 45, fill_blue),
        ("Clinical/Reg GNews Status", 18, fill_blue),
        ("Clinical/Reg GNews Headline", 45, fill_blue),
        ("CT.gov Live Registry Status", 25, fill_blue),
        ("Latest CT.gov Protocol Registered", 50, fill_blue),
        ("Radar Scope & CI Observations", 55, fill_crimson),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(ind_headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws_ind.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_ind.column_dimensions[col_letter].width = width
    ws_ind.row_dimensions[1].height = 28

    # Write Data for Indications Tab
    for r in range(2, ws_ind.max_row + 1):
        i_res = ind_results.get(r)
        if not i_res:
            continue

        # Col A-G: Existing styling
        for col_i in range(1, 8):
            cell = ws_ind.cell(row=r, column=col_i)
            cell.font = font_data
            cell.border = thin_border
            if col_i == 7 and str(cell.value or "").startswith("http"):
                cell.font = font_link
                cell.hyperlink = str(cell.value)

        # Col H: Broad GNews Status
        cell_h = ws_ind.cell(row=r, column=8, value=f"{i_res['b_code']} OK ({i_res['b_count']} items)" if i_res['b_code'] == 200 else f"HTTP {i_res['b_code']}")
        cell_h.font = font_data
        cell_h.fill = fill_ok if i_res['b_code'] == 200 else fill_failed
        cell_h.border = thin_border
        cell_h.alignment = Alignment(horizontal="center")

        # Col I: Broad GNews Headline
        cell_i = ws_ind.cell(row=r, column=9, value=i_res["b_title"])
        cell_i.font = font_title
        cell_i.fill = fill_ok if i_res['b_code'] == 200 else fill_failed
        cell_i.border = thin_border
        cell_i.alignment = Alignment(wrap_text=True)
        if i_res["b_url"]:
            cell_i.hyperlink = i_res["b_url"]

        # Col J: Clinical/Reg GNews Status
        cell_j = ws_ind.cell(row=r, column=10, value=f"{i_res['c_code']} OK ({i_res['c_count']} items)" if i_res['c_code'] == 200 else f"HTTP {i_res['c_code']}")
        cell_j.font = font_data
        cell_j.fill = fill_ok if i_res['c_code'] == 200 else fill_failed
        cell_j.border = thin_border
        cell_j.alignment = Alignment(horizontal="center")

        # Col K: Clinical/Reg GNews Headline
        cell_k = ws_ind.cell(row=r, column=11, value=i_res["c_title"])
        cell_k.font = font_title
        cell_k.fill = fill_ok if i_res['c_code'] == 200 else fill_failed
        cell_k.border = thin_border
        cell_k.alignment = Alignment(wrap_text=True)
        if i_res["c_url"]:
            cell_k.hyperlink = i_res["c_url"]

        # Col L: CT.gov Live Status
        cell_l = ws_ind.cell(row=r, column=12, value=i_res["ct_status"])
        cell_l.font = font_data
        cell_l.fill = fill_ok if "200 OK" in i_res["ct_status"] else fill_note if "N/A" in i_res["ct_status"] else fill_failed
        cell_l.border = thin_border
        cell_l.alignment = Alignment(horizontal="center")

        # Col M: Latest CT.gov Protocol
        cell_m = ws_ind.cell(row=r, column=13, value=i_res["ct_latest"])
        cell_m.font = font_code
        cell_m.fill = fill_ok if "200 OK" in i_res["ct_status"] else fill_note if "N/A" in i_res["ct_status"] else fill_failed
        cell_m.border = thin_border
        cell_m.alignment = Alignment(wrap_text=True)

        # Col N: Observations
        cell_n = ws_ind.cell(row=r, column=14, value=i_res["observations"])
        cell_n.font = font_flag
        cell_n.fill = fill_white
        cell_n.border = thin_border
        cell_n.alignment = Alignment(wrap_text=True)

        ws_ind.row_dimensions[r].height = 24

    ws_ind.freeze_panes = "A2"
    ws_ind.auto_filter.ref = f"A1:N{ws_ind.max_row}"

    # -------------------------------------------------------------------------
    # STEP 4: Save and Polish
    # -------------------------------------------------------------------------
    print(f"\nSaving final Master Excel workbook to: {XLSX_PATH}...")
    wb.save(XLSX_PATH)
    print("Master Excel workbook successfully updated and verified!")

if __name__ == "__main__":
    run()
