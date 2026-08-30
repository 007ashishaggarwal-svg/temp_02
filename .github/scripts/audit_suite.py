#!/usr/bin/env python3
"""
Deep Forensic Audit Suite for RSSFeedChecker.
Executes empirical verification across Batches 1 to 11.
Extracts hard statistics, distributions, failure points, and edge-case samples.
"""

import os
import sys
import re
import json
import time
import hashlib
import openpyxl
from datetime import datetime, timezone
from urllib.parse import urlparse

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORKSPACE = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

audit_results = {}

# =============================================================================
# BATCH 1: WORKBOOK STRUCTURAL AUDIT
# =============================================================================
print("=== [BATCH 1] WORKBOOK STRUCTURAL AUDIT ===")
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
b1_data = {
    "total_sheets": len(wb.sheetnames),
    "sheet_names": wb.sheetnames,
    "sheet_metrics": {}
}

for sname in wb.sheetnames:
    ws = wb[sname]
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    
    empty_cells_count = 0
    total_data_cells = (ws.max_row - 1) * ws.max_column if ws.max_row > 1 else 0
    
    # Check data types in column 1 to max_col
    col_types = {}
    for c in range(1, ws.max_column + 1):
        c_types = set()
        c_blanks = 0
        for r in range(2, min(ws.max_row + 1, 100)): # sample up to 100
            val = ws.cell(row=r, column=c).value
            if val is None or str(val).strip() == "":
                c_blanks += 1
            else:
                c_types.add(type(val).__name__)
        col_types[headers[c-1] if c-1 < len(headers) else f"Col_{c}"] = {
            "types": list(c_types),
            "sample_blanks": c_blanks
        }

    b1_data["sheet_metrics"][sname] = {
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "data_rows": ws.max_row - 1,
        "headers": headers,
        "column_analysis": col_types
    }

audit_results["batch_1_structure"] = b1_data
print(f"Verified {len(wb.sheetnames)} sheets across {sum(s['max_row'] for s in b1_data['sheet_metrics'].values())} total rows.")

# =============================================================================
# BATCH 2: 00_Unified_Intelligence_Feed FORENSIC AUDIT
# =============================================================================
print("\n=== [BATCH 2] 00_Unified_Intelligence_Feed FORENSIC AUDIT ===")
ws_u = wb["00_Unified_Intelligence_Feed"]
u_rows = []
for r in range(2, ws_u.max_row + 1):
    vals = [ws_u.cell(row=r, column=c).value for c in range(1, 13)]
    u_rows.append({
        "row_idx": r,
        "date": str(vals[0] or ""),
        "source": str(vals[1] or ""),
        "entity": str(vals[2] or ""),
        "indication": str(vals[3] or ""),
        "desk": str(vals[4] or ""),
        "priority": str(vals[5] or ""),
        "headline": str(vals[6] or ""),
        "pillar": str(vals[7] or ""),
        "vector": str(vals[8] or ""),
        "url": str(vals[9] or ""),
        "snippet": str(vals[10] or ""),
        "full_text": str(vals[11] or "")
    })

b2_data = {
    "total_items": len(u_rows),
    "priority_distribution": {},
    "desk_distribution": {},
    "pillar_distribution": {},
    "vector_distribution": {},
    "url_analysis": {
        "google_news_wrapper_count": 0,
        "decoded_authoritative_count": 0,
        "domains": {}
    },
    "text_metrics": {
        "empty_snippet_count": 0,
        "empty_full_text_count": 0,
        "snippet_equals_full_text_count": 0,
        "deep_full_text_extracted_count": 0,
        "avg_snippet_len": 0,
        "avg_full_text_len": 0,
        "js_code_leak_count": 0
    },
    "date_analysis": {
        "invalid_date_count": 0,
        "earliest_date": None,
        "latest_date": None,
        "date_distribution_by_day": {}
    }
}

snippet_lens = []
full_lens = []
dates_parsed = []

for item in u_rows:
    # Priority
    p = item["priority"]
    b2_data["priority_distribution"][p] = b2_data["priority_distribution"].get(p, 0) + 1
    
    # Desk
    d = item["desk"]
    b2_data["desk_distribution"][d] = b2_data["desk_distribution"].get(d, 0) + 1
    
    # Pillar
    pil = item["pillar"]
    b2_data["pillar_distribution"][pil] = b2_data["pillar_distribution"].get(pil, 0) + 1
    
    # Vector
    vec = item["vector"]
    b2_data["vector_distribution"][vec] = b2_data["vector_distribution"].get(vec, 0) + 1
    
    # URL
    u = item["url"]
    if "news.google.com" in u.lower():
        b2_data["url_analysis"]["google_news_wrapper_count"] += 1
    else:
        b2_data["url_analysis"]["decoded_authoritative_count"] += 1
    
    try:
        dom = urlparse(u).netloc.lower()
        if dom:
            b2_data["url_analysis"]["domains"][dom] = b2_data["url_analysis"]["domains"].get(dom, 0) + 1
    except Exception:
        pass

    # Text
    snip = item["snippet"]
    full = item["full_text"]
    if not snip:
        b2_data["text_metrics"]["empty_snippet_count"] += 1
    if not full:
        b2_data["text_metrics"]["empty_full_text_count"] += 1
    if snip and full and snip == full:
        b2_data["text_metrics"]["snippet_equals_full_text_count"] += 1
    elif full and len(full) > len(snip):
        b2_data["text_metrics"]["deep_full_text_extracted_count"] += 1
        
    snippet_lens.append(len(snip))
    full_lens.append(len(full))
    
    # Check JS leak
    if any(k in full for k in ["firstChild", "setAttribute", "function(", "jQuery", "typeof "]):
        b2_data["text_metrics"]["js_code_leak_count"] += 1

    # Date
    d_str = item["date"]
    try:
        dt = datetime.strptime(d_str[:16], "%Y-%m-%d %H:%M")
        dates_parsed.append(dt)
        day_key = d_str[:10]
        b2_data["date_analysis"]["date_distribution_by_day"][day_key] = b2_data["date_analysis"]["date_distribution_by_day"].get(day_key, 0) + 1
    except Exception:
        try:
            dt = datetime.strptime(d_str[:10], "%Y-%m-%d")
            dates_parsed.append(dt)
            day_key = d_str[:10]
            b2_data["date_analysis"]["date_distribution_by_day"][day_key] = b2_data["date_analysis"]["date_distribution_by_day"].get(day_key, 0) + 1
        except Exception:
            b2_data["date_analysis"]["invalid_date_count"] += 1

if snippet_lens:
    b2_data["text_metrics"]["avg_snippet_len"] = round(sum(snippet_lens) / len(snippet_lens), 1)
if full_lens:
    b2_data["text_metrics"]["avg_full_text_len"] = round(sum(full_lens) / len(full_lens), 1)

if dates_parsed:
    dates_parsed.sort()
    b2_data["date_analysis"]["earliest_date"] = dates_parsed[0].strftime("%Y-%m-%d %H:%M")
    b2_data["date_analysis"]["latest_date"] = dates_parsed[-1].strftime("%Y-%m-%d %H:%M")

audit_results["batch_2_unified_feed"] = b2_data
print(f"Total Unified Items: {b2_data['total_items']}")
print(f"Priorities: {b2_data['priority_distribution']}")
print(f"Desks: {b2_data['desk_distribution']}")
print(f"URL Status: {b2_data['url_analysis']['decoded_authoritative_count']} Direct URLs vs {b2_data['url_analysis']['google_news_wrapper_count']} Google News Wrappers")
print(f"JS Leaks Found: {b2_data['text_metrics']['js_code_leak_count']}")

# =============================================================================
# BATCH 3: SOURCE COVERAGE AUDIT
# =============================================================================
print("\n=== [BATCH 3] SOURCE COVERAGE AUDIT ===")
ws_f = wb["03_Feeds_Master (459)"]
ws_c = wb["04_Companies_Master (616)"]
ws_i = wb["05_Indications_Radar (18)"]

feed_cats = {}
feed_urls = set()
dup_feed_urls = []
for r in range(2, ws_f.max_row + 1):
    cat = str(ws_f.cell(row=r, column=4).value or "Uncategorized")
    u = str(ws_f.cell(row=r, column=2).value or "").strip()
    feed_cats[cat] = feed_cats.get(cat, 0) + 1
    if u in feed_urls:
        dup_feed_urls.append((r, u))
    feed_urls.add(u)

comp_strategies = {}
for r in range(2, ws_c.max_row + 1):
    strat = str(ws_c.cell(row=r, column=9).value or "Unassigned")
    comp_strategies[strat] = comp_strategies.get(strat, 0) + 1

b3_data = {
    "total_feeds": ws_f.max_row - 1,
    "feed_categories": feed_cats,
    "unique_feed_urls": len(feed_urls),
    "duplicate_feed_urls_count": len(dup_feed_urls),
    "total_companies": ws_c.max_row - 1,
    "company_strategies": comp_strategies,
    "total_indications": ws_i.max_row - 1
}
audit_results["batch_3_sources"] = b3_data
print(f"Feed Categories: {feed_cats}")
print(f"Company Strategies: {comp_strategies}")

# =============================================================================
# BATCH 4: KEYWORD MATCHER & ADVERSARIAL STRESS TEST
# =============================================================================
print("\n=== [BATCH 4] KEYWORD PRECISION & ADVERSARIAL STRESS TEST ===")
from match_keywords import KeywordMatcher

matcher = KeywordMatcher(XLSX_PATH)

adversarial_test_cases = [
    # 1. False Positive Traps (Should be rejected or tagged General)
    {
        "id": "ADV-01",
        "title": "PureTech Health's Celea Therapeutics to Unveil SURPASS-IPF Phase 3 Trial Design in Pulmonary Fibrosis",
        "snippet": "Studying deupirfenidone in idiopathic pulmonary fibrosis.",
        "expected_entity_not": "Tirzepatide (Eli Lilly)",
        "expected_desk_not": "Metabolic & Obesity Desk",
        "description": "Naked SURPASS acronym colliding with non-pharma pulmonary trial"
    },
    {
        "id": "ADV-02",
        "title": "Endovia Health Sciences, formerly Splash Beverage, Achieves FDA Milestone for CannEpil Veterinary Program",
        "snippet": "Canine epilepsy veterinary development program achieves regulatory step.",
        "expected_is_noise": True,
        "description": "Common English verb 'achieves' matching ACHIEVE trial + veterinary program"
    },
    {
        "id": "ADV-03",
        "title": "Texas Instruments Unveils Ultra-High Speed 16-Bit ADC with Low Noise Payload",
        "snippet": "Analog-to-digital converter designed for high-frequency signal processing.",
        "expected_entity_not": "Antibody-Drug Conjugates (ADC)",
        "description": "Electronics Analog-to-Digital Converter acronym collision"
    },
    {
        "id": "ADV-04",
        "title": "SMA Solar Technology AG Expands Commercial Inverter Production in Germany",
        "snippet": "Solar power inverter manufacturer announces expansion.",
        "expected_entity_not": "Rare Neuromuscular (DMD & SMA)",
        "description": "Solar inverter SMA company name collision"
    },
    {
        "id": "ADV-05",
        "title": "Dr. John Smith, DMD, Opens New Aesthetic Dentistry and Implant Clinic in Boston",
        "snippet": "General dental medicine practice expands services.",
        "expected_entity_not": "Rare Neuromuscular (DMD & SMA)",
        "description": "Dental Doctor of Dental Medicine (DMD) degree collision"
    },
    {
        "id": "ADV-06",
        "title": "Commercial Real Estate Firm Secures Agreement to Acquire 5-Story Office Building in London",
        "snippet": "Definitive agreement to acquire office property for 50 million pounds.",
        "expected_desk_not": "Business Development & M&A Desk",
        "description": "Real estate property acquisition matching generic M&A rule"
    },
    # 2. Genuine Positive Tests (Must match high tier)
    {
        "id": "ADV-07",
        "title": "Eli Lilly Announces Positive Topline Results from SURPASS-4 Phase 3 Trial for Tirzepatide in Type 2 Diabetes",
        "snippet": "Tirzepatide demonstrated significant HbA1c reduction and weight loss.",
        "expected_entity": "Tirzepatide (Eli Lilly)",
        "expected_priority": "🔴 Tier 1 (Urgent / Immediate)",
        "expected_desk": "Metabolic & Obesity Desk",
        "description": "True Eli Lilly Tirzepatide SURPASS-4 Phase 3 trial readout"
    },
    {
        "id": "ADV-08",
        "title": "FDA Approves Merck's Keytruda (pembrolizumab) in Combination with Chemotherapy for First-Line Advanced Gastric Cancer",
        "snippet": "FDA marketing authorization granted following Phase 3 KEYNOTE-522 trial.",
        "expected_entity": "Keytruda (Merck & Co)",
        "expected_priority": "🔴 Tier 1 (Urgent / Immediate)",
        "expected_desk": "Oncology & Immuno-Oncology Desk",
        "description": "True Merck Keytruda FDA approval"
    },
    {
        "id": "ADV-09",
        "title": "AstraZeneca and Daiichi Sankyo Submit sBLA for Enhertu (T-DXd) in HER2-Ultralow Metastatic Breast Cancer Following DESTINY-Breast06",
        "snippet": "Phase 3 trial demonstrated progression-free survival benefit in HER2-ultralow patients.",
        "expected_entity": "Enhertu (Daiichi Sankyo / AstraZeneca)",
        "expected_priority": "🔴 Tier 1 (Urgent / Immediate)",
        "expected_desk": "Oncology & Immuno-Oncology Desk",
        "description": "True Enhertu DESTINY-Breast06 sBLA filing"
    },
    {
        "id": "ADV-10",
        "title": "FDA Places Full Clinical Hold on Phase 1 Solid Tumor Trial Due to Unanticipated Neurotoxicity",
        "snippet": "Agency suspended patient dosing pending safety investigation into Grade 4 adverse events.",
        "expected_entity": "Clinical Holds & Safety Alerts",
        "expected_priority": "🔴 Tier 1 (Urgent / Immediate)",
        "expected_desk": "Regulatory & Strategy Desk",
        "description": "True FDA clinical hold regulatory catalyst"
    }
]

adversarial_results = []
for tc in adversarial_test_cases:
    m_res = matcher.match(tc["title"], tc.get("snippet", ""))
    passed = True
    reasons = []
    
    if tc.get("expected_is_noise"):
        if not m_res.get("is_noise"):
            passed = False
            reasons.append(f"Failed to flag as noise (got: {m_res.get('matches')})")
    if tc.get("expected_entity_not"):
        if tc["expected_entity_not"] in m_res.get("matches", []):
            passed = False
            reasons.append(f"Falsely matched entity '{tc['expected_entity_not']}'")
    if tc.get("expected_desk_not"):
        if m_res.get("assigned_desk") == tc["expected_desk_not"]:
            passed = False
            reasons.append(f"Falsely routed to desk '{tc['expected_desk_not']}'")
    if tc.get("expected_entity"):
        if tc["expected_entity"] not in m_res.get("matches", []):
            passed = False
            reasons.append(f"Missed expected entity '{tc['expected_entity']}' (got: {m_res.get('matches')})")
    if tc.get("expected_priority"):
        if tc["expected_priority"] != m_res.get("top_priority"):
            passed = False
            reasons.append(f"Wrong priority tier (expected {tc['expected_priority']}, got {m_res.get('top_priority')})")
    if tc.get("expected_desk"):
        if tc["expected_desk"] != m_res.get("assigned_desk"):
            passed = False
            reasons.append(f"Wrong desk (expected {tc['expected_desk']}, got {m_res.get('assigned_desk')})")

    adversarial_results.append({
        "id": tc["id"],
        "description": tc["description"],
        "passed": passed,
        "reasons": reasons,
        "matched_entities": m_res.get("matches", []),
        "assigned_desk": m_res.get("assigned_desk"),
        "priority": m_res.get("top_priority"),
        "is_noise": m_res.get("is_noise")
    })

b4_data = {
    "total_adversarial_tests": len(adversarial_test_cases),
    "passed_tests": sum(1 for r in adversarial_results if r["passed"]),
    "failed_tests": sum(1 for r in adversarial_results if not r["passed"]),
    "details": adversarial_results
}
audit_results["batch_4_keywords"] = b4_data
print(f"Adversarial Matcher Pass Rate: {b4_data['passed_tests']}/{b4_data['total_adversarial_tests']} ({b4_data['passed_tests']/b4_data['total_adversarial_tests']*100:.0f}%)")

# =============================================================================
# BATCH 5: SMART DEDUPLICATOR ADVERSARIAL AUDIT
# =============================================================================
print("\n=== [BATCH 5] SMART DEDUPLICATOR AUDIT ===")
from run_unified_intelligence_pipeline import SmartDeduplicator

dedup_test_cases = [
    # Scenario 1: Exact Headline with Publisher Suffix Variant (Must Merge)
    (
        {
            "headline": "Pfizer Reports Positive Phase 3 Data for Dual Incretin Agonist - STAT News",
            "url": "https://news.google.com/rss/articles/CBMi12345",
            "source": "STAT News",
            "vector": "Google News Query",
            "full_text": "Short snippet"
        },
        {
            "headline": "Pfizer Reports Positive Phase 3 Data for Dual Incretin Agonist",
            "url": "https://www.statnews.com/2026/08/25/pfizer-incretin-data",
            "source": "STAT News (Native)",
            "vector": "Native RSS",
            "full_text": "Full deep paragraph text from native article..."
        },
        "MUST_MERGE",
        "Exact headline with publisher suffix vs native link"
    ),
    # Scenario 2: Two Completely Different Headlines on Different Drugs (Must NOT Merge)
    (
        {
            "headline": "FDA Approves Keytruda for First-Line Gastric Cancer",
            "url": "https://www.fda.gov/news/keytruda-approval",
            "source": "FDA",
            "vector": "Native RSS",
            "full_text": "FDA approval details for Keytruda."
        },
        {
            "headline": "FDA Approves Enhertu for HER2-Ultralow Breast Cancer",
            "url": "https://www.fda.gov/news/enhertu-approval",
            "source": "FDA",
            "vector": "Native RSS",
            "full_text": "FDA approval details for Enhertu."
        },
        "MUST_NOT_MERGE",
        "Two distinct FDA approvals on different drugs"
    ),
    # Scenario 3: Same Company, Different Trial Outcomes (Must NOT Merge)
    (
        {
            "headline": "Novo Nordisk Announces Positive Topline Results for CagriSema in REDEFINE-1",
            "url": "https://www.novonordisk.com/news/redefine-1-success",
            "source": "Novo Nordisk",
            "vector": "Native RSS",
            "full_text": "REDEFINE-1 met all primary endpoints."
        },
        {
            "headline": "Novo Nordisk Announces Phase 3 Trial Readout for Oral Semaglutide in PIONEER-12",
            "url": "https://www.novonordisk.com/news/pioneer-12-results",
            "source": "Novo Nordisk",
            "vector": "Native RSS",
            "full_text": "PIONEER-12 results announced."
        },
        "MUST_NOT_MERGE",
        "Same company with two different clinical trial readouts"
    )
]

dedup_results = []
for item_a, item_b, expected_action, desc in dedup_test_cases:
    s_dedup = SmartDeduplicator()
    dup_a, obj_a = s_dedup.is_duplicate_or_merge(item_a)
    dup_b, obj_b = s_dedup.is_duplicate_or_merge(item_b)
    
    merged = dup_b
    passed = (merged and expected_action == "MUST_MERGE") or (not merged and expected_action == "MUST_NOT_MERGE")
    
    dedup_results.append({
        "description": desc,
        "expected": expected_action,
        "actual_merged": merged,
        "passed": passed,
        "final_url": obj_b.get("url") if merged else None
    })

b5_data = {
    "total_dedup_tests": len(dedup_test_cases),
    "passed_tests": sum(1 for r in dedup_results if r["passed"]),
    "details": dedup_results
}
audit_results["batch_5_deduplication"] = b5_data
print(f"SmartDeduplicator Pass Rate: {b5_data['passed_tests']}/{b5_data['total_dedup_tests']}")

# =============================================================================
# SAVE JSON EVIDENCE AUDIT
# =============================================================================
evidence_path = os.path.join(WORKSPACE, "audit_evidence_ledger.json")
with open(evidence_path, "w", encoding="utf-8") as f:
    json.dump(audit_results, f, indent=2, default=str)

print(f"\n✨ Empirical Forensic Audit Complete! Saved evidence to: {evidence_path}")
