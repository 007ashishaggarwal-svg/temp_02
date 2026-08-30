#!/usr/bin/env python3
"""
Re-categorize all feeds in '03_Feeds_Master (518)' with accurate,
CI-relevant categories based on URL patterns, labels, and domain knowledge.
"""

import sys
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"
SHEET = "03_Feeds_Master (518)"

# -----------------------------------------------------------------------
# CATEGORY DEFINITIONS (from a Competitive Intelligence lens)
# -----------------------------------------------------------------------
# 1. "Company IR / Press Release"  — Direct investor relations or corporate press release feeds
# 2. "Trade Press"                 — Industry trade publications (STAT, Endpoints, BioWorld, Fierce, etc.)
# 3. "Google News Aggregation"     — Google News RSS search feeds
# 4. "Medical Journal"             — Peer-reviewed journal feeds (Nature, Lancet, NEJM, JAMA, Cell, etc.)
# 5. "Regulatory & Government"     — FDA, EMA, MHRA, SEC, CDC, WHO
# 6. "Newswire"                    — PRNewswire, GlobeNewswire, BusinessWire
# 7. "Preprint Server"             — medRxiv, bioRxiv
# 8. "Patient Advocacy & Disease"  — Disease-focused non-profit / patient advocacy orgs
# 9. "Industry News & Analysis"    — Multi-source industry news aggregators (MedPage, Drugs.com, etc.)
# 10. "CDMO / CRO / Services"     — Contract research, manufacturing, lab services
# -----------------------------------------------------------------------

# --- Large Pharma / known Big Pharma companies whose feeds are IR/corporate ---
BIG_PHARMA_DOMAINS = {
    "lilly.com", "abbvie.com", "pfizer.com", "roche.com", "amgen.com",
    "astrazeneca.com", "gsk.com", "sanofi.us", "sanofi.com", "novartis.com",
    "biogen.com", "regeneron.com", "bms.com", "merck.com", "gilead.com",
    "modernatx.com", "incyte.com", "vertex.com", "vrtx.com",
    "sunpharma.com", "biocon.com", "ipsen.com", "sobi.com",
    "neurocrine.gcs-web.com", "eisai.com", "eapharma.co.jp",
    "ono-pharma.com", "teijin.com", "orix.co.jp",
}

# --- Regulatory / Government domains ---
REGULATORY_DOMAINS = {
    "fda.gov", "ema.europa.eu", "gov.uk", "sec.gov", "cdc.gov", "who.int",
    "g-ba.de",
}

# --- Trade Press domains ---
TRADE_PRESS_DOMAINS = {
    "statnews.com", "endpoints.news", "endpts.com", "bioworld.com",
    "fiercepharma.com", "fiercebiotech.com", "fiercehealthcare.com",
    "biospace.com", "genengnews.com", "labiotech.eu",
    "pharmexec.com", "pharmaphorum.com", "pharmafile.com",
    "pharmatimes.com", "pharmavoice.com", "healthcaredive.com",
    "biopharmadive.com", "medtechdive.com", "medcitynews.com",
    "drugdiscoverytrends.com", "pharmaceuticalprocessingworld.com",
    "worldpharmanews.com", "insideprecisionmedicine.com",
    "neurologylive.com", "hemostasistoday.com",
    "citeline.com", "businesskorea.co.kr",
}

# --- Medical Journal domains ---
JOURNAL_DOMAINS = {
    "nature.com", "nejm.org", "thelancet.com", "jamanetwork.com",
    "cell.com", "annalsofoncology.org", "ascopubs.org",
    "medrxiv.org", "biorxiv.org",
}

# --- Preprint server domains ---
PREPRINT_DOMAINS = {
    "medrxiv.org", "biorxiv.org",
}

# --- Industry News / Aggregator domains ---
INDUSTRY_NEWS_DOMAINS = {
    "medpagetoday.com", "drugs.com", "sciencedaily.com",
    "forbes.com", "icer.org",
}

# --- Patient Advocacy / Disease-focused orgs ---
PATIENT_ADVOCACY_KEYWORDS = [
    "newstoday.com", "cureduchenne", "parentprojectmd", "curesma",
    "famigliesma", "smauk.org", "smabenimleyuru", "ehdn.org",
    "hdsa.org", "hdbuzz.net", "wfh.org", "eahad.org",
    "f-sma.ru", "fsma.pl", "esperare.org", "thebionews.net",
]

# --- CDMO / CRO / Services ---
CDMO_CRO_KEYWORDS = [
    "catalent", "criver.com", "charles river", "lonza", "siegfried",
    "bachem", "evonik", "recipharm", "theemmesgroup",
]

# --- Newswire domains ---
NEWSWIRE_DOMAINS = {
    "prnewswire.com", "globenewswire.com", "businesswire.com",
}

def get_domain(url):
    """Extract domain from URL."""
    m = re.search(r"https?://(?:www\.)?([^/]+)", url)
    return m.group(1).lower() if m else ""

def categorize(feed_id, url, label, current_cat):
    """Determine the correct category for a feed."""
    url_lower = url.lower()
    label_lower = label.lower()
    domain = get_domain(url)
    
    # 1. Google News feeds
    if "news.google.com" in url_lower:
        return "Google News Aggregation"
    
    # 2. Preprint servers (before journals since they share some patterns)
    if any(d in domain for d in PREPRINT_DOMAINS):
        return "Preprint Server"
    
    # 3. Medical Journals
    if any(d in domain for d in JOURNAL_DOMAINS):
        return "Medical Journal"
    
    # 4. Regulatory & Government
    if any(d in domain for d in REGULATORY_DOMAINS):
        return "Regulatory & Government"
    if "sec.gov" in url_lower or "data.sec.gov" in url_lower:
        return "Regulatory & Government"
    
    # 5. Newswires
    if any(d in domain for d in NEWSWIRE_DOMAINS):
        return "Newswire"
    
    # 6. Trade Press
    if any(d in domain for d in TRADE_PRESS_DOMAINS):
        return "Trade Press"
    
    # 7. Patient Advocacy & Disease-focused
    if any(kw in url_lower or kw in label_lower for kw in PATIENT_ADVOCACY_KEYWORDS):
        return "Patient Advocacy & Disease"
    
    # 8. CDMO / CRO / Services
    if any(kw in url_lower or kw in label_lower for kw in CDMO_CRO_KEYWORDS):
        return "CDMO / CRO / Services"
    
    # 9. Industry News Aggregators
    if any(d in domain for d in INDUSTRY_NEWS_DOMAINS):
        return "Industry News & Analysis"
    
    # 10. Company IR / Press Release detection
    # Key signals: "investor", "ir.", "/rss.xml", company-specific domain with press/investor feed
    ir_signals = [
        "investor" in url_lower,
        "ir." in domain,
        "/press-releases" in url_lower,
        "/news-events" in url_lower,
        "/news-releases" in url_lower,
        "gcs-web.com" in domain,
        "-official" in label_lower,
    ]
    if any(ir_signals):
        return "Company IR / Press Release"
    
    # Check if it's a known big pharma domain
    if any(d in domain for d in BIG_PHARMA_DOMAINS):
        return "Company IR / Press Release"
    
    # If label ends with "-Official" pattern or URL is a company feed
    if label_lower.endswith("-official"):
        return "Company IR / Press Release"
    
    # Detect company feeds by common URL patterns
    company_url_patterns = [
        r"^https?://[^/]*\.(com|co\.kr|bio|co\.jp|eu|ch|nl|com\.au)/feed/?$",
        r"^https?://[^/]*\.(com|co\.kr|bio|co\.jp|eu|ch|nl|com\.au)/rss\.xml$",
        r"^https?://[^/]*\.(com|co\.kr|bio|co\.jp|eu|ch|nl|com\.au)/rss/?$",
        r"/sitemap\.rss$",
        r"/blog-feed\.xml$",
        r"/news/feed/?$",
        r"/news\?format=rss$",
        r"/feed/rss2$",
    ]
    if any(re.search(pat, url_lower) for pat in company_url_patterns):
        # But exclude known non-company domains
        non_company = TRADE_PRESS_DOMAINS | JOURNAL_DOMAINS | INDUSTRY_NEWS_DOMAINS
        if not any(d in domain for d in non_company):
            return "Company IR / Press Release"
    
    # Default: keep current or mark as Industry News & Analysis
    return current_cat if current_cat and current_cat != "Industry News" else "Industry News & Analysis"


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    
    changes = 0
    category_counts = {}
    
    for row in range(2, ws.max_row + 1):
        feed_id = str(ws.cell(row=row, column=1).value or "")
        url = str(ws.cell(row=row, column=2).value or "")
        label = str(ws.cell(row=row, column=3).value or "")
        old_cat = str(ws.cell(row=row, column=4).value or "")
        
        new_cat = categorize(feed_id, url, label, old_cat)
        
        category_counts[new_cat] = category_counts.get(new_cat, 0) + 1
        
        if new_cat != old_cat:
            ws.cell(row=row, column=4, value=new_cat)
            changes += 1
            print(f"  Row {row}: {feed_id} | {label[:30]:<30} | {old_cat:<30} -> {new_cat}")
    
    print(f"\n{'='*78}")
    print(f"TOTAL CHANGES: {changes} feeds re-categorized")
    print(f"{'='*78}")
    print("\nNew Category Distribution:")
    for cat, count in sorted(category_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat:<35} {count:>4}")
    
    # Save
    try:
        wb.save(XLSX)
        print(f"\nSaved to: {XLSX}")
    except PermissionError:
        alt = XLSX.replace(".xlsx", "_recategorized.xlsx")
        wb.save(alt)
        print(f"\nExcel open! Saved to: {alt}")


if __name__ == "__main__":
    main()
