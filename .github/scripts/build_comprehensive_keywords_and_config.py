#!/usr/bin/env python3
"""
Build Comprehensive '02_Keywords_and_Rules' and '03_Config_and_Settings' Tabs
=============================================================================
Features:
1. 15 Detailed Therapeutic Projects from Match logic.xlsx with Primary Identifiers,
   Secondary Drug/Modality Whitelists, and Negative Preclinical/Animal Rejectors.
2. Signal Type Classification Matrix (clinical_pos, clinical_neg, regulatory, corporate, leadership_change, commercial, general).
3. Configurable Scoring & Relevance Parameters (Title Match, Phase 3, Approvals, Thresholds).
4. Event Clustering Configuration (Cluster ID, Entity Hints, Time Window).
5. Clean User-Tweakable Layout with clear definitions.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")
ALT_MATCH_PATH = os.path.join(WORKSPACE, "Alternative", "Match logic.xlsx")

def build_keywords_and_config():
    print("=" * 90)
    print(" 🌟 BUILDING COMPREHENSIVE KEYWORDS, SCORING & CONFIGURATION MATRICES")
    print("=" * 90)

    wb = openpyxl.load_workbook(XLSX_PATH)

    # -------------------------------------------------------------------------
    # TAB 02: 02_Keywords_and_Rules
    # -------------------------------------------------------------------------
    TAB_K = "02_Keywords_and_Rules"
    if TAB_K in wb.sheetnames:
        del wb[TAB_K]
    ws_k = wb.create_sheet(title=TAB_K, index=2)
    ws_k.views.sheetView[0].showGridLines = True

    # Styling Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    PURPLE_DARK = "382D5C"
    TEAL_DARK = "0E5A5E"
    GREEN_DARK = "155724"
    WHITE = "FFFFFF"
    BORDER_COLOR = "D0D7DE"
    ICE_BLUE = "F0F5FA"
    LIGHT_GRAY = "F8F9FA"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_purple = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
    fill_teal = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
    fill_green_dark = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
    fill_ice_blue = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_gray_section = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_section_header = Font(name="Calibri", size=12, bold=True, color=NAVY_DARK)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_code = Font(name="Consolas", size=9, color="333333")

    fill_green_badge = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_green_badge = Font(name="Calibri", size=10, bold=True, color="155724")

    fill_red_badge = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_red_badge = Font(name="Calibri", size=10, bold=True, color="721C24")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # 1. Load Projects from Match logic.xlsx
    wb_alt_m = openpyxl.load_workbook(ALT_MATCH_PATH, data_only=True)
    ws_alt_m = wb_alt_m.active
    
    projects_data = []
    desk_mapping = {
        "Alzheimer": "Neuroscience & Neurology Desk",
        "Obesity": "Metabolic & Obesity Desk",
        "Oncology": "Oncology & Immuno-Oncology Desk",
        "Atopic Dermatitis": "Immunology & Respiratory Desk",
        "Schizophrenia": "Neuroscience & Neurology Desk",
        "Parkinson MSA": "Neuroscience & Neurology Desk",
        "Myeloma": "Oncology & Immuno-Oncology Desk",
        "NSCLC": "Oncology & Immuno-Oncology Desk",
        "Breast Cancer": "Oncology & Immuno-Oncology Desk",
        "IBD": "Immunology & Respiratory Desk",
        "MASH": "Metabolic & Obesity Desk",
        "Lupus": "Immunology & Respiratory Desk",
        "Small Cell Lung Cancer": "Oncology & Immuno-Oncology Desk"
    }

    for r in range(2, ws_alt_m.max_row + 1):
        p_name = str(ws_alt_m.cell(row=r, column=1).value or "").strip()
        p_prim = str(ws_alt_m.cell(row=r, column=2).value or "").strip()
        p_sec = str(ws_alt_m.cell(row=r, column=3).value or "").strip()
        p_neg = str(ws_alt_m.cell(row=r, column=4).value or "").strip()

        if not p_name:
            continue

        if not p_neg:
            p_neg = "mouse model, rat model, murine, in vitro, cell culture, cell line, xenograft, zebrafish, drosophila, caenorhabditis, preclinical only, review article, systematic review, meta analysis, case report, editorial, letter to editor"

        desk = desk_mapping.get(p_name, "Executive Briefing Desk")
        priority = "🔴 Tier 1 (Urgent)" if p_name in ["Obesity", "Oncology", "NSCLC", "Breast Cancer", "Alzheimer"] else "🟡 Tier 2 (Daily)"

        projects_data.append((p_name, p_prim, p_sec, p_neg, desk, priority, "Active"))

    # Add Rare Disease & General Biopharma projects
    projects_data.append((
        "Rare Diseases",
        "spinal muscular atrophy,SMA,Duchenne muscular dystrophy,DMD,sickle cell disease,SCD,beta thalassemia,cystic fibrosis,huntington disease,amyotrophic lateral sclerosis,ALS",
        "Casgevy,Spinraza,Evrysdi,Elevidys,Zolgensma,gene therapy,CRISPR,AAV,exon skipping,Sarepta,Vertex,Biogen,Ionis,Novartis",
        "mouse model, rat model, murine, in vitro, cell culture, xenograft, preclinical only, review article, case report",
        "Rare Disease Desk",
        "🔴 Tier 1 (Urgent)",
        "Active"
    ))
    projects_data.append((
        "General Biopharma & M&A",
        "biopharma,pharmaceuticals,biotech,FDA,EMA,CHMP,PDUFA,NDA,BLA,CRL,clinical trial,Phase 1,Phase 2,Phase 3",
        "acquisition,merger,buyout,licensing agreement,partnership,deal,asset purchase,restructuring,topline results,primary endpoint",
        "class action lawsuit,shareholder litigation,stock split,options trading,motley fool,zacks,simply wall st,veterinary,canine,feline",
        "Business Development & M&A Desk",
        "🟡 Tier 2 (Daily)",
        "Active"
    ))

    # Write Table 1: Therapeutic Projects & Asset Match Logic
    headers_t1 = [
        ("Project / Indication Theme", 26, fill_navy),
        ("Primary Inclusion Identifiers (Synonyms / Receptors)", 45, fill_navy),
        ("Secondary Whitelist (Drugs / Modalities / Sponsors)", 50, fill_blue),
        ("Negative Exclusion Rejectors (Preclinical / Non-Human / Animals)", 45, fill_teal),
        ("Assigned CI Desk", 28, fill_purple),
        ("Base Priority Tier", 20, fill_green_dark),
        ("Active Toggle", 16, fill_green_dark)
    ]

    for c_idx, (h_name, width, h_fill) in enumerate(headers_t1, start=1):
        col_letter = get_column_letter(c_idx)
        cell = ws_k.cell(row=1, column=c_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_k.column_dimensions[col_letter].width = width
    ws_k.row_dimensions[1].height = 30

    for r_idx, proj in enumerate(projects_data, start=2):
        row_fill = fill_ice_blue if (r_idx % 2 == 0) else PatternFill(fill_type=None)
        
        ws_k.cell(row=r_idx, column=1, value=proj[0]).font = font_bold
        ws_k.cell(row=r_idx, column=2, value=proj[1]).font = font_code
        ws_k.cell(row=r_idx, column=3, value=proj[2]).font = font_code
        ws_k.cell(row=r_idx, column=4, value=proj[3]).font = font_code
        ws_k.cell(row=r_idx, column=5, value=proj[4]).font = font_data
        
        c_p = ws_k.cell(row=r_idx, column=6, value=proj[5])
        c_p.font = font_bold
        c_p.alignment = Alignment(horizontal="center")
        if "Tier 1" in proj[5]:
            c_p.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            c_p.font = Font(name="Calibri", size=10, bold=True, color="721C24")
        else:
            c_p.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            c_p.font = Font(name="Calibri", size=10, bold=True, color="856404")

        c_act = ws_k.cell(row=r_idx, column=7, value=proj[6])
        c_act.fill = fill_green_badge
        c_act.font = font_green_badge
        c_act.alignment = Alignment(horizontal="center")

        for c_i in range(1, 8):
            cell_c = ws_k.cell(row=r_idx, column=c_i)
            cell_c.border = thin_border
            if cell_c.fill.fill_type is None and row_fill.fill_type is not None:
                cell_c.fill = row_fill
        ws_k.row_dimensions[r_idx].height = 24

    ws_k.freeze_panes = "A2"
    ws_k.auto_filter.ref = f"A1:G{len(projects_data)+1}"

    # -------------------------------------------------------------------------
    # TAB 03: 03_Config_and_Settings (Expanded with Signal Types & Scoring)
    # -------------------------------------------------------------------------
    TAB_C = "03_Config_and_Settings"
    if TAB_C in wb.sheetnames:
        del wb[TAB_C]
    ws_c = wb.create_sheet(title=TAB_C, index=3)
    ws_c.views.sheetView[0].showGridLines = True

    # Section 1: Signal Type Classification Table
    ws_c.cell(row=1, column=1, value="1. SIGNAL TYPE CLASSIFICATION & DETECTION RULES").font = font_section_header
    ws_c.row_dimensions[1].height = 25

    signal_headers = [
        ("Signal Type Code", 20, fill_navy),
        ("Signal Category Name", 26, fill_navy),
        ("Detection Keywords / Triggers", 45, fill_blue),
        ("Relevance Bonus (+Pts)", 22, fill_teal),
        ("Description & Strategic Impact", 45, fill_purple)
    ]
    for c_idx, (h_name, width, h_fill) in enumerate(signal_headers, start=1):
        col_letter = get_column_letter(c_idx)
        cell = ws_c.cell(row=2, column=c_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_c.column_dimensions[col_letter].width = width
    ws_c.row_dimensions[2].height = 26

    signals_data = [
        ("clinical_pos", "Positive Clinical Trial Readout", "topline,met primary endpoint,positive results,statistically significant,overall survival benefit,PFS improvement,Phase 3 win,Phase 2 success", "+35", "Trial meets primary efficacy endpoint with positive data"),
        ("clinical_neg", "Negative Clinical Readout / Hold", "failed,missed primary endpoint,futility,did not meet,clinical hold,discontinued,adverse event,halted,toxicity", "+30", "Trial failure, clinical hold by agency, or pipeline termination"),
        ("regulatory", "Regulatory Filing / Approval / CRL", "FDA approval,approved by FDA,EMA approval,CHMP positive opinion,BLA accepted,NDA accepted,Priority Review,Breakthrough Therapy,Fast Track,CRL,complete response letter", "+40", "Binding regulatory milestone, marketing authorization or CRL"),
        ("corporate", "M&A, In-Licensing & Partnerships", "acquire,acquisition,to buy,merger,in-licensing,collaboration agreement,strategic partnership,buyout,asset purchase", "+25", "Strategic transaction, asset acquisition, or co-development deal"),
        ("leadership_change", "Executive Leadership Transition", "appoints,named CEO,new CEO,chief medical officer,CSO,leadership shuffle,executive appointment,resigns,steps down", "+15", "C-suite executive appointment or leadership transition"),
        ("commercial", "Commercial Launch & Pricing", "commercial launch,market availability,reimbursement,formulary,NICE recommendation,pricing agreement,label expansion", "+20", "Post-approval commercialization, market access, or label expansion"),
        ("general", "General Biopharma Intelligence", "publishes,presents,conference abstract,pipeline update,preclinical,scientific update", "+10", "General biopharma news and conference presentations")
    ]

    for r_i, sig in enumerate(signals_data, start=3):
        ws_c.cell(row=r_i, column=1, value=sig[0]).font = font_bold
        ws_c.cell(row=r_i, column=2, value=sig[1]).font = font_bold
        ws_c.cell(row=r_i, column=3, value=sig[2]).font = font_code
        ws_c.cell(row=r_i, column=4, value=sig[3]).alignment = Alignment(horizontal="center")
        ws_c.cell(row=r_i, column=4).font = font_bold
        ws_c.cell(row=r_i, column=5, value=sig[4]).font = font_data

        for c_i in range(1, 6):
            ws_c.cell(row=r_i, column=c_i).border = thin_border
        ws_c.row_dimensions[r_i].height = 22

    # Section 2: Configurable Scoring & Relevance Parameters Table
    start_sec2 = len(signals_data) + 5
    ws_c.cell(row=start_sec2, column=1, value="2. RELEVANCE SCORING (0-100) & USER THRESHOLD CONFIGURATION").font = font_section_header
    ws_c.row_dimensions[start_sec2].height = 25

    scoring_headers = [
        ("Scoring Parameter", 30, fill_navy),
        ("Configured Value / Weight", 24, fill_teal),
        ("Default Setting", 20, fill_blue),
        ("Parameter Definition & Impact", 45, fill_purple)
    ]
    for c_idx, (h_name, width, h_fill) in enumerate(scoring_headers, start=1):
        cell = ws_c.cell(row=start_sec2+1, column=c_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_c.row_dimensions[start_sec2+1].height = 26

    scoring_data = [
        ("Title Match Weight", "+35 points", "+35 points", "Keyword / drug / asset match located directly in article headline"),
        ("Pivotal Phase 3 / Approval Match", "+25 points", "+25 points", "Content refers to Phase 3, pivotal study, or regulatory decision"),
        ("Therapeutic Project Asset Whitelist Match", "+20 points", "+20 points", "Matched drug asset listed in Sheet 02 project whitelist"),
        ("Primary Source Authority Bonus", "+10 points", "+10 points", "Source is Primary Newsroom, SEC EDGAR, or Top PR Wire"),
        ("Preclinical / Animal Model Penalty", "-100 points (Discard)", "-100 points", "Mentions mouse model, in vitro, or murine -> instant discard"),
        ("Retail Stock Chatter Penalty", "-100 points (Discard)", "-100 points", "Mentions options trading, Motley Fool, Zacks -> instant discard"),
        ("Tier 1 Relevance Threshold", ">= 80 points", ">= 80 points", "Score threshold required to trigger 🔴 Tier 1 (Urgent)"),
        ("Tier 2 Relevance Threshold", ">= 60 points", ">= 60 points", "Score threshold required to trigger 🟡 Tier 2 (Daily)"),
        ("Tier 3 Relevance Threshold", ">= 40 points", ">= 40 points", "Score threshold required to trigger 🟢 Tier 3 (Weekly)"),
        ("Noise Discard Cutoff", "< 40 points", "< 40 points", "Any article scoring below this threshold is dropped from feed"),
        ("Event Clustering Time Window", "72 hours", "72 hours", "Max time gap to group multi-outlet syndications of the same news"),
        ("Event Clustering Fuzzy Threshold", "0.85 (85%)", "0.85", "Fuzzy token similarity required to merge duplicate headlines")
    ]

    for r_i, sc in enumerate(scoring_data, start=start_sec2+2):
        ws_c.cell(row=r_i, column=1, value=sc[0]).font = font_bold
        ws_c.cell(row=r_i, column=2, value=sc[1]).font = font_bold
        ws_c.cell(row=r_i, column=2).alignment = Alignment(horizontal="center")
        ws_c.cell(row=r_i, column=3, value=sc[2]).alignment = Alignment(horizontal="center")
        ws_c.cell(row=r_i, column=4, value=sc[3]).font = font_data

        for c_i in range(1, 5):
            ws_c.cell(row=r_i, column=c_i).border = thin_border
        ws_c.row_dimensions[r_i].height = 22

    # Save
    for attempt in range(5):
        try:
            wb.save(XLSX_PATH)
            print(f"✅ Successfully updated '02_Keywords_and_Rules' and '03_Config_and_Settings' in {XLSX_PATH}!")
            break
        except PermissionError:
            time.sleep(2)

if __name__ == "__main__":
    build_keywords_and_config()
