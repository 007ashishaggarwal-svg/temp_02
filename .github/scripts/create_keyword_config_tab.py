#!/usr/bin/env python3
"""
Create and format the dedicated '07_Keywords_Match_Config' tab in RSSFeedChecker_Master_Guide_and_Data.xlsx.
Includes rich pre-populated tracking rules across Blockbuster Assets, Next-Gen Modalities,
Regulatory Catalysts, Disease Radar Themes, Corporate M&A Deals, and Universal Noise Negations.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX_PATH = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"
TAB_NAME = "07_Keywords_Match_Config"

CONFIG_RULES = [
    # -------------------------------------------------------------------------
    # TRACK 1: TOP COMPETITOR BLOCKBUSTER ASSETS & PIPELINE CANDIDATES
    # -------------------------------------------------------------------------
    (
        "1. Drug Asset & Pipeline",
        "CagriSema (Novo Nordisk)",
        "CagriSema, cagrilintide, semaglutide combination, NN9535, REDEFINE-1, REDEFINE-2, REDEFINE-3",
        "trial OR phase OR efficacy OR topline OR weight loss OR obesity OR primary endpoint OR FDA",
        "counterfeit, compounding, med spa, weight loss clinic",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Tracks Novo Nordisk's dual amylin/GLP-1 combination against Mounjaro/Zepbound in Phase 3 readouts."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Tirzepatide (Eli Lilly)",
        "tirzepatide, Mounjaro, Zepbound, SURMOUNT, SURPASS, dual GIP/GLP-1",
        "FDA OR approval OR phase 3 OR topline OR label expansion OR trial OR shortage",
        "stock options, options market, class action, coupon code",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Monitors Eli Lilly's market-leading dual GIP/GLP-1 receptor agonist for approvals, trial readouts, and supply."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Orforglipron (Eli Lilly)",
        "orforglipron, LY3502970, oral GLP-1, non-peptide GLP-1, ATTAIN, ACHIEVE",
        "Phase 3 OR trial OR data OR tolerability OR weight loss OR FDA",
        "stock options, lock-up agreement",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Tracks next-gen daily oral non-peptide GLP-1 agonist in Phase 3 global obesity trials."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Keytruda (Merck & Co)",
        "Keytruda, pembrolizumab, MK-3475, anti-PD-1, KEYNOTE",
        "FDA OR approval OR Phase 3 OR OS OR PFS OR overall survival OR CRL OR label expansion",
        "stock options, options market, patent cliff litigation generic",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks oncology blockbuster Keytruda across clinical readouts, regulatory filings, and subcutaneous formulations."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Enhertu (Daiichi Sankyo / AstraZeneca)",
        "Enhertu, trastuzumab deruxtecan, T-DXd, DS-8201, DESTINY-Breast, DESTINY-Lung, DESTINY-PanTumor",
        "FDA OR approval OR Phase 3 OR readout OR HER2-low OR HER2-ultralow OR survival OR safety",
        "stock options, share buy-back",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors breakthrough HER2-directed antibody-drug conjugate (ADC) clinical milestones and line extensions."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Leqembi & Kisunla (Alzheimer's)",
        "Leqembi, lecanemab, BAN2401, Kisunla, donanemab, TRAILBLAZER-ALZ, CLARITY-AD",
        "FDA OR EMA OR approval OR CMS OR coverage OR ARIA OR amyloid OR Phase 3 OR blood biomarker",
        "stock options, Zacks rank",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Neurology & Neurodegenerative Desk",
        "TRUE",
        "Tracks anti-amyloid monoclonal antibodies in Alzheimer's (Eisai/Biogen and Eli Lilly) for approvals, safety (ARIA), and CMS reimbursement."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Dupixent (Sanofi / Regeneron)",
        "Dupixent, dupilumab, anti-IL-4R, anti-IL-13, BOREAS, NOTUS, COPD, atopic dermatitis, asthma",
        "FDA OR approval OR Phase 3 OR label expansion OR sBLA OR PDUFA OR efficacy",
        "stock options, lock-up",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Immunology & Respiratory Desk",
        "TRUE",
        "Monitors blockbuster Type 2 inflammatory biologic for major expansions into COPD, pediatric indications, and competitive threats."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Casgevy & Lyfgenia (CRISPR / Sickle Cell)",
        "Casgevy, exa-cel, exagamglogene autotemcel, Lyfgenia, lovo-cel, CRISPR gene editing, sickle cell, beta-thalassemia",
        "launch OR access OR reimbursement OR safety OR durability OR patient treatment OR commercial",
        "stock options, option volume",
        "🟡 Tier 2 (Daily Briefing)",
        "Cell & Gene Therapy Desk",
        "TRUE",
        "Monitors commercial launch, hospital activation, patient intake, and long-term durability of first approved CRISPR gene editing therapies."
    ),

    # -------------------------------------------------------------------------
    # TRACK 2: NEXT-GEN MODALITIES & MECHANISMS OF ACTION
    # -------------------------------------------------------------------------
    (
        "2. Modality & Biology",
        "Antibody-Drug Conjugates (ADC)",
        "Antibody-Drug Conjugate, ADC, topoisomerase payload, MMAE, DXd, exatecan, bispecific ADC, target-to-toxin",
        "oncology OR solid tumor OR phase OR clinical OR trial OR toxicity OR payload OR linker",
        "analog-to-digital, audio converter, analog digital, ADC software",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks competitive clinical trials and licensing deals in the high-growth ADC landscape while filtering out engineering/electronics noise."
    ),
    (
        "2. Modality & Biology",
        "Targeted Radioligand Therapy (RLT)",
        "radiopharmaceutical, radioligand, lutetium-177, actinium-225, lead-212, copper-64, PSMA, alpha emitter, beta emitter",
        "clinical OR patient OR phase OR trial OR FDA OR supply OR isotope OR cancer",
        "nuclear energy, uranium mining, nuclear power plant",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors targeted radiopharmaceutical pipelines (Novartis, Point, RayzeBio/BMS, Mariana/Lilly) and clinical supply chains."
    ),
    (
        "2. Modality & Biology",
        "Bispecific Antibodies & T-Cell Engagers",
        "bispecific antibody, T-cell engager, BiTE, CD3 bispecific, CD20xCD3, DLL3, BCMAxCD3, trispecific",
        "Phase OR clinical OR trial OR oncology OR hematology OR cytokine release OR CRS OR FDA",
        "computer hardware, electronic components",
        "🟡 Tier 2 (Daily Briefing)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks dual-targeting antibodies in solid tumors and hematology with safety/CRS monitoring."
    ),
    (
        "2. Modality & Biology",
        "In Vivo Cell & Gene Therapy",
        "in vivo CAR-T, in vivo gene editing, LNP delivery, viral vector, AAV capsid, non-viral delivery, mRNA editing",
        "preclinical OR IND OR Phase 1 OR clinical OR delivery OR target OR efficacy",
        "stock options, retail investment",
        "🟡 Tier 2 (Daily Briefing)",
        "Cell & Gene Therapy Desk",
        "TRUE",
        "Tracks cutting-edge in vivo delivery platforms aimed at eliminating ex vivo manufacturing bottlenecks."
    ),
    (
        "2. Modality & Biology",
        "Targeted Protein Degradation (PROTACs)",
        "PROTAC, protein degrader, molecular glue, E3 ligase, cereblon, VHL degrader, targeted degradation",
        "clinical OR Phase OR trial OR oncology OR oral OR degrader OR drug",
        "glue manufacturing, adhesive retail",
        "🟡 Tier 2 (Daily Briefing)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors molecular glues and PROTAC clinical advancements (Arvinas, Kymera, Nurix, BMS)."
    ),

    # -------------------------------------------------------------------------
    # TRACK 3: HIGH-IMPACT REGULATORY CATALYSTS & STATUTORY ACTIONS
    # -------------------------------------------------------------------------
    (
        "3. Regulatory Catalysts",
        "FDA Approval & PDUFA Decisions",
        "PDUFA, FDA approval, sBLA approval, NDA approval, Complete Response Letter, CRL, tentative approval",
        "FDA OR agency OR pharmaceutical OR drug OR therapeutic OR approval",
        "food recall, pet food, consumer cosmetic, medical device drill",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "High-priority alert for all FDA marketing authorization decisions, PDUFA target dates, and Complete Response Letters."
    ),
    (
        "3. Regulatory Catalysts",
        "Expedited Designation Awards",
        "Breakthrough Therapy Designation, Fast Track Designation, Priority Review, Accelerated Approval, RMAT, PRIME designation",
        "FDA OR EMA OR granted OR drug OR clinical OR therapeutic",
        "Amazon Prime, fast track train, breakthrough idea",
        "🟡 Tier 2 (Daily Briefing)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Tracks FDA/EMA regulatory speed pathways that compress time-to-market for competitor assets."
    ),
    (
        "3. Regulatory Catalysts",
        "EMA CHMP Opinions & Marketing Authorizations",
        "CHMP positive opinion, CHMP negative opinion, European Commission marketing authorisation, EMA validation, Type II variation",
        "EMA OR European Medicines Agency OR committee OR recommendation OR medicinal product",
        "stock options, cryptocurrency",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Captures European regulatory milestones from monthly CHMP meetings."
    ),
    (
        "3. Regulatory Catalysts",
        "Clinical Holds & Safety Alerts",
        "clinical hold, partial clinical hold, black box warning, boxed warning, REMS, safety alert, trial paused, patient death",
        "FDA OR trial OR clinical OR patient OR safety OR investigation",
        "holding company, holding pattern, sports box",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Instant critical alert for clinical study pauses, toxicity holds, or FDA safety labeling revisions."
    ),

    # -------------------------------------------------------------------------
    # TRACK 4: HIGH-STAKES THERAPEUTIC INDICATIONS
    # -------------------------------------------------------------------------
    (
        "4. Indication & Disease",
        "Obesity, T2D & MASH / NASH",
        "Obesity, weight loss, overweight, Type 2 Diabetes, MASH, NASH, metabolic dysfunction-associated steatohepatitis, GLP-1, GIP, glucagon",
        "trial OR Phase OR FDA OR approval OR efficacy OR topline OR cardiovascular OR liver fibrosis",
        "diet blog, fitness workout, spa, cosmetic weight loss",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Monitors the multi-billion-dollar metabolic competitive landscape across injectable and oral incretin and non-incretin therapies."
    ),
    (
        "4. Indication & Disease",
        "Non-Small Cell Lung Cancer (NSCLC)",
        "NSCLC, non-small cell lung cancer, EGFR mutation, KRAS G12C, ALK, ROS1, PD-L1 high, MET exon 14",
        "trial OR Phase OR FDA OR PFS OR OS OR overall survival OR line of therapy",
        "smoking cessation patch, air pollution retail",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks frontline and refractory NSCLC targeted therapies and immune checkpoint combinations."
    ),
    (
        "4. Indication & Disease",
        "Multiple Myeloma & Hematologic Malignancies",
        "multiple myeloma, relapsed refractory multiple myeloma, RRMM, BCMA, CAR-T, bispecific, GPRC5D, quadruplet regimen",
        "trial OR Phase OR FDA OR MRD negative OR progression-free survival OR CR",
        "stock options, option alert",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks competitive frontline quadruplets, BCMA bispecifics (Tecvayli, Elrexfio), and CAR-T therapies."
    ),
    (
        "4. Indication & Disease",
        "Inflammatory Bowel Disease (IBD: Crohn's & UC)",
        "Inflammatory Bowel Disease, Crohn's Disease, Ulcerative Colitis, anti-TL1A, anti-IL-23, JAK1, S1P modulator",
        "trial OR Phase OR clinical OR remission OR endoscopic OR FDA OR approval",
        "diet supplement, herbal remedy",
        "🟡 Tier 2 (Daily Briefing)",
        "Immunology & Respiratory Desk",
        "TRUE",
        "Monitors competitive advanced biologics and oral therapies in moderate-to-severe UC and Crohn's."
    ),
    (
        "4. Indication & Disease",
        "Rare Neuromuscular (DMD & SMA)",
        "Duchenne Muscular Dystrophy, DMD, microdystrophin, exon skipping, Spinal Muscular Atrophy, SMA, SMN2, gene therapy",
        "trial OR Phase OR clinical OR ambulant OR FDA OR functional score OR NSAA OR CHOP-INTEND",
        "SMA solar inverter, DMD dentist, DMD tools",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Rare Disease Desk",
        "TRUE",
        "Tracks high-value gene therapies and exon-skipping drugs in Duchenne and SMA, strictly filtering out solar/dental acronym noise."
    ),

    # -------------------------------------------------------------------------
    # TRACK 5: CORPORATE TRANSACTIONS, M&A & STRATEGIC DISCONTINUATIONS
    # -------------------------------------------------------------------------
    (
        "5. Transactions & Strategy",
        "Biopharma M&A & Asset Acquisitions",
        "to acquire, definitive agreement to acquire, acquisition of, buyout, takeover, merger agreement, tender offer",
        "biotech OR biopharmaceutical OR therapeutics OR billion OR million OR per share OR transaction",
        "real estate, commercial property, tech merger, banking acquisition",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Business Development & M&A Desk",
        "TRUE",
        "Detects multi-million and multi-billion-dollar biopharma buyout announcements and asset purchases."
    ),
    (
        "5. Transactions & Strategy",
        "Licensing Deals & Collaboration Biobucks",
        "exclusive license agreement, global collaboration, upfront payment, development milestones, biobucks, option agreement",
        "biotech OR pharmaceutical OR therapeutics OR rights OR commercialization",
        "software license, driver license, gaming license",
        "🟡 Tier 2 (Daily Briefing)",
        "Business Development & M&A Desk",
        "TRUE",
        "Monitors early and late-stage in-licensing deals, platform collaborations, and co-development rights."
    ),
    (
        "5. Transactions & Strategy",
        "Pipeline Terminations & Strategic Reprioritizations",
        "discontinue development of, terminate study, pipeline reprioritization, strategic review, halted trial, failed primary endpoint",
        "Phase OR clinical OR asset OR development OR patient OR program",
        "computer program, software termination",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Business Development & M&A Desk",
        "TRUE",
        "Immediate alert when a competitor drops or discontinues a clinical development asset following trial failure or strategic shift."
    ),
    (
        "5. Transactions & Strategy",
        "Biotech Restructuring & Workforce Reductions",
        "workforce reduction, laying off, headcount reduction, corporate restructuring, cost-saving initiatives, runway extension",
        "biotech OR biopharma OR pharmaceutical OR employees OR workforce OR operations",
        "tech layoffs, retail layoffs, manufacturing auto",
        "🟢 Tier 3 (Horizon / Weekly)",
        "Business Development & M&A Desk",
        "TRUE",
        "Tracks operational distress, cash runway constraints, and corporate restructuring across the biotech ecosystem."
    ),

    # -------------------------------------------------------------------------
    # TRACK 6: GLOBAL UNIVERSAL NOISE & FALSE-POSITIVE NEGATIONS
    # -------------------------------------------------------------------------
    (
        "6. Noise Negation Filters",
        "Financial & Algorithmic Stock Spam",
        "undervalued by 10%, options market predicting spike, stock options trading, lock-up agreement ending, Zacks rank upgrade, short interest increased, share buy-back announcement",
        "N/A (Global Noise Rule)",
        "N/A",
        "🚫 Noise Suppression Rule",
        "Automated Filtering Engine",
        "TRUE",
        "Universal negative keyword suppression pattern to eliminate algorithmic retail trading noise across Google News feeds."
    ),
    (
        "6. Noise Negation Filters",
        "Consumer, Food & Pet Recalls",
        "dog food recall, cat food recall, undeclared peanut, beef contamination, organic spinach recall, cosmetic eye drops recall",
        "N/A (Global Noise Rule)",
        "N/A",
        "🚫 Noise Suppression Rule",
        "Automated Filtering Engine",
        "TRUE",
        "Filters out non-pharmaceutical food, agricultural, and veterinary FDA notices."
    ),
]


def create_config_tab():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    # If tab exists, remove and recreate fresh
    if TAB_NAME in wb.sheetnames:
        print(f"  Replacing existing '{TAB_NAME}' tab...")
        del wb[TAB_NAME]
    
    ws = wb.create_sheet(title=TAB_NAME)
    ws.views.sheetView[0].showGridLines = True

    print(f"Creating '{TAB_NAME}' with {len(CONFIG_RULES)} comprehensive tracking rules...")

    # Visual Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    CRIMSON_DARK = "8B0000"
    WHITE = "FFFFFF"
    ICE_BLUE = "EBF2FA"
    LIGHT_GRAY = "F4F6F8"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_ice = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    # Tier fills
    fill_tier1 = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid") # soft red/pink
    font_tier1 = Font(name="Calibri", size=10, bold=True, color="C2185B")
    
    fill_tier2 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # soft yellow
    font_tier2 = Font(name="Calibri", size=10, bold=True, color="B78103")
    
    fill_tier3 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    font_tier3 = Font(name="Calibri", size=10, bold=True, color="2E7D32")
    
    fill_noise = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid") # grey
    font_noise = Font(name="Calibri", size=10, bold=True, color="495057")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_italic = Font(name="Calibri", size=9, italic=True, color="444444")
    font_status_active = Font(name="Calibri", size=10, bold=True, color="155724")
    fill_status_active = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Category / Track Type", 24, fill_navy),
        ("Primary Focus Entity", 30, fill_navy),
        ("Synonyms, Aliases & Code Names", 45, fill_navy),
        ("Mandatory Context Qualifiers (Co-occurrence)", 45, fill_navy),
        ("Negative Exclude Terms (Negations)", 42, fill_navy),
        ("Alert Priority Tier", 25, fill_blue),
        ("Assigned CI Desk / Specialist", 28, fill_blue),
        ("Active Status", 15, fill_blue),
        ("Operational Role & Intelligence Focus", 55, fill_blue),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    for r_idx, rule in enumerate(CONFIG_RULES, start=2):
        cat, entity, syn, qual, neg, tier, desk, active, role = rule

        # Col A: Category
        cell_a = ws.cell(row=r_idx, column=1, value=cat)
        cell_a.font = font_bold
        cell_a.border = thin_border
        cell_a.fill = fill_ice if r_idx % 2 == 0 else fill_white

        # Col B: Entity
        cell_b = ws.cell(row=r_idx, column=2, value=entity)
        cell_b.font = font_bold
        cell_b.border = thin_border

        # Col C: Synonyms
        cell_c = ws.cell(row=r_idx, column=3, value=syn)
        cell_c.font = font_code
        cell_c.border = thin_border
        cell_c.alignment = Alignment(wrap_text=True)

        # Col D: Qualifiers
        cell_d = ws.cell(row=r_idx, column=4, value=qual)
        cell_d.font = font_code
        cell_d.border = thin_border
        cell_d.alignment = Alignment(wrap_text=True)

        # Col E: Negations
        cell_e = ws.cell(row=r_idx, column=5, value=neg)
        cell_e.font = font_code
        cell_e.border = thin_border
        cell_e.alignment = Alignment(wrap_text=True)

        # Col F: Priority Tier
        cell_f = ws.cell(row=r_idx, column=6, value=tier)
        cell_f.border = thin_border
        cell_f.alignment = Alignment(horizontal="center", vertical="center")
        if "Tier 1" in tier:
            cell_f.font = font_tier1
            cell_f.fill = fill_tier1
        elif "Tier 2" in tier:
            cell_f.font = font_tier2
            cell_f.fill = fill_tier2
        elif "Tier 3" in tier:
            cell_f.font = font_tier3
            cell_f.fill = fill_tier3
        else:
            cell_f.font = font_noise
            cell_f.fill = fill_noise

        # Col G: Assigned Desk
        cell_g = ws.cell(row=r_idx, column=7, value=desk)
        cell_g.font = font_data
        cell_g.border = thin_border
        cell_g.alignment = Alignment(vertical="center")

        # Col H: Active Status
        cell_h = ws.cell(row=r_idx, column=8, value=active)
        cell_h.border = thin_border
        cell_h.alignment = Alignment(horizontal="center", vertical="center")
        if active == "TRUE":
            cell_h.font = font_status_active
            cell_h.fill = fill_status_active
        else:
            cell_h.font = font_data

        # Col I: Role
        cell_i = ws.cell(row=r_idx, column=9, value=role)
        cell_i.font = font_italic
        cell_i.border = thin_border
        cell_i.alignment = Alignment(wrap_text=True, vertical="center")

        ws.row_dimensions[r_idx].height = 36

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    # Also update 01_System_Overview to reference this new tab
    if "01_System_Overview" in wb.sheetnames:
        ws_over = wb["01_System_Overview"]
        # Find if there is a section for configs
        ws_over.cell(row=24, column=2, value="4. KEYWORD MATCHING & RELEVANCE ENGINE")
        ws_over.cell(row=24, column=2).font = Font(name="Calibri", size=12, bold=True, color=WHITE)
        ws_over.cell(row=24, column=2).fill = PatternFill(start_color="2D5584", end_color="2D5584", fill_type="solid")
        ws_over.cell(row=25, column=2, value="Interactive Tab 07")
        ws_over.cell(row=25, column=3, value="07_Keywords_Match_Config provides an analyst-driven rule matrix across 6 dimensions: Assets, Modalities, Catalysts, Indications, Deals, and Noise Negations. Drives automated relevance scoring, priority tiering (Tier 1/2/3), and team routing.")
        ws_over.cell(row=25, column=2).font = font_bold
        ws_over.cell(row=25, column=3).font = font_data

    wb.save(XLSX_PATH)
    print(f"\nSuccessfully saved Master Excel workbook: {XLSX_PATH}")
    print(f"Current Sheets in Master Workbook: {wb.sheetnames}")


if __name__ == "__main__":
    create_config_tab()
