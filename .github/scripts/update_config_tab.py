#!/usr/bin/env python3
"""
Update 06_Config_and_Settings in RSSFeedChecker_Master_Guide_and_Data.xlsx
to include Universal Time-Windowing parameters, timezones, and the 7-day Admin Guardrail.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX_PATH = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"

CONFIG_ROWS = [
    # Universal Time-Windowing & Guardrails
    ("Universal Time-Windowing", "TIME_WINDOW / --window", "72h", "Relative ('24h','48h','72h','7d'), Absolute Start ('From 24 Aug 2026 01:00 AM IST onwards'), Bounded Range ('START to END')", "Universal time-window filter applied across ALL 3 methods (Feeds, Newsrooms, CT.gov). Controls exact timeframe for data ingestion."),
    ("Universal Time-Windowing", "MAX_STANDARD_WINDOW_DAYS", "7", "Integer (1-7 days / 168 hours)", "Security & Performance Guardrail. Strictly prevents standard non-admin users from running queries exceeding 7 days across any source."),
    ("Universal Time-Windowing", "ADMIN_OVERRIDE / --admin", "false", "Boolean (true/false)", "Administrative privilege flag. Unlocks extended historical queries beyond 7 days (e.g. 14d, 30d, or historical archive lookups)."),
    ("Universal Time-Windowing", "DEFAULT_TIMEZONE", "IST (+05:30)", "Timezone string ('IST', 'UTC', 'EST', 'EDT', 'PST', 'CET', 'GMT')", "Default timezone used when timestamps are provided without explicit timezone offset suffixes."),

    # Pillar 1: Publisher Feeds
    ("Pillar 1: Publisher Feeds", "FEED_TIMEOUT", "30", "Seconds (int/float)", "HTTP timeout per feed attempt. Prevents slow servers from hanging the job."),
    ("Pillar 1: Publisher Feeds", "FEED_WORKERS", "8", "Integer (1-32)", "Number of parallel threads fetching feeds simultaneously."),
    ("Pillar 1: Publisher Feeds", "FEED_HOST_GAP", "1.0", "Seconds (float)", "Minimum delay between consecutive requests to the same host domain (prevents 429 rate limits)."),
    ("Pillar 1: Publisher Feeds", "FEED_RETRIES", "1", "Integer (0-3)", "Extra retry attempts for transient errors (HTTP 0, 401, 403, 429, 5xx)."),

    # Pillar 2: Company Newsroom Watch
    ("Pillar 2: Company Newsrooms", "NEWS_DAYS_BACK", "3", "Integer / Window String", "Strict time-window filter for corporate press releases. Sitemaps and RSS are filtered to this window."),
    ("Pillar 2: Company Newsrooms", "NEWS_TIMEOUT", "25", "Seconds (int/float)", "Per-request timeout for company homepages, robots.txt, sitemaps, and PR pages."),
    ("Pillar 2: Company Newsrooms", "NEWS_WORKERS", "6", "Integer (1-16)", "Parallel thread pool size for company newsroom crawling."),
    ("Pillar 2: Company Newsrooms", "NEWS_MAX_SUBSITEMAPS", "20", "Integer", "Max child sitemaps inspected per company (prevents runaway indexing)."),
    ("Pillar 2: Company Newsrooms", "NEWS_STATE_CAP", "6000", "Integer", "Maximum historical release URLs remembered per company domain in seen.json."),
    ("Pillar 2: Company Newsrooms", "SEC_UA", "None", "String ('Name email@domain.com')", "Custom user-agent header required by SEC.gov EDGAR feeds to prevent 403 blocks."),

    # Pillar 3: ClinicalTrials.gov & Indication Radar
    ("Pillar 3: ClinicalTrials.gov", "CT_INGESTION_SCOPE", "New Registrations Only", "'New Registrations Only' / 'Full Protocol Delta'", "Simplified baseline ingestion: absorbs only newly registered trials within the time window. Multi-version diffing separated."),
    ("Pillar 3: ClinicalTrials.gov", "CT_MAX_STUDIES_PER_COND", "50", "Integer (1-100)", "Maximum number of recent studies fetched per indication per scan."),
    ("Pillar 3: ClinicalTrials.gov", "IND_TIMEOUT", "30", "Seconds (int/float)", "Per-request verification timeout for generated Google News and CT.gov feeds."),
    ("Pillar 3: ClinicalTrials.gov", "IND_WORKERS", "4", "Integer (1-8)", "Worker threads for parallel feed verification."),
]


def update_config_tab():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)

    ws_name = next((s for s in wb.sheetnames if "Config" in s), "06_Config_and_Settings")
    ws = wb[ws_name]

    # Clear existing rows below header
    for r in range(ws.max_row, 1, -1):
        ws.delete_rows(r)

    # Styles
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    WHITE = "FFFFFF"
    ICE_BLUE = "EBF2FA"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_ice = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_data = Font(name="Calibri", size=10, color="000000")

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Workflow / Component", 28, fill_navy),
        ("Setting / Parameter", 32, fill_navy),
        ("Default Value", 22, fill_blue),
        ("Valid Range / Options", 42, fill_navy),
        ("Operational Role, Impact & Security Guardrails", 65, fill_blue),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    for r_idx, row_data in enumerate(CONFIG_ROWS, start=2):
        comp, param, default_val, valid_range, desc = row_data

        cell_a = ws.cell(row=r_idx, column=1, value=comp)
        cell_a.font = font_bold
        cell_a.border = thin_border
        cell_a.fill = fill_ice if "Universal" in comp else fill_white

        cell_b = ws.cell(row=r_idx, column=2, value=param)
        cell_b.font = font_code
        cell_b.border = thin_border

        cell_c = ws.cell(row=r_idx, column=3, value=default_val)
        cell_c.font = font_code
        cell_c.border = thin_border
        cell_c.alignment = Alignment(horizontal="center")
        if "72h" in default_val or "7" in default_val or "New Registrations" in default_val:
            cell_c.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

        cell_d = ws.cell(row=r_idx, column=4, value=valid_range)
        cell_d.font = font_data
        cell_d.border = thin_border
        cell_d.alignment = Alignment(wrap_text=True)

        cell_e = ws.cell(row=r_idx, column=5, value=desc)
        cell_e.font = font_data
        cell_e.border = thin_border
        cell_e.alignment = Alignment(wrap_text=True)

        ws.row_dimensions[r_idx].height = 28

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    wb.save(XLSX_PATH)
    print(f"Updated '{ws_name}' with {len(CONFIG_ROWS)} configuration settings and guardrails.")


if __name__ == "__main__":
    update_config_tab()
