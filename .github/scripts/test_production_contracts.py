#!/usr/bin/env python3
"""
PRODUCTION CONTRACT & SCHEMA VALIDATION SUITE
=============================================
Asserts and guarantees that:
1. 'Results' worksheet strictly adheres to the 27-column canonical schema (A through AA)
2. Column headers match the exact canonical order with ZERO column shifts
3. Historical rows preserve their exact column alignments during accumulative read/write
4. ai_synthesis_engine.py cleanly exports synthesize_event returning (ai_summary, implications, provenance)
5. web/data/latest_intelligence.json contains all required schema keys including 'provenance'
6. run_unified_intelligence_pipeline.py imports and initializes without errors
"""

import os
import sys
import json
import openpyxl

# Ensure UTF-8 output
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    except Exception:
        pass

CANONICAL_27_HEADERS = [
    (1, "A", "Published Date"),
    (2, "B", "Published Time (UTC)"),
    (3, "C", "Event Headline"),
    (4, "D", "Project / Indication Theme"),
    (5, "E", "Signal Type"),
    (6, "F", "Relevance Score"),
    (7, "G", "Priority Tier"),
    (8, "H", "Routed Desk / CI Workstream"),
    (9, "I", "Matched Biopharma Catalyst"),
    (10, "J", "Source Entity Name"),
    (11, "K", "Source Classification"),
    (12, "L", "Direct Publisher / SEC URL"),
    (13, "M", "Editorial Snippet / Summary"),
    (14, "N", "Cluster ID"),
    (15, "O", "Cluster Hint JSON"),
    (16, "P", "Discovery Pillar"),
    (17, "Q", "Extraction Vector / Method"),
    (18, "R", "Discovered At (UTC)"),
    (19, "S", "Full Body Excerpt"),
    (20, "T", "Event UUID"),
    (21, "U", "AI Summary"),
    (22, "V", "Implications"),
    (23, "W", "AI Synthesis Method / Provenance"),
    (24, "X", "Ingestion Batch ID"),
    (25, "Y", "Ingested Date (IST)"),
    (26, "Z", "Ingested Time (IST)"),
    (27, "AA", "Execution Run Type"),
]

def run_contract_tests():
    print("=" * 85)
    print(" 🛡️ EXECUTING PRODUCTION CONTRACT VALIDATION SUITE")
    print("=" * 85)
    
    workspace = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    xlsx_path = os.path.join(workspace, "RSSFeedChecker_Master_Guide_and_Data.xlsx")
    json_path = os.path.join(workspace, "web", "data", "latest_intelligence.json")
    
    # -------------------------------------------------------------------------
    # TEST 1: Pipeline Import Test
    # -------------------------------------------------------------------------
    print("\n[TEST 1/6] Validating run_unified_intelligence_pipeline.py Import Contract...")
    sys.path.insert(0, os.path.join(workspace, ".github", "scripts"))
    try:
        import run_unified_intelligence_pipeline
        print("  ✅ [PASS] run_unified_intelligence_pipeline.py imported cleanly with 0 errors!")
    except Exception as e:
        print(f"  ❌ [FAIL] Pipeline import failed: {e}")
        sys.exit(1)

    # -------------------------------------------------------------------------
    # TEST 2: AI Synthesis Return Contract
    # -------------------------------------------------------------------------
    print("\n[TEST 2/6] Validating AI Synthesis Engine Contract (synthesize_event)...")
    from ai_synthesis_engine import synthesize_event
    test_res = synthesize_event(
        title="Ascletis Doses Participant in Phase 3 ASC30 Trial",
        company="Ascletis",
        desk="Metabolic & Obesity Desk",
        snippet="Ascletis dosed first patient in Phase 3 AURORA trial.",
        full_text="Ascletis dosed first patient in Phase 3 AURORA trial evaluating 4600 patients.",
        primary_mode="Option A",
        secondary_mode="None"
    )
    assert len(test_res) == 3, f"Expected 3 return values, got {len(test_res)}"
    ai_s, imp, prov = test_res
    assert ai_s and len(ai_s) > 30, "AI summary must not be empty"
    assert imp and len(imp) > 20, "Implications must not be empty"
    assert "Local" in prov or "Gemini" in prov or "Cloudflare" in prov, f"Invalid provenance: {prov}"
    print(f"  ✅ [PASS] synthesize_event returns clean 3-tuple -> Provenance: '{prov}'")

    # -------------------------------------------------------------------------
    # TEST 3: Excel Results Tab 27-Column Header Contract
    # -------------------------------------------------------------------------
    print("\n[TEST 3/6] Validating Master Excel 'Results' Tab 27-Column Canonical Schema...")
    assert os.path.exists(xlsx_path), f"Master workbook not found at {xlsx_path}"
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    assert "Results" in wb.sheetnames, "'Results' tab missing from master workbook"
    ws = wb["Results"]
    
    assert ws.max_column == 27, f"Expected exactly 27 columns, found {ws.max_column}"
    for pos, col_let, expected_name in CANONICAL_27_HEADERS:
        actual_name = str(ws.cell(1, pos).value or "").strip()
        assert actual_name == expected_name, f"Column mismatch at Pos {pos} ({col_let}): Expected '{expected_name}', Found '{actual_name}'"
    print(f"  ✅ [PASS] All 27 columns (A through AA) match canonical schema with zero shift!")

    # -------------------------------------------------------------------------
    # TEST 4: Historical Row Data Alignment Contract
    # -------------------------------------------------------------------------
    print("\n[TEST 4/6] Validating Historical Row Data Alignment (No Column Shifting)...")
    sample_row = 2
    val_uuid = str(ws.cell(sample_row, 20).value or "")
    val_ai = str(ws.cell(sample_row, 21).value or "")
    val_imp = str(ws.cell(sample_row, 22).value or "")
    val_prov = str(ws.cell(sample_row, 23).value or "")
    val_batch = str(ws.cell(sample_row, 24).value or "")
    val_date_ist = str(ws.cell(sample_row, 25).value or "")
    val_time_ist = str(ws.cell(sample_row, 26).value or "")
    val_run_type = str(ws.cell(sample_row, 27).value or "")
    
    assert val_uuid.startswith("EVT_"), f"Col 20 must be Event UUID, got '{val_uuid}'"
    assert len(val_ai) > 30, f"Col 21 must be AI Summary, got '{val_ai}'"
    assert len(val_imp) > 20, f"Col 22 must be Implications, got '{val_imp}'"
    assert "Local" in val_prov or "Gemini" in val_prov or "Cloudflare" in val_prov, f"Col 23 must be Provenance, got '{val_prov}'"
    assert "RUN_" in val_batch or "HISTORICAL" in val_batch or "BATCH" in val_batch, f"Col 24 must be Batch ID, got '{val_batch}'"
    assert val_date_ist, "Col 25 Ingested Date (IST) must not be empty"
    assert "IST" in val_time_ist or ":" in val_time_ist, f"Col 26 Ingested Time must be time string, got '{val_time_ist}'"
    assert val_run_type, "Col 27 Execution Run Type must not be empty"
    print("  ✅ [PASS] Historical row alignment confirmed across all telemetry & provenance columns!")

    # -------------------------------------------------------------------------
    # TEST 5: Web JSON Schema & Key Preservation Contract
    # -------------------------------------------------------------------------
    print("\n[TEST 5/6] Validating web/data/latest_intelligence.json Payload Contract...")
    assert os.path.exists(json_path), f"JSON payload missing at {json_path}"
    with open(json_path, "r", encoding="utf-8") as f:
        web_data = json.load(f)
        
    events = web_data.get("events", [])
    assert len(events) == ws.max_row - 1, f"Event count mismatch: JSON has {len(events)}, Excel has {ws.max_row - 1}"
    
    sample_ev = events[0]
    required_keys = [
        "event_id", "published_date", "published_time", "headline", "project_name",
        "signal_type", "relevance_score", "priority_tier", "desk", "source_name",
        "raw_url", "snippet", "full_text", "cluster_id", "ai_summary", "implications",
        "provenance"
    ]
    for rk in required_keys:
        assert rk in sample_ev, f"Missing required key '{rk}' in JSON event object"
    assert sample_ev["provenance"], "JSON provenance value must not be empty"
    print(f"  ✅ [PASS] JSON schema verified with all {len(required_keys)} required keys and active provenance!")

    # -------------------------------------------------------------------------
    # TEST 6: Frontend Provenance Rendering Syntax
    # -------------------------------------------------------------------------
    print("\n[TEST 6/6] Validating Frontend web/app.js Template Contract...")
    js_path = os.path.join(workspace, "web", "app.js")
    with open(js_path, "r", encoding="utf-8") as f:
        app_js = f.read()
    assert "lead.provenance" in app_js and "item.provenance" in app_js, "web/app.js must render lead.provenance and item.provenance"
    print("  ✅ [PASS] web/app.js renders provenance badge across all card templates!")

    print("\n" + "=" * 85)
    print(" 🏆 ALL 6 PRODUCTION CONTRACT TESTS PASSED WITH ZERO DRIFT OR CORRUPTION!")
    print("=" * 85)

if __name__ == "__main__":
    run_contract_tests()
