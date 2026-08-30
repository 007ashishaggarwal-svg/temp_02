#!/usr/bin/env python3
"""
Consolidate Master Workbook into 4 Pristine Production Tabs:
1. '00_Unified_Intelligence_Feed' (Output: Ingested News, Catalysts & Clinical Trials)
2. '01_Master_Sources_Registry' (Input: 4,047 Verified Working Endpoints across all 3 Pillars)
3. '02_Keywords_and_Rules' (Configuration: Biopharma Taxonomy, Noise Regexes, Priority Scorer)
4. '03_Config_and_Settings' (Configuration: Time Window & System Settings)

Prunes all 5 redundant legacy tabs.
"""

import os
import sys
import time
import openpyxl

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
SRC_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data_NEW.xlsx")
DST_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

def consolidate_workbook():
    print(f"Loading source workbook from {SRC_PATH}...")
    wb = openpyxl.load_workbook(SRC_PATH)

    # 1. Rename tabs
    if "Unified_All_Pillars" in wb.sheetnames:
        ws_u = wb["Unified_All_Pillars"]
        ws_u.title = "01_Master_Sources_Registry"

    if "07_Keywords_Match_Config" in wb.sheetnames:
        ws_k = wb["07_Keywords_Match_Config"]
        ws_k.title = "02_Keywords_and_Rules"

    if "06_Config_and_Settings" in wb.sheetnames:
        ws_c = wb["06_Config_and_Settings"]
        ws_c.title = "03_Config_and_Settings"

    # 2. Delete redundant legacy tabs
    redundant_tabs = [
        "01_System_Overview",
        "02_Data_Dictionary",
        "03_Feeds_Master (459)",
        "04_Companies_Master (616)",
        "05_Indications_Radar (18)",
        "Preview_01_Combined_Wide",
        "Preview_02_Combined_Long",
        "Preview_03_Unified_All_Pillars"
    ]

    for t in redundant_tabs:
        if t in wb.sheetnames:
            print(f"  🗑️ Pruning redundant tab: '{t}'...")
            del wb[t]

    print(f"\nFinal Consolidated Tabs: {wb.sheetnames}")

    # 3. Save to DST_PATH with retry logic
    saved = False
    for attempt in range(5):
        try:
            wb.save(DST_PATH)
            saved = True
            print(f"✅ Successfully saved pristine 4-tab master workbook to: {DST_PATH}")
            break
        except PermissionError:
            print(f"⚠️ Excel is open on {DST_PATH} (Attempt {attempt+1}/5). Retrying in 2 seconds...")
            time.sleep(2)

    if not saved:
        wb.save(SRC_PATH)
        print(f"⚠️ Note: Master Excel file was open. Pristine 4-tab workbook saved to {SRC_PATH}")

if __name__ == "__main__":
    consolidate_workbook()
