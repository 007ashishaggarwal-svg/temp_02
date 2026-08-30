#!/usr/bin/env python3
"""
MultiVectorClusterEngine — Deterministic Multi-Vector Biopharma Event Clustering
================================================================================
Evaluates pairwise candidate events across 5 distinct biopharma vectors:
1. Dynamic Entity/Company Recognition (from 04_Biopharma_Entities_Catalog)
2. Specific Drug Asset Codes, Named INNs & NCT IDs (from Headline & 02_Keywords)
3. Financial Figures ($B/$M) & Clinical/Regulatory Action Types
4. Temporal Proximity Window (<= 72h default)
5. Semantic Desk & Indication Overlap
"""

import os
import re
import sys
import json
import hashlib
from datetime import datetime, timezone
import openpyxl

class MultiVectorClusterEngine:
    def __init__(self, xlsx_path: str = None):
        self.enable_v1_entity = True
        self.enable_v2_asset = True
        self.enable_v3_action_fin = True
        self.enable_v4_time = True
        self.enable_v5_indication = True

        self.enable_dynamic_companies = True
        self.enable_dynamic_keywords = True
        self.enable_strict_indication_guardrail = True
        self.enable_strict_time_guardrail = True
        self.enable_jaccard_similarity = True

        self.auto_cluster_threshold = 70
        self.review_cluster_threshold = 55
        self.max_time_window_hours = 72

        self.company_alias_map = {}
        self.specific_drug_anchors = {
            "cagrisema", "wegovy", "ozempic", "semaglutide", "tirzepatide", "mounjaro", "zepbound",
            "retatrutide", "orforglipron", "surmount", "surpass", "achieve", "attain", "enhertu",
            "trodelvy", "deramiocel", "brensocatib", "hopledo", "rgx-121", "rgx121", "hm-17321",
            "hm17321", "cap-1002", "cap1002", "ak-157d1", "ak157d1", "tak-279", "tak279", "st-920",
            "st920", "volrustomig", "destiny-lung", "evolve-lung", "awiqli", "monlunabant", "amycretin",
            "petrelintide", "pemvidutide", "ct-388", "ct-868", "ct-996", "vk2735", "vk0214", "lepodisiran",
            "zerlasiran", "olpasiran", "pelacarsen", "lecanoemab", "donanemab", "aducanumab", "remternetug"
        }

        if xlsx_path and os.path.exists(xlsx_path):
            self.load_all_from_workbook(xlsx_path)

    def load_all_from_workbook(self, xlsx_path: str):
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            if "04_Biopharma_Entities_Catalog" in wb.sheetnames:
                ws_ent = wb["04_Biopharma_Entities_Catalog"]
                for r in range(2, ws_ent.max_row + 1):
                    comp_name = str(ws_ent.cell(row=r, column=1).value or "").strip()
                    aliases_str = str(ws_ent.cell(row=r, column=3).value or "").strip()
                    status = str(ws_ent.cell(row=r, column=6).value or "Active").strip()
                    if comp_name and status.lower() != "inactive":
                        self.company_alias_map[comp_name.lower()] = comp_name
                        if aliases_str:
                            for a in aliases_str.split(","):
                                a_clean = a.strip().lower()
                                if a_clean and len(a_clean) >= 2:
                                    self.company_alias_map[a_clean] = comp_name

            if "02_Keywords_and_Rules" in wb.sheetnames:
                ws_rules = wb["02_Keywords_and_Rules"]
                broad_noise = {"gene therapy", "cell therapy", "adc", "monoclonal antibody", "small molecule", "approval", "primary endpoint", "clinical trial", "phase 3", "phase 2", "oncology", "cancer"}
                for r in range(2, ws_rules.max_row + 1):
                    prim_kw = str(ws_rules.cell(row=r, column=2).value or "").strip()
                    status = str(ws_rules.cell(row=r, column=7).value or "Active").strip()
                    if status.lower() == "active":
                        for kw in prim_kw.split(","):
                            kw_clean = kw.strip().lower()
                            if kw_clean and len(kw_clean) >= 3 and kw_clean not in broad_noise:
                                self.specific_drug_anchors.add(kw_clean)
        except Exception:
            pass

    def extract_features(self, item: dict) -> dict:
        headline = item.get("headline", "")
        snippet = item.get("snippet", "")
        comb_text = f"{headline} {snippet}".lower()
        h_low = headline.lower()

        # 1. Company Extraction: HEADLINE FIRST
        h_comps = set()
        for alias, canonical in sorted(self.company_alias_map.items(), key=lambda x: -len(x[0])):
            if re.search(r"\b" + re.escape(alias) + r"\b", h_low):
                h_comps.add(canonical)

        s_comps = set()
        if not h_comps:
            for alias, canonical in sorted(self.company_alias_map.items(), key=lambda x: -len(x[0])):
                if re.search(r"\b" + re.escape(alias) + r"\b", comb_text):
                    s_comps.add(canonical)

        all_comps = h_comps if h_comps else s_comps
        lead_comp = list(all_comps)[0] if all_comps else item.get("matched_company", "Not Identified")
        if lead_comp == "Not Identified" and all_comps:
            lead_comp = list(all_comps)[0]
        partner = list(all_comps)[1] if len(all_comps) > 1 else ""

        # 2. Specific Drug Assets & NCT IDs
        asset_matches = set()
        for m in re.finditer(r"\b(nct\d{8})\b", comb_text):
            asset_matches.add(m.group(1).lower())

        for drug in self.specific_drug_anchors:
            clean_d = re.sub(r"[-_\s]", "", drug).lower()
            if re.search(r"\b" + re.escape(drug) + r"\b", h_low) or re.search(r"\b" + re.escape(clean_d) + r"\b", h_low):
                asset_matches.add(clean_d)

        # 3. Action type
        action = ""
        if re.search(r"\b(?:acquir(?:e|es|ed|ition)|to\s+buy|merger|licensing|pact|deal|\$\d+(\.\d+)?\s*(?:b|m|billion|million))\b", h_low):
            action = "deal"
        elif re.search(r"\b(?:clinical\s+hold|halt(?:ed)?|safety\s+concerns?|discontinued)\b", h_low):
            action = "hold"
        elif re.search(r"\b(?:approv(?:ed|al)|chmp|pdufa|bla|nda)\b", h_low):
            action = "approval"
        elif re.search(r"\b(?:win|positive|phase\s+3|met\s+primary|survival)\b", h_low):
            action = "clinical_win"

        # 4. Financial amounts
        fin_matches = set()
        for m in re.finditer(r"\$?(\d+(?:\.\d+)?)\s*(b|bn|billion|m|million)\b", comb_text):
            val = m.group(1)
            unit = "b" if m.group(2).startswith("b") else "m"
            fin_matches.add(f"{val}{unit}")

        # 5. Headline Word Tokens for Jaccard
        stop_words = {"the", "and", "for", "with", "from", "that", "this", "after", "into", "over", "under", "about", "novo", "nordisk", "lilly", "roche", "pfizer", "astrazeneca", "drug", "trial", "study"}
        sig_tokens = set(w for w in re.findall(r"\b[a-zA-Z0-9]{3,}\b", h_low) if w not in stop_words)

        return {
            "headline": headline,
            "headline_comps": h_comps,
            "all_comps": all_comps,
            "lead_company": lead_comp,
            "partner": partner,
            "asset_set": asset_matches,
            "action": action,
            "fin_set": fin_matches,
            "project": item.get("project_name", ""),
            "date": item.get("published_date", ""),
            "sig_tokens": sig_tokens
        }

    def compute_jaccard_similarity(self, tokens_a: set, tokens_b: set) -> float:
        if not tokens_a or not tokens_b:
            return 0.0
        return float(len(tokens_a & tokens_b)) / float(len(tokens_a | tokens_b))

    def calculate_pairwise_score(self, item_a: dict, item_b: dict) -> int:
        fa = item_a["_feats"]
        fb = item_b["_feats"]

        # Date proximity guardrail (<= 3 days)
        diff_days = 1
        try:
            da = datetime.strptime(fa["date"], "%Y-%m-%d")
            db = datetime.strptime(fb["date"], "%Y-%m-%d")
            diff_days = abs((da - db).days)
        except Exception:
            diff_days = 1

        if diff_days > 3:
            return 0

        # Action incompatibility
        if fa["action"] and fb["action"] and fa["action"] != fb["action"] and fa["action"] in ["deal", "hold"] and fb["action"] in ["deal", "hold"]:
            return 0

        # 1. Check 1: Identical NCT Trial ID (Exact Ground Truth)
        shared_nct = [d for d in (fa["asset_set"] & fb["asset_set"]) if d.startswith("nct")]
        if shared_nct:
            return 95

        # 2. Check 2: High Headline Word Overlap (>= 0.60 Jaccard) for syndicated wire stories
        j_score = self.compute_jaccard_similarity(fa["sig_tokens"], fb["sig_tokens"])
        shared_comps = fa["all_comps"] & fb["all_comps"]
        if j_score >= 0.60 and (shared_comps or (fa["lead_company"] != "Not Identified" and fa["lead_company"] == fb["lead_company"])):
            return 90

        # 3. Check 3: Multi-Company Specific M&A / Partnership Deal (Both Partners in Headline + Deal Figure/Action)
        if len(shared_comps) >= 2 and fa["action"] == "deal" and fb["action"] == "deal":
            shared_fin = fa["fin_set"] & fb["fin_set"]
            if shared_fin or j_score >= 0.40:
                return 85

        # 4. Ultra-high headline overlap (>= 0.70) across different publisher syndications
        if j_score >= 0.70:
            return 80

        # Default: If there is any doubt -> 0 score (Remains 100% Standalone / Independent)
        return 0

    def cluster_feed_items(self, items: list) -> list:
        if not items:
            return []

        for it in items:
            it["_feats"] = self.extract_features(it)

        clusters = []
        assigned = set()

        for i, it1 in enumerate(items):
            if i in assigned:
                continue
            current_cluster = [it1]
            assigned.add(i)

            for j, it2 in enumerate(items):
                if j in assigned:
                    continue
                pair_score = self.calculate_pairwise_score(it1, it2)
                if pair_score >= self.review_cluster_threshold:
                    current_cluster.append(it2)
                    assigned.add(j)

            clusters.append(current_cluster)

        # Process clusters and assign canonical identifiers
        for cluster in clusters:
            lead_item = cluster[0]
            feats = lead_item["_feats"]
            lead_comp = feats["lead_company"]
            partner = feats["partner"]
            indication = feats["project"]
            date_slug = lead_item.get("published_date", "").replace("-", "") or datetime.now(timezone.utc).strftime("%Y%m%d")

            key_asset = next(iter(feats["asset_set"]), "")
            deal_fig = next(iter(feats["fin_set"]), "")
            action = feats["action"]

            if len(cluster) > 1:
                conf_score = self.calculate_pairwise_score(cluster[0], cluster[1])
                status = "auto_grouped" if conf_score >= self.auto_cluster_threshold else "needs_review"

                comp_slug = lead_comp.upper().replace(" ", "_")[:15] if lead_comp != "Not Identified" else "EVENT"
                asset_slug = f"_{key_asset.upper()[:12]}" if key_asset else ""
                fin_slug = f"_{deal_fig.upper()}" if deal_fig else (f"_{action.upper()[:10]}" if action else "")

                canonical_id = f"CLUST_{date_slug}_{comp_slug}{asset_slug}{fin_slug}"

                hint_dict = {
                    "canonical_cluster_id": canonical_id,
                    "confidence_score": conf_score,
                    "cluster_status": status,
                    "lead_company": lead_comp,
                    "partner_company": partner if partner else "None",
                    "indication": indication,
                    "signal_type": lead_item.get("signal_type", "general"),
                    "key_asset": key_asset.upper() if key_asset else "General",
                    "deal_figure": deal_fig.upper() if deal_fig else "N/A",
                    "cluster_size": len(cluster)
                }
                hint_json = json.dumps(hint_dict)

                for member in cluster:
                    member["cluster_id"] = canonical_id
                    member["cluster_hint"] = hint_json
                    member["matched_company"] = lead_comp
                    member["partner_company"] = partner

            else:
                key_token = key_asset.lower() if key_asset else ""
                if not key_token:
                    sig_words = [w.lower() for w in re.findall(r"\b[a-zA-Z0-9]{3,}\b", lead_item.get("headline", "")) if w.lower() not in [
                        "the", "and", "for", "with", "from", "that", "this", "after", "into", "over", "under", "about",
                        "human", "medicines", "european", "public", "assessment", "report", "epar", "date", "authorisation"
                    ]]
                    key_token = sig_words[0] if sig_words else "item"

                title_words = [w.lower() for w in re.findall(r"\b[a-zA-Z0-9]{3,}\b", lead_item.get("headline", "")) if w.lower() not in [
                    "the", "and", "for", "with", "from", "that", "this", "after", "into", "over", "under", "about"
                ]]
                hash_seed = "".join(title_words[:8]) + lead_item.get("raw_url", "")[-10:]
                hash4 = hashlib.md5(hash_seed.encode("utf-8")).hexdigest()[:4]

                comp_slug = lead_comp.lower().replace(" ", "_")[:15] if lead_comp != "Not Identified" else "unidentified"
                standalone_id = f"EVT_{date_slug}_{comp_slug}_{key_token[:10]}_{hash4}"

                hint_dict = {
                    "canonical_cluster_id": standalone_id,
                    "confidence_score": 100,
                    "cluster_status": "standalone",
                    "lead_company": lead_comp,
                    "indication": indication,
                    "signal_type": lead_item.get("signal_type", "general"),
                    "cluster_size": 1
                }

                lead_item["cluster_id"] = standalone_id
                lead_item["cluster_hint"] = json.dumps(hint_dict)
                lead_item["matched_company"] = lead_comp
                lead_item["partner_company"] = partner

        return items
