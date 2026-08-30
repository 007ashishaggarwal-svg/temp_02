#!/usr/bin/env python3
"""
Inspect all sheets, user changes, headers, and rows in RSSFeedChecker_Master_Guide_and_Data.xlsx.
"""

import os
import sys
import openpyxl

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker\RSSFeedChecker_Master_Guide_and_Data.xlsx"

def inspect():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    print(f"Workbook loaded: {XLSX}")
    print(f"Total Sheets: {len(wb.sheetnames)}")
    print(f"Sheet Names: {wb.sheetnames}\n")

    for idx, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        print(f"[{idx}] Tab: '{name}' | Rows: {ws.max_row} | Cols: {ws.max_column}")
        
        # Print top 3 rows
        for r in range(1, min(4, ws.max_row + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(14, ws.max_column + 1))]
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            print(f"    Row {r}: {row_vals}")
        print("-" * 60)

if __name__ == "__main__":
    inspect()
