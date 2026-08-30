#!/usr/bin/env python3
"""
Regenerates and cleans '07_Keywords_Match_Config' in RSSFeedChecker_Master_Guide_and_Data.xlsx.
Eliminates all noisy keywords, overly broad trial acronyms, loose English verbs, and ambiguous terms.
Establishes precise, disambiguated biopharma intelligence tracking rules across 6 dimensions.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

XLSX_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "RSSFeedChecker_Master_Guide_and_Data.xlsx"))

# -----------------------------------------------------------------------------
# PRISTINE, HIGH-PRECISION RULES (Zero Noise, Ambiguity-Gated)
# -----------------------------------------------------------------------------
CLEANED_KEYWORD_RULES = [
    # -------------------------------------------------------------------------
    # 1. DRUG ASSETS & PIPELINE (Tiers 1 & 2)
    # -------------------------------------------------------------------------
    (
        "1. Drug Asset & Pipeline",
        "CagriSema (Novo Nordisk)",
        "CagriSema, cagrilintide, semaglutide combination, NN9535, REDEFINE-1, REDEFINE-2, REDEFINE-3",
        "trial OR phase OR efficacy OR topline OR weight loss OR obesity OR primary endpoint OR FDA",
        "counterfeit, compounding, med spa, weight loss clinic, diet blog",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Tracks Novo Nordisk's dual amylin/GLP-1 combination against Mounjaro/Zepbound in Phase 3 readouts."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Tirzepatide (Eli Lilly)",
        "tirzepatide, Mounjaro, Zepbound, SURMOUNT-1, SURMOUNT-2, SURMOUNT-3, SURMOUNT-4, SURMOUNT-5, SURPASS-1, SURPASS-2, SURPASS-3, SURPASS-4, SURPASS-5, SURPASS-6, LY3298176, dual GIP/GLP-1",
        "FDA OR approval OR phase 3 OR topline OR label expansion OR trial OR shortage OR Lilly",
        "stock options, options market, class action, coupon code, SURPASS-IPF, celea, deupirfenidone, veterinary",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Monitors Eli Lilly's market-leading dual GIP/GLP-1 receptor agonist for approvals, trial readouts, and supply."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Orforglipron (Eli Lilly)",
        "orforglipron, LY3502970, oral GLP-1, non-peptide GLP-1, ATTAIN-1, ATTAIN-2, ACHIEVE-1, ACHIEVE-2, ACHIEVE-3, ACHIEVE-4",
        "Phase 3 OR trial OR data OR tolerability OR weight loss OR FDA OR Lilly OR incretin",
        "stock options, lock-up agreement, veterinary, canine, beverage, splash beverage",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Tracks next-gen daily oral non-peptide GLP-1 agonist in Phase 3 global obesity trials."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Keytruda (Merck & Co)",
        "Keytruda, pembrolizumab, MK-3475, anti-PD-1, KEYNOTE-001, KEYNOTE-189, KEYNOTE-522, KEYNOTE-671, KEYNOTE-A18, subcutaneous pembrolizumab",
        "FDA OR approval OR Phase 3 OR OS OR PFS OR overall survival OR CRL OR label expansion OR Merck",
        "stock options, options market, patent cliff litigation generic, legal settlement",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks oncology blockbuster Keytruda across clinical readouts, regulatory filings, and subcutaneous formulations."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Enhertu (Daiichi Sankyo / AstraZeneca)",
        "Enhertu, trastuzumab deruxtecan, T-DXd, DS-8201, DESTINY-Breast, DESTINY-Lung, DESTINY-PanTumor",
        "FDA OR approval OR Phase 3 OR readout OR HER2-low OR HER2-ultralow OR survival OR safety OR AstraZeneca",
        "stock options, share buy-back, stock pump",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors breakthrough HER2-directed antibody-drug conjugate (ADC) clinical milestones and line extensions."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Leqembi & Kisunla (Alzheimer's)",
        "Leqembi, lecanemab, BAN2401, Kisunla, donanemab, TRAILBLAZER-ALZ, CLARITY-AD",
        "FDA OR EMA OR approval OR CMS OR coverage OR ARIA OR amyloid OR Phase 3 OR blood biomarker OR Biogen OR Eisai OR Lilly",
        "stock options, Zacks rank, options activity",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Neurology & Neurodegenerative Desk",
        "TRUE",
        "Tracks anti-amyloid monoclonal antibodies in Alzheimer's (Eisai/Biogen and Eli Lilly) for approvals, safety (ARIA), and CMS reimbursement."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Dupixent (Sanofi / Regeneron)",
        "Dupixent, dupilumab, anti-IL-4R, REGN668, SAR440340, BOREAS trial, NOTUS trial",
        "COPD OR atopic dermatitis OR asthma OR eosinophilic OR prurigo OR sBLA OR FDA OR Phase 3 OR Regeneron OR Sanofi",
        "stock options, lock-up, coupon promo",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Immunology & Respiratory Desk",
        "TRUE",
        "Monitors blockbuster Type 2 inflammatory biologic for major expansions into COPD, pediatric indications, and competitive threats."
    ),
    (
        "1. Drug Asset & Pipeline",
        "Casgevy & Lyfgenia (CRISPR / Sickle Cell)",
        "Casgevy, exa-cel, exagamglogene autotemcel, Lyfgenia, lovo-cel, CTX001, bb1111",
        "sickle cell OR beta-thalassemia OR CRISPR OR gene therapy OR launch OR reimbursement OR Vertex OR Bluebird",
        "stock options, option volume, penny stock",
        "🟡 Tier 2 (Daily Briefing)",
        "Cell & Gene Therapy Desk",
        "TRUE",
        "Monitors commercial launch, hospital activation, patient intake, and long-term durability of first approved CRISPR gene editing therapies."
    ),

    # -------------------------------------------------------------------------
    # 2. MODALITIES & BIOLOGY (Tiers 1 & 2)
    # -------------------------------------------------------------------------
    (
        "2. Modality & Biology",
        "Antibody-Drug Conjugates (ADC)",
        "Antibody-Drug Conjugate, antibody drug conjugates, ADC payload, topoisomerase payload, MMAE payload, DXd payload, exatecan payload, bispecific ADC, target-to-toxin",
        "oncology OR solid tumor OR phase OR clinical OR trial OR toxicity OR payload OR linker OR HER2 OR TROP2 OR Claudin",
        "analog-to-digital, audio converter, converter, analog digital, ADC software, apple developer, darts championship",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks competitive clinical trials and licensing deals in the high-growth ADC landscape while filtering out engineering/electronics noise."
    ),
    (
        "2. Modality & Biology",
        "Targeted Radioligand Therapy (RLT)",
        "radiopharmaceutical, radioligand therapy, lutetium-177, actinium-225, lead-212, copper-64, PSMA radioligand, alpha emitter, beta emitter",
        "clinical OR patient OR phase OR trial OR FDA OR supply OR isotope OR cancer OR oncology",
        "nuclear energy, uranium mining, nuclear power plant, weapon",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors targeted radiopharmaceutical pipelines (Novartis, Point, RayzeBio/BMS, Mariana/Lilly) and clinical supply chains."
    ),
    (
        "2. Modality & Biology",
        "Bispecific Antibodies & T-Cell Engagers",
        "bispecific antibody, T-cell engager, BiTE antibody, CD3 bispecific, CD20xCD3, DLL3 bispecific, BCMAxCD3, trispecific engager",
        "Phase OR clinical OR trial OR oncology OR hematology OR cytokine release OR CRS OR FDA",
        "computer hardware, electronic components, binary",
        "🟡 Tier 2 (Daily Briefing)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks dual-targeting antibodies in solid tumors and hematology with safety/CRS monitoring."
    ),
    (
        "2. Modality & Biology",
        "In Vivo Cell & Gene Therapy",
        "in vivo CAR-T, in vivo gene editing, LNP delivery platform, viral vector capsid, AAV capsid engineering, non-viral delivery, mRNA editing",
        "preclinical OR IND OR Phase 1 OR clinical OR delivery OR target OR efficacy OR gene therapy",
        "stock options, retail investment, crypto",
        "🟡 Tier 2 (Daily Briefing)",
        "Cell & Gene Therapy Desk",
        "TRUE",
        "Tracks cutting-edge in vivo delivery platforms aimed at eliminating ex vivo manufacturing bottlenecks."
    ),
    (
        "2. Modality & Biology",
        "Targeted Protein Degradation (PROTACs)",
        "targeted protein degrader, PROTAC molecule, molecular glue degrader, E3 ligase degrader, cereblon binder, VHL degrader, ARV-110, ARV-471, CFT7455, NX-2127",
        "clinical OR Phase OR trial OR oncology OR oral OR degrader OR drug OR Arvinas OR Kymera OR Nurix",
        "glue manufacturing, adhesive retail, school glue, super glue, craft glue",
        "🟡 Tier 2 (Daily Briefing)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Monitors molecular glues and PROTAC clinical advancements (Arvinas, Kymera, Nurix, BMS)."
    ),

    # -------------------------------------------------------------------------
    # 3. REGULATORY CATALYSTS (Tier 1 & 2)
    # -------------------------------------------------------------------------
    (
        "3. Regulatory Catalysts",
        "FDA Approval & PDUFA Decisions",
        "PDUFA target date, FDA approval, FDA approved, FDA approves, sBLA approval, NDA approval, Complete Response Letter, FDA CRL, tentative FDA approval",
        "FDA OR agency OR pharmaceutical OR drug OR therapeutic OR approval OR oncology OR medicine",
        "food recall, pet food, consumer cosmetic, medical device drill, certificate revocation list",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "High-priority alert for all FDA marketing authorization decisions, PDUFA target dates, and Complete Response Letters."
    ),
    (
        "3. Regulatory Catalysts",
        "Expedited Designation Awards",
        "Breakthrough Therapy Designation, Breakthrough Designation, Fast Track Designation, FDA Fast Track, Priority Review Voucher, Accelerated Approval, RMAT designation, EMA PRIME designation",
        "FDA OR EMA OR granted OR drug OR clinical OR therapeutic OR biotech",
        "Amazon Prime, prime video, fast track train, breakthrough idea, marketing buzzword",
        "🟡 Tier 2 (Daily Briefing)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Tracks FDA/EMA regulatory speed pathways that compress time-to-market for competitor assets."
    ),
    (
        "3. Regulatory Catalysts",
        "EMA CHMP Opinions & Marketing Authorizations",
        "CHMP positive opinion, CHMP negative opinion, European Commission marketing authorisation, EMA validation, Type II variation approval",
        "EMA OR European Medicines Agency OR committee OR recommendation OR medicinal product",
        "stock options, cryptocurrency, forex",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Captures European regulatory milestones from monthly CHMP meetings."
    ),
    (
        "3. Regulatory Catalysts",
        "Clinical Holds & Safety Alerts",
        "clinical hold, partial clinical hold, FDA clinical hold, black box warning, boxed warning, FDA REMS, study paused due to toxicity, trial suspended for safety",
        "FDA OR trial OR clinical OR patient OR safety OR investigation OR toxicity",
        "holding company, holding pattern, sports box, court trial, traffic safety alert",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Regulatory & Strategy Desk",
        "TRUE",
        "Instant critical alert for clinical study pauses, toxicity holds, or FDA safety labeling revisions."
    ),

    # -------------------------------------------------------------------------
    # 4. INDICATIONS & RADARS (Tier 1 & 2)
    # -------------------------------------------------------------------------
    (
        "4. Indication & Disease",
        "Obesity, T2D & MASH / NASH",
        "incretin therapy, GLP-1 receptor agonist, GIP receptor, dual incretin, triple agonist, MASH fibrosis, NASH trial",
        "trial OR Phase OR FDA OR approval OR efficacy OR topline OR weight loss OR diabetes OR liver fibrosis",
        "diet blog, fitness workout, spa, cosmetic weight loss, gym program",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Metabolic & Obesity Desk",
        "TRUE",
        "Monitors the multi-billion-dollar metabolic competitive landscape across injectable and oral incretin and non-incretin therapies."
    ),
    (
        "4. Indication & Disease",
        "Non-Small Cell Lung Cancer (NSCLC)",
        "non-small cell lung cancer, NSCLC, EGFR mutation, KRAS G12C, ALK rearrangement, ROS1 fusion, PD-L1 high, MET exon 14",
        "trial OR Phase OR FDA OR PFS OR OS OR overall survival OR line of therapy OR oncology",
        "smoking cessation patch, air pollution retail, tobacco",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks frontline and refractory NSCLC targeted therapies and immune checkpoint combinations."
    ),
    (
        "4. Indication & Disease",
        "Multiple Myeloma & Hematologic Malignancies",
        "multiple myeloma, relapsed refractory multiple myeloma, RRMM, BCMA CAR-T, BCMA bispecific, GPRC5D bispecific, quadruplet myeloma",
        "trial OR Phase OR FDA OR MRD negative OR progression-free survival OR CR OR remission",
        "stock options, option alert, trading signal",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Oncology & Immuno-Oncology Desk",
        "TRUE",
        "Tracks competitive frontline quadruplets, BCMA bispecifics (Tecvayli, Elrexfio), and CAR-T therapies."
    ),
    (
        "4. Indication & Disease",
        "Inflammatory Bowel Disease (IBD: Crohn's & UC)",
        "anti-TL1A antibody, anti-IL-23 antibody, JAK1 inhibitor in IBD, S1P modulator in UC, moderate to severe Crohn's, active Ulcerative Colitis",
        "trial OR Phase OR clinical OR remission OR endoscopic OR FDA OR approval OR mucosal healing",
        "diet supplement, herbal remedy, holistic tea",
        "🟡 Tier 2 (Daily Briefing)",
        "Immunology & Respiratory Desk",
        "TRUE",
        "Monitors competitive advanced biologics and oral therapies in moderate-to-severe UC and Crohn's."
    ),
    (
        "4. Indication & Disease",
        "Rare Neuromuscular (DMD & SMA)",
        "Duchenne Muscular Dystrophy, Duchenne dystrophy, microdystrophin gene, exon skipping therapy, Spinal Muscular Atrophy, SMA disease, SMN2 splicing, Spinraza, Evrysdi, Zolgensma, Elevidys",
        "trial OR Phase OR clinical OR ambulant OR dystrophin OR SMN OR gene therapy OR FDA OR NSAA score",
        "SMA solar inverter, solar panel, DMD dentist, dental medicine, DMD tools, shape memory alloy",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Rare Disease Desk",
        "TRUE",
        "Tracks high-value gene therapies and exon-skipping drugs in Duchenne and SMA, strictly filtering out solar/dental acronym noise."
    ),

    # -------------------------------------------------------------------------
    # 5. TRANSACTIONS & STRATEGY (Tier 1, 2, 3)
    # -------------------------------------------------------------------------
    (
        "5. Transactions & Strategy",
        "Biopharma M&A & Asset Acquisitions",
        "biopharma buyout, biotech acquisition, definitive agreement to acquire, to acquire biopharmaceutical, takeover offer for biotech, merger agreement with therapeutics, tender offer for shares of",
        "biotech OR biopharmaceutical OR therapeutics OR pharma OR clinical pipeline OR per share OR cash transaction",
        "real estate, commercial property, tech merger, banking acquisition, software buyout, retail chain",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Business Development & M&A Desk",
        "TRUE",
        "Detects multi-million and multi-billion-dollar biopharma buyout announcements and asset purchases."
    ),
    (
        "5. Transactions & Strategy",
        "Licensing Deals & Collaboration Biobucks",
        "exclusive license agreement, global biopharma collaboration, upfront payment of, development milestones and royalties, biobucks deal, option and license agreement",
        "biotech OR pharmaceutical OR therapeutics OR rights OR commercialization OR drug discovery",
        "software license, driver license, gaming license, music license",
        "🟡 Tier 2 (Daily Briefing)",
        "Business Development & M&A Desk",
        "TRUE",
        "Monitors early and late-stage in-licensing deals, platform collaborations, and co-development rights."
    ),
    (
        "5. Transactions & Strategy",
        "Pipeline Terminations & Strategic Reprioritizations",
        "discontinue development of, terminate clinical study, pipeline reprioritization, strategic review of pipeline, halted clinical trial, failed primary endpoint in Phase",
        "Phase OR clinical OR asset OR development OR patient OR program OR therapeutic",
        "computer program, software termination, employee termination",
        "🔴 Tier 1 (Urgent / Immediate)",
        "Business Development & M&A Desk",
        "TRUE",
        "Immediate alert when a competitor drops or discontinues a clinical development asset following trial failure or strategic shift."
    ),
    (
        "5. Transactions & Strategy",
        "Biotech Restructuring & Workforce Reductions",
        "workforce reduction in biopharma, laying off employees in biotech, corporate restructuring and cost reduction, cost-saving initiatives, runway extension into",
        "biotech OR biopharma OR pharmaceutical OR drug development OR pipeline",
        "tech layoffs, retail layoffs, manufacturing auto, banking job cuts",
        "🟢 Tier 3 (Horizon / Weekly)",
        "Business Development & M&A Desk",
        "TRUE",
        "Tracks operational distress, cash runway constraints, and corporate restructuring across the biotech ecosystem."
    ),

    # -------------------------------------------------------------------------
    # 6. GLOBAL NOISE SUPPRESSION RULES
    # -------------------------------------------------------------------------
    (
        "6. Noise Negation Filters",
        "Financial & Algorithmic Stock Spam",
        "undervalued by 10%, options market predicting, stock options trading, options volume spike, lock-up agreement ending, Zacks rank upgrade, short interest increased, share buy-back announcement, penny stock alert",
        "N/A (Global Noise Rule)",
        "N/A",
        "🚫 Noise Suppression Rule",
        "Automated Filtering Engine",
        "TRUE",
        "Universal negative keyword suppression pattern to eliminate algorithmic retail trading noise across Google News feeds."
    ),
    (
        "6. Noise Negation Filters",
        "Consumer, Food & Pet Recalls",
        "dog food recall, cat food recall, pet food recall, veterinary development, canine epilepsy, animal health program, livestock vaccine, beef contamination, organic spinach recall, cosmetic eye drops recall, med spa, weight loss clinic, counterfeit ozempic",
        "N/A (Global Noise Rule)",
        "N/A",
        "🚫 Noise Suppression Rule",
        "Automated Filtering Engine",
        "TRUE",
        "Filters out non-pharmaceutical food, agricultural, cosmetic, and veterinary notices."
    ),
]


def rebuild_keywords_tab():
    wb = openpyxl.load_workbook(XLSX_PATH)
    TAB_NAME = "07_Keywords_Match_Config"
    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]

    ws = wb.create_sheet(title=TAB_NAME)
    ws.views.sheetView[0].showGridLines = True

    # Palette
    PURPLE_HEADER = "4A235A"
    PURPLE_SUB = "5B2C6F"
    WHITE = "FFFFFF"
    fill_header = PatternFill(start_color=PURPLE_HEADER, end_color=PURPLE_HEADER, fill_type="solid")
    fill_ice = PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"), right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"), bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Category / Track Type", 24),
        ("Primary Focus Entity", 32),
        ("Synonyms, Aliases & Code Names", 45),
        ("Mandatory Context Qualifiers (Co-occurrence)", 45),
        ("Negative Exclude Terms (Negations)", 40),
        ("Alert Priority Tier", 24),
        ("Assigned CI Desk / Specialist", 28),
        ("Active Status", 14),
        ("Operational Role & Intelligence Focus", 45),
    ]

    for col_idx, (h_name, width) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = fill_header
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    for r_idx, rule in enumerate(CLEANED_KEYWORD_RULES, start=2):
        ws.cell(row=r_idx, column=1, value=rule[0]).font = font_bold
        ws.cell(row=r_idx, column=2, value=rule[1]).font = font_bold
        ws.cell(row=r_idx, column=3, value=rule[2]).font = font_code
        ws.cell(row=r_idx, column=4, value=rule[3]).font = font_code
        ws.cell(row=r_idx, column=5, value=rule[4]).font = font_code

        # Priority badge
        cell_p = ws.cell(row=r_idx, column=6, value=rule[5])
        cell_p.alignment = Alignment(horizontal="center")
        if "Tier 1" in rule[5]:
            cell_p.font = Font(name="Calibri", size=10, bold=True, color="C2185B")
            cell_p.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        elif "Tier 2" in rule[5]:
            cell_p.font = Font(name="Calibri", size=10, bold=True, color="B78103")
            cell_p.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif "Noise" in rule[5]:
            cell_p.font = Font(name="Calibri", size=10, bold=True, color="7F8C8D")
            cell_p.fill = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
        else:
            cell_p.font = font_data

        ws.cell(row=r_idx, column=7, value=rule[6]).font = font_data
        ws.cell(row=r_idx, column=8, value=rule[7]).alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=9, value=rule[8]).font = font_data

        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).border = thin_border
        ws.row_dimensions[r_idx].height = 24

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    
    from robust_fetcher import safe_save_workbook
    saved_file = safe_save_workbook(wb, XLSX_PATH)
    print(f"Tab '{TAB_NAME}' successfully rebuilt with {len(CLEANED_KEYWORD_RULES)} clean, noise-free rules in '{saved_file}'!")


if __name__ == "__main__":
    rebuild_keywords_tab()
