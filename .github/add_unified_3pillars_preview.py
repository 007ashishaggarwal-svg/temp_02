#!/usr/bin/env python3
"""
Create a sample master unified 3-Pillar Long Registry tab in RSSFeedChecker_Master_Guide_and_Data.xlsx.
Brings Publishers (Pillar 1), Companies (Pillar 2), and Indications + ClinicalTrials.gov (Pillar 3)
together into a single, unified Long Schema enriched with user configuration controls.
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
XLSX_PATH = os.path.join(WORKSPACE, "RSSFeedChecker_Master_Guide_and_Data.xlsx")

def build_unified_3pillars_sample():
    print(f"Loading workbook from {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    TAB_NAME = "Preview_03_Unified_All_Pillars"
    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]

    ws = wb.create_sheet(title=TAB_NAME)
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    PURPLE_DARK = "382D5C"
    TEAL_DARK = "0E5A5E"
    WHITE = "FFFFFF"
    BORDER_COLOR = "D0D7DE"
    ICE_BLUE = "EBF2FA"
    ICE_GREEN = "EAF5EA"

    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_purple = PatternFill(start_color=PURPLE_DARK, end_color=PURPLE_DARK, fill_type="solid")
    fill_teal = PatternFill(start_color=TEAL_DARK, end_color=TEAL_DARK, fill_type="solid")
    fill_ice_blue = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
    fill_ice_green = PatternFill(start_color=ICE_GREEN, end_color=ICE_GREEN, fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    font_code = Font(name="Consolas", size=9, color="111111")

    # Status Badges
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_green = Font(name="Calibri", size=10, bold=True, color="155724")

    fill_amber = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_amber = Font(name="Calibri", size=10, bold=True, color="856404")

    fill_blue_badge = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
    font_blue_badge = Font(name="Calibri", size=10, bold=True, color="0C5460")

    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_red = Font(name="Calibri", size=10, bold=True, color="721C24")

    fill_purple_badge = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    font_purple_badge = Font(name="Calibri", size=10, bold=True, color="6A1B9A")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    headers = [
        # Identity & Origin
        ("Entity ID", 14, fill_navy),
        ("Entity / Target Name", 28, fill_navy),
        ("Source Classification", 22, fill_navy),
        ("Pillar Origin", 24, fill_navy),
        ("Ingestion Vector / Method", 30, fill_blue),
        ("Endpoint URL / Query Definition", 50, fill_blue),
        
        # User Configuration Controls
        ("Fetch Frequency (Hours)", 22, fill_teal),
        ("Active Toggle", 18, fill_teal),
        ("Desk Route Override", 24, fill_teal),
        ("Priority Booster", 22, fill_teal),
        ("Max Items / Scan", 18, fill_teal),
        
        # Engine Telemetry & Observability
        ("Execution Status", 22, fill_purple),
        ("Last Verified Response", 28, fill_purple),
        ("Technical Notes & Bypass Configuration", 42, fill_navy),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 30

    sample_data = [
        # ---------------------------------------------------------------------
        # PILLAR 1: MEDIA & REGULATORY AGENCIES
        # ---------------------------------------------------------------------
        # FDA Press Office
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "Pillar 1: Publisher Feeds", "1. Native RSS Feed", "https://www.fda.gov/newsroom/press-announcements/rss.xml", 1, "Active", "Regulatory & Strategy", "Always Tier 1 (Urgent)", 25, "ACTIVE (Primary)", "HTTP 200 OK (12 approvals)", "Official FDA regulatory press stream. Hourly sweep."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "Pillar 1: Publisher Feeds", "2. XML Sitemap Index", "https://www.fda.gov/sitemap.xml", 12, "Standby (Backup)", "Regulatory & Strategy", "Always Tier 1 (Urgent)", 20, "STANDBY (Backup)", "HTTP 200 OK (Gov Sitemap)", "Standby gov sitemap."),
        ("REG_FDA", "FDA Press Announcements", "Regulatory Agency", "Pillar 1: Publisher Feeds", "3. openFDA Drug API", "https://api.fda.gov/drug/event.json", 6, "Standby (Backup)", "Regulatory & Strategy", "Always Tier 1 (Urgent)", 50, "STANDBY (Backup)", "HTTP 200 OK (REST API)", "Structured label/event API."),

        # STAT News
        ("PUB_STAT", "STAT News", "Industry Trade Media", "Pillar 1: Publisher Feeds", "1. Native RSS Feed", "https://www.statnews.com/feed/", 2, "Active", "Auto (Sheet 07 Rules)", "Default", 30, "ACTIVE (Primary)", "HTTP 200 OK (30 stories)", "High-velocity media stream. 2-hour scan cycle."),
        ("PUB_STAT", "STAT News", "Industry Trade Media", "Pillar 1: Publisher Feeds", "2. XML Sitemap Index", "https://www.statnews.com/sitemap.xml", 24, "Standby (Backup)", "Auto (Sheet 07 Rules)", "Default", 20, "STANDBY (Backup)", "HTTP 200 OK (News Sitemap)", "Standby news sitemap."),

        # Endpoints News
        ("PUB_ENDP", "Endpoints News", "Industry Trade Media", "Pillar 1: Publisher Feeds", "1. Native RSS Feed", "https://endpts.com/feed/", 2, "Active", "Auto (Sheet 07 Rules)", "Default", 25, "ACTIVE (Primary)", "HTTP 200 OK (20 stories)", "R&D and VC coverage. 2-hour scan cycle."),

        # ---------------------------------------------------------------------
        # PILLAR 2: CORPORATE BIOPHARMA DRUGMAKERS
        # ---------------------------------------------------------------------
        # Novartis
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "1. Native RSS Feed", "https://www.novartis.com/news/media-releases/feed", 2, "Active", "Auto (Sheet 07 Rules)", "Default", 15, "ACTIVE (Primary)", "HTTP 200 OK (22 releases)", "Native RSS discovered & active."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "2. XML Sitemap Index", "https://www.novartis.com/sitemap.xml", 12, "Standby (Backup)", "Auto (Sheet 07 Rules)", "Default", 10, "STANDBY (Backup)", "HTTP 200 OK (<lastmod> active)", "Capped at 20 sub-sitemaps."),
        ("COMP_NVS", "Novartis", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Anovartis.com+press&hl=en-US", 4, "Standby (Backup)", "Auto (Sheet 07 Rules)", "Default", 10, "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Bypasses corporate CDN."),

        # Eli Lilly (Akamai WAF -> GNews Fallback Active)
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "1. Native RSS Feed", "https://investor.lilly.com/rss/news-releases.xml", 1, "Degraded", "Auto (Sheet 07 Rules)", "Always Tier 1 (Urgent)", 20, "BLOCKED (Akamai WAF)", "HTTP 403 Forbidden", "Akamai challenge blocks cloud IP."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "2. XML Sitemap Index", "https://www.lilly.com/sitemap.xml", 6, "Degraded", "Auto (Sheet 07 Rules)", "Always Tier 1 (Urgent)", 10, "BLOCKED (WAF Challenge)", "HTTP 403 Forbidden", "Sitemap endpoint behind WAF."),
        ("COMP_LLY", "Eli Lilly and Company", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Alilly.com+press&hl=en-US", 1, "Active", "Auto (Sheet 07 Rules)", "Always Tier 1 (Urgent)", 20, "ACTIVE (Fallback)", "HTTP 200 OK (14 releases)", "Active primary route. 100% reliable bypass."),

        # Pfizer (Dead RSS -> Sitemap Active)
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "1. Native RSS Feed", "https://www.pfizer.com/rss.xml", 2, "Degraded", "Auto (Sheet 07 Rules)", "Default", 15, "DEGRADED (404 Dead)", "HTTP 404 Not Found (0 items)", "Feed removed during CMS migration."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "2. XML Sitemap Index", "https://www.pfizer.com/sitemap.xml", 2, "Active", "Auto (Sheet 07 Rules)", "Default", 15, "ACTIVE (Fallback)", "HTTP 200 OK (8 fresh PRs)", "Active primary route. Reads /news/press-release/<lastmod>."),
        ("COMP_PFE", "Pfizer", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "3. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Apfizer.com+press&hl=en-US", 4, "Standby (Backup)", "Auto (Sheet 07 Rules)", "Default", 10, "STANDBY (Backup)", "HTTP 200 OK (GNews Search)", "Standby mirror."),

        # Alnylam (No RSS -> GNews Active)
        ("COMP_ALNY", "Alnylam Pharmaceuticals", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "1. Native RSS Feed", "None (No RSS Exposed)", 4, "Paused (Muted)", "Auto (Sheet 07 Rules)", "Default", 10, "STANDBY (Backup)", "None", "Company exposes 0 public RSS feeds."),
        ("COMP_ALNY", "Alnylam Pharmaceuticals", "Corporate Drugmaker", "Pillar 2: Company Newsroom", "2. Google News 'site:' Mirror", "https://news.google.com/rss/search?q=site%3Aalnylam.com+press&hl=en-US", 2, "Active", "Auto (Sheet 07 Rules)", "Default", 15, "ACTIVE (Primary)", "HTTP 200 OK (6 releases)", "Active primary route."),

        # ---------------------------------------------------------------------
        # PILLAR 3: THERAPEUTIC INDICATION RADARS & CLINICALTRIALS.GOV
        # ---------------------------------------------------------------------
        # Obesity & GLP-1 Incretins
        ("IND_OBES", "Obesity & Weight Loss", "Therapy Area Radar", "Pillar 3: Indication Radar", "1. Google News Broad Query", "q=(semaglutide+OR+tirzepatide+OR+retatrutide)+AND+(obesity+OR+weight-loss)", 2, "Active", "Clinical & Pipeline", "Default", 40, "ACTIVE (Primary)", "HTTP 200 OK (35 articles)", "Broad therapeutic landscape monitoring."),
        ("IND_OBES", "Obesity & Weight Loss", "Therapy Area Radar", "Pillar 3: Indication Radar", "2. Google News Regulatory", "q=(obesity+OR+GLP-1+OR+incretin)+AND+(FDA+OR+approval+OR+PDUFA+OR+CRL)", 1, "Active", "Regulatory & Strategy", "Always Tier 1 (Urgent)", 20, "ACTIVE (Primary)", "HTTP 200 OK (10 articles)", "Regulatory decision radar. Hourly cadence."),
        ("IND_OBES", "Obesity & Weight Loss", "Clinical Trials Registry", "Pillar 3: ClinicalTrials.gov", "3. ClinicalTrials.gov API v2", "https://clinicaltrials.gov/api/v2/studies?query.cond=Obesity&filter.advanced=AREA[FirstPostDate]RANGE[2026-08-20,MAX]", 6, "Active", "Clinical & Pipeline", "Always Tier 1 (Urgent)", 25, "ACTIVE (Primary)", "HTTP 200 OK (14 new trials)", "REST API v2. Absorbs newly registered studies."),

        # HER2+ Oncology
        ("IND_HER2", "HER2+ Breast & Gastric Cancer", "Therapy Area Radar", "Pillar 3: Indication Radar", "1. Google News Broad Query", "q=(HER2+OR+Enhertu+OR+T-DXd)+AND+(breast+OR+gastric+OR+GEA)", 2, "Active", "Oncology & Immunotherapy", "Default", 30, "ACTIVE (Primary)", "HTTP 200 OK (22 articles)", "HER2 antibody-drug conjugate radar."),
        ("IND_HER2", "HER2+ Breast & Gastric Cancer", "Clinical Trials Registry", "Pillar 3: ClinicalTrials.gov", "3. ClinicalTrials.gov API v2", "https://clinicaltrials.gov/api/v2/studies?query.cond=HER2-Positive+Breast+Cancer", 6, "Active", "Oncology & Immunotherapy", "Always Tier 1 (Urgent)", 20, "ACTIVE (Primary)", "HTTP 200 OK (8 new trials)", "REST API v2 new HER2 interventional trials."),

        # Alzheimer's Disease
        ("IND_ALZH", "Alzheimer's Disease", "Therapy Area Radar", "Pillar 3: Indication Radar", "1. Google News Broad Query", "q=(lecanemab+OR+donanemab+OR+amyloid)+AND+(Alzheimer+OR+cognitive)", 4, "Active", "Neurology & Rare Disease", "Default", 25, "ACTIVE (Primary)", "HTTP 200 OK (18 articles)", "Anti-amyloid & tau therapeutic tracking."),
        ("IND_ALZH", "Alzheimer's Disease", "Clinical Trials Registry", "Pillar 3: ClinicalTrials.gov", "3. ClinicalTrials.gov API v2", "https://clinicaltrials.gov/api/v2/studies?query.cond=Alzheimer+Disease", 6, "Active", "Neurology & Rare Disease", "Always Tier 1 (Urgent)", 20, "ACTIVE (Primary)", "HTTP 200 OK (6 new trials)", "REST API v2 Alzheimer's registrations.")
    ]

    current_entity = ""
    is_alt = False

    for r_idx, row in enumerate(sample_data, start=2):
        if row[1] != current_entity:
            current_entity = row[1]
            is_alt = not is_alt

        row_fill = fill_ice_blue if is_alt else PatternFill(fill_type=None)

        # Col 1: Entity ID
        ws.cell(row=r_idx, column=1, value=row[0]).font = font_bold
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")

        # Col 2: Entity Name
        ws.cell(row=r_idx, column=2, value=row[1]).font = font_bold

        # Col 3: Classification
        ws.cell(row=r_idx, column=3, value=row[2]).font = font_data

        # Col 4: Pillar Origin Badge
        cell_p = ws.cell(row=r_idx, column=4, value=row[3])
        cell_p.font = font_bold
        if "Publisher" in row[3]:
            cell_p.font = Font(name="Calibri", size=9, bold=True, color="155724")
        elif "Company" in row[3]:
            cell_p.font = Font(name="Calibri", size=9, bold=True, color="1F4E79")
        else:
            cell_p.font = Font(name="Calibri", size=9, bold=True, color="6A1B9A")

        # Col 5: Ingestion Vector
        ws.cell(row=r_idx, column=5, value=row[4]).font = font_bold

        # Col 6: URL / Query
        cell_url = ws.cell(row=r_idx, column=6, value=row[5])
        if str(row[5]).startswith("http"):
            cell_url.font = font_link
        else:
            cell_url.font = font_code

        # Col 7: Fetch Frequency (Hours)
        cell_freq = ws.cell(row=r_idx, column=7, value=f"{row[6]}h")
        cell_freq.font = font_bold
        cell_freq.alignment = Alignment(horizontal="center")
        if row[6] == 1:
            cell_freq.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            cell_freq.font = Font(name="Calibri", size=10, bold=True, color="E65100")

        # Col 8: Active Toggle
        cell_act = ws.cell(row=r_idx, column=8, value=row[7])
        cell_act.alignment = Alignment(horizontal="center")
        if row[7] == "Active":
            cell_act.font = font_green
            cell_act.fill = fill_green
        elif "Standby" in row[7]:
            cell_act.font = font_amber
            cell_act.fill = fill_amber
        else:
            cell_act.font = font_red
            cell_act.fill = fill_red

        # Col 9: Desk Override
        cell_desk = ws.cell(row=r_idx, column=9, value=row[8])
        cell_desk.font = font_data
        cell_desk.alignment = Alignment(horizontal="center")

        # Col 10: Priority Booster
        cell_boost = ws.cell(row=r_idx, column=10, value=row[9])
        cell_boost.alignment = Alignment(horizontal="center")
        if "Always Tier 1" in row[9]:
            cell_boost.font = Font(name="Calibri", size=10, bold=True, color="C2185B")
            cell_boost.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        else:
            cell_boost.font = font_data

        # Col 11: Max Items
        cell_max = ws.cell(row=r_idx, column=11, value=row[10])
        cell_max.font = font_data
        cell_max.alignment = Alignment(horizontal="center")

        # Col 12: Execution Status
        cell_st = ws.cell(row=r_idx, column=12, value=row[11])
        cell_st.alignment = Alignment(horizontal="center")
        if "ACTIVE (Primary)" in row[11]:
            cell_st.font = font_green
            cell_st.fill = fill_green
        elif "ACTIVE (Fallback)" in row[11]:
            cell_st.font = font_blue_badge
            cell_st.fill = fill_blue_badge
        elif "STANDBY" in row[11]:
            cell_st.font = font_amber
            cell_st.fill = fill_amber
        else:
            cell_st.font = font_red
            cell_st.fill = fill_red

        # Col 13: Last Verified Response
        ws.cell(row=r_idx, column=13, value=row[12]).font = font_code

        # Col 14: Technical Notes
        ws.cell(row=r_idx, column=14, value=row[13]).font = font_data

        for col_c in range(1, 15):
            cell_c = ws.cell(row=r_idx, column=col_c)
            cell_c.border = thin_border
            if cell_c.fill.fill_type is None and row_fill.fill_type is not None:
                cell_c.fill = row_fill

        ws.row_dimensions[r_idx].height = 24

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(sample_data)+1}"

    # Save workbook
    wb.save(XLSX_PATH)
    print(f"✅ Successfully created '{TAB_NAME}' with 14 rich columns in {XLSX_PATH}!")

if __name__ == "__main__":
    build_unified_3pillars_sample()
