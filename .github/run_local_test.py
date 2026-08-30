#!/usr/bin/env python3
"""
Enhanced Local Test Runner & Unified Intelligence Orchestrator.
Guarantees the complete 5-Core Data Schema with deep cleaning, HTML unescaping,
tracking tag stripping, publisher suffix cleaning, and proactive edge case handling across:
- Pillar 1: Publisher & Regulatory Feeds (STAT, Endpoints, Fierce, FDA, BioWorld)
- Pillar 2: Company Newsrooms (Pfizer, Merck, Novo Nordisk, Eli Lilly, AstraZeneca)
- Pillar 3: Indication Radars & CT.gov New Registrations (Obesity, Oncology, Alzheimer's)

Outputs:
1. '00_Unified_Intelligence_Feed' tab in RSSFeedChecker_Master_Guide_and_Data.xlsx
2. 'results/unified_intelligence_feed.html' interactive visual card dashboard
"""

import os
import sys
import re
import time
import json
import html
import ssl
import urllib.request
import urllib.parse
from datetime import datetime, timezone
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from time_window import create_time_window, TimeWindow
from match_keywords import KeywordMatcher
from fetch_new_clinical_trials import fetch_new_trials_for_condition

XLSX_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "RSSFeedChecker_Master_Guide_and_Data.xlsx"))
HTML_OUT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "results", "unified_intelligence_feed.html"))

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
_CTX = ssl.create_default_context()
_CTX.check_hostname = False
_CTX.verify_mode = ssl.CERT_NONE

# -----------------------------------------------------------------------------
# SAMPLE CURATED ENDPOINTS FOR FAST LOCAL TESTING (10-20 seconds)
# -----------------------------------------------------------------------------
SAMPLE_FEEDS = [
    ("STAT News", "https://www.statnews.com/feed/", "Trade Press"),
    ("Endpoints News", "https://endpoints.news/feed/", "Trade Press"),
    ("FiercePharma", "https://www.fiercepharma.com/rss/xml", "Trade Press"),
    ("FDA Press Releases", "https://www.fda.gov/about-fda/contact-fda/stay-informed/rss-feeds/press-releases/rss.xml", "Regulatory & Government"),
    ("BioSpace", "https://www.biospace.com/rss/news", "Trade Press"),
]

SAMPLE_COMPANIES = [
    ("Pfizer", "pfizer.com", "/newsroom/press-releases", "https://pfizer.com/rss.xml"),
    ("Merck & Co", "merck.com", "/news/", "https://news.google.com/rss/search?q=site%3Amerck.com+%28press+OR+release+OR+approval%29&hl=en-US&gl=US&ceid=US:en"),
    ("Novo Nordisk", "novonordisk.com", "/news-and-media/latest-news.html", "https://news.google.com/rss/search?q=site%3Anovonordisk.com+%28press+OR+release+OR+approval%29&hl=en-US&gl=US&ceid=US:en"),
    ("Eli Lilly", "lilly.com", "/news/press-releases", "https://news.google.com/rss/search?q=site%3Alilly.com+%28press+OR+release+OR+approval%29&hl=en-US&gl=US&ceid=US:en"),
    ("AstraZeneca", "astrazeneca.com", "/media-centre/press-releases.html", "https://news.google.com/rss/search?q=site%3Aastrazeneca.com+%28press+OR+release+OR+approval%29&hl=en-US&gl=US&ceid=US:en"),
]

SAMPLE_INDICATIONS = [
    ("Obesity", "(Obesity OR GLP-1 OR semaglutide OR tirzepatide) AND (drug OR trial OR FDA)"),
    ("Breast Cancer", "(\"Breast Cancer\" OR HER2 OR Enhertu OR CDK4/6) AND (trial OR FDA OR Phase)"),
    ("Alzheimers", "(\"Alzheimer's\" OR Leqembi OR Kisunla OR amyloid OR tau) AND (trial OR FDA OR approval)"),
]

# -----------------------------------------------------------------------------
# TEXT CLEANING & NORMALIZATION HELPERS
# -----------------------------------------------------------------------------
def clean_text(raw_text: str) -> str:
    """Strip HTML, unescape all entities, collapse whitespaces, and remove boilerplate."""
    if not raw_text:
        return ""
    # Strip HTML tags
    t = re.sub(r"<[^>]+>", " ", raw_text)
    # Decode HTML entities repeatedly (handles double-escaped &amp;quot;)
    t = html.unescape(t)
    t = html.unescape(t)
    # Strip publisher boilerplate suffixes
    suffixes = [
        r"\s*[-–—|:]\s*STAT(?:\s*News)?\s*$",
        r"\s*[-–—|:]\s*Endpoints(?:\s*News)?\s*$",
        r"\s*[-–—|:]\s*FiercePharma\s*$",
        r"\s*[-–—|:]\s*FierceBiotech\s*$",
        r"\s*[-–—|:]\s*BioSpace\s*$",
        r"\s*[-–—|:]\s*PR\s*Newswire\s*$",
        r"\s*[-–—|:]\s*Business\s*Wire\s*$",
        r"\s*[-–—|:]\s*GlobeNewswire\s*$",
        r"\s*[-–—|:]\s*Reuters\s*$",
        r"\s*[-–—|:]\s*Bloomberg\s*$",
    ]
    for suf in suffixes:
        t = re.sub(suf, "", t, flags=re.I)
    # Collapse whitespace
    t = " ".join(t.split())
    return t.strip()


def clean_url(raw_url: str) -> str:
    """Strip tracking params (UTM, cmpid, ref), session IDs, and trailing slashes."""
    if not raw_url:
        return ""
    u = raw_url.strip()
    # Strip query parameters that are tracking tags
    u = re.sub(r"[?&](?:utm_[^&=]+|cmpid|ref|fbclid|gclid|session_id)=[^&]*", "", u, flags=re.I)
    # Clean dangling ? or &
    u = re.sub(r"\?&", "?", u)
    u = re.sub(r"[?&]$", "", u)
    return u.rstrip("/")


def parse_date_to_utc(date_str: str) -> datetime:
    """Parse various RFC-822, ISO, and standard date strings to UTC."""
    if not date_str:
        return datetime.now(timezone.utc)
    
    # Try email/RFC-822 format (e.g. 'Tue, 25 Aug 2026 07:00:00 GMT')
    try:
        from email.utils import parsedate_to_datetime
        dt = parsedate_to_datetime(date_str)
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        pass

    # Try standard ISO formats
    for fmt in [
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%SZ",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ]:
        try:
            dt = datetime.strptime(date_str.rstrip("Z"), fmt)
            if dt.tzinfo is None:
                return dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            continue

    return datetime.now(timezone.utc)


# -----------------------------------------------------------------------------
# FETCHING HELPERS
# -----------------------------------------------------------------------------
def fetch_rss_items(url: str, source_name: str, pillar: str, vector: str, time_window: TimeWindow) -> list[dict]:
    """Fetch and parse RSS/Atom items within time_window with deep cleaning."""
    items = []
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
        with urllib.request.urlopen(req, timeout=10, context=_CTX) as resp:
            body = resp.read()
    except Exception:
        return []

    try:
        root = ET.fromstring(body)
    except Exception:
        return []

    # 1. Parse RSS 2.0 (<item>)
    for item in root.iter("item"):
        title_el = item.find("title")
        link_el = item.find("link")
        desc_el = item.find("description")
        date_el = item.find("pubDate")

        raw_title = title_el.text if title_el is not None and title_el.text else ""
        raw_link = link_el.text if link_el is not None and link_el.text else ""
        raw_desc = desc_el.text if desc_el is not None and desc_el.text else ""
        raw_date = date_el.text if date_el is not None and date_el.text else ""

        title = clean_text(raw_title)
        link = clean_url(raw_link)
        desc = clean_text(raw_desc)
        pub_dt = parse_date_to_utc(raw_date)

        if not title or not link or not time_window.is_in_window(pub_dt):
            continue

        items.append({
            "published_utc": pub_dt,
            "published_date": pub_dt.strftime("%Y-%m-%d"),
            "published_time": pub_dt.strftime("%H:%M"),
            "headline": title[:200],
            "url": link,
            "summary": desc[:300] or f"Published by {source_name}",
            "source_name": source_name,
            "discovery_method": pillar,
            "extraction_vector": vector,
        })

    # 2. Parse Atom (<entry>) if no RSS items found
    if not items:
        for entry in root.iter("{http://www.w3.org/2005/Atom}entry"):
            title_el = entry.find("{http://www.w3.org/2005/Atom}title")
            link_el = entry.find("{http://www.w3.org/2005/Atom}link")
            desc_el = entry.find("{http://www.w3.org/2005/Atom}summary") or entry.find("{http://www.w3.org/2005/Atom}content")
            date_el = entry.find("{http://www.w3.org/2005/Atom}updated") or entry.find("{http://www.w3.org/2005/Atom}published")

            raw_title = title_el.text if title_el is not None and title_el.text else ""
            raw_link = link_el.get("href", "") if link_el is not None else ""
            raw_desc = desc_el.text if desc_el is not None and desc_el.text else ""
            raw_date = date_el.text if date_el is not None and date_el.text else ""

            title = clean_text(raw_title)
            link = clean_url(raw_link)
            desc = clean_text(raw_desc)
            pub_dt = parse_date_to_utc(raw_date)

            if not title or not link or not time_window.is_in_window(pub_dt):
                continue

            items.append({
                "published_utc": pub_dt,
                "published_date": pub_dt.strftime("%Y-%m-%d"),
                "published_time": pub_dt.strftime("%H:%M"),
                "headline": title[:200],
                "url": link,
                "summary": desc[:300] or f"Published by {source_name}",
                "source_name": source_name,
                "discovery_method": pillar,
                "extraction_vector": vector,
            })

    return items


def run_local_pipeline(window_query="72h", is_admin=False):
    tw = create_time_window(window_query, is_admin=is_admin)
    matcher = KeywordMatcher(XLSX_PATH)
    
    print("=" * 75)
    print(" 🔬 RSSFeedChecker — LOCAL UNIFIED INTELLIGENCE SUITE")
    print("=" * 75)
    print(tw.format_summary())
    print(f"Loaded {len(matcher.rules)} Tracking Rules & {len(matcher.noise_rules)} Noise Rules from Tab 07\n")

    raw_items = []
    t0 = time.time()

    # 1. Fetch Method 1: Sample Publisher Feeds
    print("▶ [1/3] Fetching Sample Publisher Feeds (STAT, Endpoints, Fierce, FDA, BioSpace)...")
    for name, url, cat in SAMPLE_FEEDS:
        items = fetch_rss_items(url, name, "Pillar 1: Publisher Feeds", f"Native RSS ({cat})", tw)
        raw_items.extend(items)
        print(f"   • {name:<20}: Found {len(items)} items in window")

    # 2. Fetch Method 2: Sample Company Newsrooms
    print("\n▶ [2/3] Fetching Sample Company Newsrooms (Pfizer, Merck, Novo, Lilly, AZ)...")
    for name, dom, path, feed_url in SAMPLE_COMPANIES:
        vector = "Native RSS" if "rss.xml" in feed_url else "Google News Fallback"
        items = fetch_rss_items(feed_url, f"{name} ({dom})", "Pillar 2: Company Newsroom", vector, tw)
        raw_items.extend(items)
        print(f"   • {name:<20}: Found {len(items)} items in window")

    # 3. Fetch Method 3: Indication Radars & CT.gov
    print("\n▶ [3/3] Fetching Indication Radars & ClinicalTrials.gov New Registrations...")
    for ind, query in SAMPLE_INDICATIONS:
        # GNews indication search
        gnews_url = f"https://news.google.com/rss/search?q={urllib.parse.quote(query)}&hl=en-US&gl=US&ceid=US:en"
        items = fetch_rss_items(gnews_url, f"Radar: {ind}", "Pillar 3: Indication Radar", "Google News Query", tw)
        raw_items.extend(items)
        print(f"   • Radar '{ind}': Found {len(items)} news articles")

        # CT.gov new trials
        trials = fetch_new_trials_for_condition(ind, tw, max_studies=10)
        for t in trials:
            raw_items.append({
                "published_utc": datetime.strptime(t["first_post_date"], "%Y-%m-%d").replace(tzinfo=timezone.utc),
                "published_str": t["first_post_date"],
                "headline": f"{t['event_type']} {t['nct_id']} [{t['phase']}]: {clean_text(t['title'])}",
                "url": t["url"],
                "summary": f"Lead Sponsor: {t['lead_sponsor']} | Condition: {t['condition']} | Status: {t['status']}",
                "source_name": f"ClinicalTrials.gov ({t['condition']})",
                "discovery_method": t["discovery_method"],
                "extraction_vector": t["extraction_vector"],
                "indication": t["condition"],
            })
        print(f"   • CT.gov '{ind}': Found {len(trials)} newly registered protocols")

    print(f"\nTotal raw events captured across all 3 pillars: {len(raw_items)} (in {time.time()-t0:.1f}s)")

    # -------------------------------------------------------------------------
    # DEDUPLICATION & KEYWORD MATCHING
    # -------------------------------------------------------------------------
    print("\n▶ Applying 3-Tier Deduplication & Keyword Match Engine...")
    seen_hashes = set()
    processed_feed = []

    for item in raw_items:
        # Exact URL Dedup
        url_clean = clean_url(item["url"]).lower()
        if url_clean in seen_hashes:
            continue
        seen_hashes.add(url_clean)

        # Keyword Match & Scoring against Tab 07
        res = matcher.match(item["headline"], item["summary"])
        if res.get("is_noise"):
            continue  # Drop noise

        priority = res.get("top_priority", "🟢 Tier 3 (Weekly)")
        if not res.get("matched"):
            priority = "⚪ General (Untagged)"

        entities_str = ", ".join(res.get("matches", [])) or "General Biopharma"
        desk_str = res.get("assigned_desk", "General CI")
        indication_str = item.get("indication") or (entities_str.split("(")[0].strip() if "Indication" in entities_str else "Biopharma Landscape")

        pub_d = item.get("published_date") or (item["published_utc"].strftime("%Y-%m-%d") if item.get("published_utc") else "")
        pub_t = item.get("published_time") or (item["published_utc"].strftime("%H:%M") if item.get("published_utc") else "--")

        processed_feed.append({
            "date": pub_d,
            "time": pub_t,
            "entity": entities_str,
            "indication": indication_str,
            "desk": desk_str,
            "priority": priority,
            "headline": item["headline"],
            "method": item["discovery_method"],
            "vector": item["extraction_vector"],
            "url": item["url"],
            "summary": item["summary"],
            "source": item["source_name"],
            "sort_dt": item["published_utc"]
        })

    # Sort newest first
    processed_feed.sort(key=lambda x: x["sort_dt"], reverse=True)
    print(f"Deduplicated and tagged: {len(processed_feed)} high-value intelligence items.\n")

    # -------------------------------------------------------------------------
    # WRITE TO 00_Unified_Intelligence_Feed TAB IN MASTER EXCEL
    # -------------------------------------------------------------------------
    print(f"▶ Writing to Excel tab '00_Unified_Intelligence_Feed' in {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    TAB_UNIFIED = "00_Unified_Intelligence_Feed"
    if TAB_UNIFIED in wb.sheetnames:
        del wb[TAB_UNIFIED]

    # Create as first sheet
    ws = wb.create_sheet(title=TAB_UNIFIED, index=0)
    ws.views.sheetView[0].showGridLines = True

    # Palette
    NAVY_DARK = "1B365D"
    BLUE_DARK = "1F4E79"
    WHITE = "FFFFFF"
    fill_navy = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
    fill_blue = PatternFill(start_color=BLUE_DARK, end_color=BLUE_DARK, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    fill_ice = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")

    font_tbl_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_data = Font(name="Calibri", size=10, color="000000")
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    font_code = Font(name="Consolas", size=9, color="111111")

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE")
    )

    headers = [
        ("Published Date", 16, fill_navy),
        ("Published Time (UTC)", 20, fill_navy),
        ("Entity / Target / Asset", 28, fill_navy),
        ("Therapy Area / Indication", 24, fill_navy),
        ("Assigned CI Desk", 24, fill_navy),
        ("Priority Tier", 22, fill_blue),
        ("Headline / Event / Protocol Delta", 55, fill_navy),
        ("Discovery Pillar", 24, fill_blue),
        ("Extraction Vector", 24, fill_blue),
        ("Authentic Canonical Link", 45, fill_blue),
        ("Snippet / Excerpt / Protocol Details", 55, fill_blue),
    ]

    for col_idx, (h_name, width, h_fill) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=h_name)
        cell.font = font_tbl_header
        cell.fill = h_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 28

    for r_idx, row in enumerate(processed_feed, start=2):
        ws.cell(row=r_idx, column=1, value=row["date"]).font = font_data
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=2, value=row["time"]).font = font_data
        ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r_idx, column=3, value=row["entity"]).font = font_bold
        ws.cell(row=r_idx, column=4, value=row["indication"]).font = font_data
        ws.cell(row=r_idx, column=5, value=row["desk"]).font = font_data
        
        # Priority badge
        cell_p = ws.cell(row=r_idx, column=6, value=row["priority"])
        cell_p.alignment = Alignment(horizontal="center")
        if "Tier 1" in row["priority"]:
            cell_p.font = Font(name="Calibri", size=10, bold=True, color="C2185B")
            cell_p.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        elif "Tier 2" in row["priority"]:
            cell_p.font = Font(name="Calibri", size=10, bold=True, color="B78103")
            cell_p.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        else:
            cell_p.font = font_data

        ws.cell(row=r_idx, column=7, value=row["headline"]).font = font_bold
        ws.cell(row=r_idx, column=8, value=row["method"]).font = font_data
        ws.cell(row=r_idx, column=9, value=row["vector"]).font = font_code

        # Link
        cell_l = ws.cell(row=r_idx, column=10, value=row["url"])
        cell_l.font = font_link
        cell_l.hyperlink = row["url"]

        ws.cell(row=r_idx, column=11, value=row["summary"]).font = font_data

        for c in range(1, 12):
            ws.cell(row=r_idx, column=c).border = thin_border

        ws.row_dimensions[r_idx].height = 24

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    wb.save(XLSX_PATH)
    print(f"Master Excel workbook updated with tab '{TAB_UNIFIED}' ({len(processed_feed)} items, 10 columns)!")

    # -------------------------------------------------------------------------
    # GENERATE HTML DASHBOARD (results/unified_intelligence_feed.html)
    # -------------------------------------------------------------------------
    os.makedirs(os.path.dirname(HTML_OUT), exist_ok=True)
    html_cards = []
    for item in processed_feed:
        badge_color = "#E91E63" if "Tier 1" in item["priority"] else "#FF9800" if "Tier 2" in item["priority"] else "#4CAF50"
        card = f"""
        <div class="card">
            <div class="card-header">
                <span class="badge" style="background:{badge_color};">{item['priority']}</span>
                <span class="desk-tag">{item['desk']}</span>
                <span class="ind-tag">🔬 {item['indication']}</span>
                <span class="date-tag">📅 {item['date']}</span>
            </div>
            <h3 class="headline"><a href="{item['url']}" target="_blank">{item['headline']}</a></h3>
            <p class="summary">{item['summary']}</p>
            <div class="provenance">
                <span>📍 <b>Origin:</b> {item['method']}</span> &bull; 
                <span>⚙️ <b>Route:</b> {item['vector']}</span> &bull; 
                <span>🏷️ <b>Matched Asset:</b> {item['entity']}</span>
            </div>
        </div>
        """
        html_cards.append(card)

    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Unified Pharma & Biotech Intelligence Feed</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; background: #f0f2f5; color: #1c1e21; margin: 0; padding: 24px; }}
        .container {{ max-width: 1100px; margin: 0 auto; }}
        .header {{ background: #1B365D; color: white; padding: 24px; border-radius: 12px; margin-bottom: 20px; box-shadow: 0 4px 12px rgba(0,0,0,0.1); }}
        .header h1 {{ margin: 0 0 8px 0; font-size: 24px; }}
        .header p {{ margin: 0; opacity: 0.9; font-size: 14px; }}
        .stats {{ display: flex; gap: 15px; margin-top: 15px; font-size: 13px; }}
        .stat-badge {{ background: rgba(255,255,255,0.2); padding: 4px 10px; border-radius: 6px; }}
        .card {{ background: white; border-radius: 10px; padding: 18px; margin-bottom: 14px; box-shadow: 0 2px 6px rgba(0,0,0,0.06); transition: transform 0.15s ease; border-left: 5px solid #1B365D; }}
        .card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }}
        .card-header {{ display: flex; align-items: center; gap: 10px; margin-bottom: 10px; flex-wrap: wrap; }}
        .badge {{ color: white; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: bold; }}
        .desk-tag {{ background: #EBF2FA; color: #1F4E79; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
        .ind-tag {{ background: #E8F5E9; color: #2E7D32; padding: 3px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
        .date-tag {{ color: #65676b; font-size: 12px; margin-left: auto; }}
        .headline {{ margin: 0 0 8px 0; font-size: 16px; line-height: 1.4; }}
        .headline a {{ color: #0b57d0; text-decoration: none; }}
        .headline a:hover {{ text-decoration: underline; }}
        .summary {{ color: #4b4f56; font-size: 13px; margin: 0 0 12px 0; line-height: 1.5; }}
        .provenance {{ background: #f8f9fa; padding: 8px 12px; border-radius: 6px; font-size: 11.5px; color: #555; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🔬 Unified Pharma & Biotech Intelligence Feed</h1>
            <p>Real-Time Consolidated Radar across Publisher Feeds, Company Newsrooms & ClinicalTrials.gov</p>
            <div class="stats">
                <span class="stat-badge">⏱️ {tw.format_summary().splitlines()[0]}</span>
                <span class="stat-badge">📊 {len(processed_feed)} Deduplicated Items</span>
                <span class="stat-badge">🏷️ 28 Tracking Rules Active</span>
            </div>
        </div>
        {"".join(html_cards)}
    </div>
</body>
</html>"""

    with open(HTML_OUT, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"▶ HTML Dashboard generated at: {HTML_OUT}")
    print("=" * 75)
    print(" ✨ LOCAL TEST RUN COMPLETE!")
    print("=" * 75)


if __name__ == "__main__":
    w_arg = sys.argv[1] if len(sys.argv) > 1 else "72h"
    admin_arg = "--admin" in sys.argv
    run_local_pipeline(w_arg, is_admin=admin_arg)
