#!/usr/bin/env python3
"""
Sample Test Suite for RSSFeedChecker.
Runs a benchmark across all intelligence methods:
  1. Method 1: Feed Check (20 feeds)
  2. Method 2: Company Newsroom Watch (22 sample companies, 3-day time window)
  3. Method 3: Indication Search Feeds (18 indications)
  4. Method 4: ClinicalTrials.gov Material Protocol Deltas & New Registrations (18 indications)
Outputs consolidated 'Sample_Test_Results.xlsx' workbook with full results, sanitized dates, & full text content.
"""

import os
import sys
import time
import subprocess

# Ensure UTF-8 stdout on Windows
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = r"c:\Users\User\Desktop\App\All_Apps\RSSFeedChecker"
RESULTS_DIR = os.path.join(WORKSPACE, "results")
OUTPUT_EXCEL = os.path.join(WORKSPACE, "Sample_Test_Results.xlsx")

def run_method_1():
    print("\n" + "="*70)
    print("[METHOD 1] Testing 20 Publisher Feeds (Trade Press, FDA, Journals)...")
    print("="*70)
    cmd = [sys.executable, os.path.join(WORKSPACE, ".github", "scripts", "check_feeds.py"), os.path.join(WORKSPACE, "feeds_sample.tsv")]
    env = os.environ.copy()
    env["FEED_OUT_DIR"] = RESULTS_DIR
    env["FEED_WORKERS"] = "6"
    env["FEED_TIMEOUT"] = "20"
    subprocess.run(cmd, env=env, check=True)

def run_method_2(days=3):
    print("\n" + "="*70)
    print(f"[METHOD 2] Watching 22 Company Newsrooms — Last {days} Days Window...")
    print("="*70)
    cmd = [sys.executable, os.path.join(WORKSPACE, ".github", "scripts", "watch_newsrooms.py"), os.path.join(WORKSPACE, "companies_sample.tsv"), "--days", str(days)]
    env = os.environ.copy()
    env["NEWS_OUT_DIR"] = RESULTS_DIR
    env["NEWS_WORKERS"] = "6"
    env["NEWS_TIMEOUT"] = "20"
    env["NEWS_DAYS_BACK"] = str(days)
    env["NEWS_STATE"] = os.path.join(WORKSPACE, "state", "seen.json")
    subprocess.run(cmd, env=env, check=True)

def run_method_3():
    print("\n" + "="*70)
    print("[METHOD 3] Generating & Verifying Indication Feeds (Obesity, Oncology, MASH)...")
    print("="*70)
    cmd = [sys.executable, os.path.join(WORKSPACE, ".github", "scripts", "make_indication_feeds.py"), os.path.join(WORKSPACE, "indications.tsv")]
    env = os.environ.copy()
    env["IND_OUT_DIR"] = RESULTS_DIR
    env["IND_WORKERS"] = "4"
    env["IND_TIMEOUT"] = "20"
    subprocess.run(cmd, env=env, check=True)

def run_method_4(days=3):
    print("\n" + "="*70)
    print(f"[METHOD 4] ClinicalTrials.gov Protocol Watcher — Last {days} Days Window...")
    print("="*70)
    cmd = [sys.executable, os.path.join(WORKSPACE, ".github", "scripts", "check_clinical_trials.py"), os.path.join(WORKSPACE, "indications.tsv"), "--days", str(days)]
    env = os.environ.copy()
    env["CT_DAYS_BACK"] = str(days)
    subprocess.run(cmd, env=env, check=True)

def build_combined_excel(days=3):
    print("\n" + "="*70)
    print("--> Compiling All Results into 'Sample_Test_Results.xlsx'...")
    print("="*70)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    NAVY = "1B365D"
    SLATE = "2D5584"
    ICE = "EBF2FA"
    WHITE = "FFFFFF"
    BORDER_CLR = "D0D7DE"
    
    font_title = Font(name="Calibri", size=15, bold=True, color=NAVY)
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_norm = Font(name="Calibri", size=10, color="000000")
    font_code = Font(name="Consolas", size=9, color="111111")
    font_link = Font(name="Calibri", size=10, color="0066CC", underline="single")
    
    fill_hdr = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    fill_ice = PatternFill(start_color=ICE, end_color=ICE, fill_type="solid")
    fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    
    fill_green = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_green = Font(name="Calibri", size=10, bold=True, color="155724")
    
    fill_yellow = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_yellow = Font(name="Calibri", size=10, bold=True, color="856404")
    
    fill_red = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    font_red = Font(name="Calibri", size=10, bold=True, color="721C24")
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_CLR),
        right=Side(style='thin', color=BORDER_CLR),
        top=Side(style='thin', color=BORDER_CLR),
        bottom=Side(style='thin', color=BORDER_CLR)
    )

    # 1. Summary Dashboard Tab
    ws_sum = wb.create_sheet(title="01_Test_Summary_Dashboard")
    ws_sum.views.sheetView[0].showGridLines = True
    ws_sum.column_dimensions["A"].width = 5
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 24
    ws_sum.column_dimensions["D"].width = 30
    ws_sum.column_dimensions["E"].width = 65

    ws_sum.merge_cells("B2:E2")
    ws_sum["B2"] = "RSSFeedChecker — Multi-Pillar Benchmark Results"
    ws_sum["B2"].font = font_title
    
    ws_sum.merge_cells("B3:E3")
    ws_sum["B3"] = f"Strict time-windowing ({days} days) + Date Sanitization (YYYY-MM-DD) + Full-Text Body Content"
    ws_sum["B3"].font = Font(name="Calibri", size=11, italic=True, color="555555")

    ws_sum.merge_cells("B5:E5")
    ws_sum["B5"] = "METHOD PERFORMANCE & CAPABILITY SUMMARY"
    ws_sum["B5"].font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    ws_sum["B5"].fill = PatternFill(start_color=SLATE, end_color=SLATE, fill_type="solid")
    ws_sum["B5"].alignment = Alignment(vertical="center", indent=1)
    ws_sum.row_dimensions[5].height = 24

    for c_idx, h in enumerate(["Intelligence Method", "Sample Tested", "Key Capability Tested", "Operational Takeaway / How to Use"], start=2):
        col_letter = get_column_letter(c_idx)
        ws_sum[f"{col_letter}6"] = h
        ws_sum[f"{col_letter}6"].font = font_tbl_hdr
        ws_sum[f"{col_letter}6"].fill = fill_hdr
        ws_sum[f"{col_letter}6"].border = thin_border
    ws_sum.row_dimensions[6].height = 22

    dash_rows = [
        ("Pillar 1: Feed Health Check", "20 Publisher Feeds (feeds_sample.tsv)", "SSL verification, chunked XML download, entity fixing, sanitized dates, and live article title verification.", "Identifies dead or misconfigured media/journal feeds in seconds so news aggregation never silently fails."),
        ("Pillar 2: Company Newsroom Watch", f"22 Companies ({days}-day window)", "Waterfall: RSS -> Sitemaps -> HTML scraping -> GNews. Deep-scrapes real article publish dates from HTML & extracts full text paragraphs. Filters out jobs/careers.", "Accurately isolates true new press releases published in the window. If a company was quiet, shows exact date of their latest historical release."),
        ("Pillar 3: Indication Radar", "18 Therapy Indications (indications.tsv)", "Dynamic construction & live validation of Google News (Broad & Regulatory) search feeds.", "Gives instant search coverage for newly researched therapeutic areas without waiting for publishers."),
        ("Pillar 4: CT.gov Protocol Diff Engine", f"18 Indications ({days}-day delta)", "REST API v2 protocol diffing: Identifies [NEW TRIAL REGISTERED] and [MATERIAL PROTOCOL CHANGE] (Enrollment 50->90, Status changes). Filters out routine administrative re-verifications.", "Pinpoints exactly what changed in clinical trials with direct side-by-side history comparison links.")
    ]

    for r_i, (m_name, m_sample, m_cap, m_use) in enumerate(dash_rows, start=7):
        ws_sum[f"B{r_i}"] = m_name
        ws_sum[f"B{r_i}"].font = font_bold
        ws_sum[f"B{r_i}"].border = thin_border
        ws_sum[f"B{r_i}"].fill = fill_ice
        
        ws_sum[f"C{r_i}"] = m_sample
        ws_sum[f"C{r_i}"].font = font_norm
        ws_sum[f"C{r_i}"].border = thin_border
        
        ws_sum[f"D{r_i}"] = m_cap
        ws_sum[f"D{r_i}"].font = font_norm
        ws_sum[f"D{r_i}"].border = thin_border
        ws_sum[f"D{r_i}"].alignment = Alignment(wrap_text=True, vertical="center")
        
        ws_sum[f"E{r_i}"] = m_use
        ws_sum[f"E{r_i}"].font = font_norm
        ws_sum[f"E{r_i}"].border = thin_border
        ws_sum[f"E{r_i}"].alignment = Alignment(wrap_text=True, vertical="center")
        ws_sum.row_dimensions[r_i].height = 42

    def load_tsv_to_sheet(tsv_file, title, badge_col=None, link_cols=[]):
        if not os.path.exists(tsv_file):
            return
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        import csv as csv_mod
        with open(tsv_file, encoding="utf-8") as f:
            reader = csv_mod.reader(f, delimiter="\t")
            lines = [r for r in reader if r and any(c.strip() for c in r)]
        if not lines:
            return
            
        for c_idx, h in enumerate(lines[0], start=1):
            col_letter = get_column_letter(c_idx)
            ws[f"{col_letter}1"] = h
            ws[f"{col_letter}1"].font = font_tbl_hdr
            ws[f"{col_letter}1"].fill = fill_hdr
            ws[f"{col_letter}1"].border = thin_border
        ws.row_dimensions[1].height = 24
        
        for r_idx, row in enumerate(lines[1:], start=2):
            for c_idx, val in enumerate(row, start=1):
                col_letter = get_column_letter(c_idx)
                cell = ws[f"{col_letter}{r_idx}"]
                cell.value = val
                cell.font = font_norm
                cell.border = thin_border
                
                if c_idx in link_cols:
                    cell.font = font_link
                    if str(val).startswith("http"):
                        cell.hyperlink = val
                elif badge_col and c_idx == badge_col:
                    val_str = str(val)
                    if "OK" in val_str or "NEW" in val_str or "MATERIAL" in val_str or "ACTIVE" in val_str:
                        cell.fill = fill_green
                        cell.font = font_green
                    elif "BLOCKED" in val_str or "FAILED" in val_str or "NO_NEW" in val_str or "NO_CHANGES" in val_str:
                        cell.fill = fill_red
                        cell.font = font_red
                    else:
                        cell.fill = fill_yellow
                        cell.font = font_yellow
            ws.row_dimensions[r_idx].height = 20
            
        for c_idx in range(1, len(lines[0])+1):
            col_letter = get_column_letter(c_idx)
            max_len = max(len(str(r[c_idx-1])) if c_idx-1 < len(r) else 10 for r in lines[:50])
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 75)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(lines[0]))}{len(lines)}"

    # Method 1
    load_tsv_to_sheet(os.path.join(RESULTS_DIR, "results.tsv"), "02_Method1_Feed_Results", badge_col=6, link_cols=[2, 8])
    # Method 2
    load_tsv_to_sheet(os.path.join(RESULTS_DIR, "newsroom.tsv"), "03_Method2_Company_Releases", badge_col=10, link_cols=[9])
    # Method 3
    load_tsv_to_sheet(os.path.join(RESULTS_DIR, "indication_sources.tsv"), "04_Method3_Indication_Feeds", badge_col=8, link_cols=[1])
    # Method 4
    load_tsv_to_sheet(os.path.join(RESULTS_DIR, "clinical_trials_deltas.tsv"), "05_Clinical_Trials_Deltas", badge_col=2, link_cols=[13, 14])

    try:
        wb.save(OUTPUT_EXCEL)
        print(f"\n[OK] Successfully compiled '{OUTPUT_EXCEL}' with all results!")
    except PermissionError:
        alt_path = os.path.join(WORKSPACE, "Sample_Test_Results_Updated.xlsx")
        wb.save(alt_path)
        print(f"\n[WARN] 'Sample_Test_Results.xlsx' is open in Excel. Saved updated workbook to: '{alt_path}'")

    # Generate interactive HTML dashboard
    try:
        from generate_trial_diff_dashboard import generate_dashboard
        generate_dashboard()
    except Exception as e:
        print(f"[WARN] Could not generate visual dashboard: {e}")

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Sample Benchmark Suite")
    parser.add_argument("--days", type=int, default=3, help="Time window in days (default 3)")
    args = parser.parse_args()

    os.makedirs(RESULTS_DIR, exist_ok=True)
    print("\n" + "="*70)
    print(f">>> RSSFeedChecker — Automated Multi-Method Benchmark (Window: {args.days} Days)")
    print("="*70)
    
    t0 = time.time()
    run_method_1()
    run_method_2(days=args.days)
    run_method_3()
    run_method_4(days=args.days)
    build_combined_excel(days=args.days)
    
    total_time = round(time.time() - t0, 1)
    print(f"\n[DONE] All benchmark tests completed in {total_time}s!")
    print(f"[FILE] Output workbook: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()
